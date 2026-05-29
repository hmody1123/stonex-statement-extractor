"""MyStoneX Positions statement parser (v76).

Extracts trade-related sections from StoneX Daily, Monthly, and IFL statement PDFs into
pandas DataFrames. This is a local/offline parser using PyMuPDF text extraction and regex
rules tuned to the sample statement layouts. It keeps source lines and page numbers for
audit/reconciliation.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment



def _daily_open_positions_backstop(pdf_bytes: bytes, include_open_positions: bool = True) -> list[dict]:
    """Parse all classic daily-statement OPEN POSITION rows across all pages.

    This is a safety net for continuation pages that repeat only the column header
    ("CONTRACT DESCRIPTION-OPEN") and not the full spaced OPEN POSITIONS title.
    """
    if not include_open_positions:
        return []
    rows: list[dict] = []
    for page_no, text in pdf_text(pdf_bytes):
        if "CONTRACT DESCRIPTION-OPEN" not in text:
            continue
        stmt_date = _statement_date(text)
        account = _account_number(text)
        broker = _broker_code(text)
        for raw in text.splitlines():
            line = " ".join(raw.strip().split())
            if not line or line.startswith("-------") or line.startswith("TRADE CARD"):
                continue
            parsed = _parse_daily_open_position_line(line, raw)
            if parsed:
                row = _base_row(parsed, stmt_date, account, page_no, line, broker)
                row["market_value_signed"] = _signed(row.get("market_value"), row.get("drcr"))
                rows.append(row)
    return rows

MONTHS = {m: i for i, m in enumerate(["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], 1)}
DATE_RE = re.compile(r"(?P<m>JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(?P<d>\d{1,2}),\s+(?P<y>\d{4})")
STATEMENT_DATE_RE = re.compile(r"STATEMENT DATE:\s+" + DATE_RE.pattern, re.I)
STATEMENT_DATE_DDMMM_RE = re.compile(r"Statement Date:\s+(?P<d>\d{1,2})-(?P<m>[A-Za-z]{3})-(?P<y>\d{4})", re.I)
ACCOUNT_RE = re.compile(r"ACCOUNT NUMBER:\s+(?P<account>\S+)", re.I)
SALESMAN_RE = re.compile(r"SALESMAN:\s+(?P<broker>\S+)", re.I)

TRADE_LINE_RE = re.compile(
    r"^(?P<trade_date>\d{1,2}/\d{2}/\d)\s+"
    r"(?P<card>[A-Z0-9]+)\s+"
    r"(?P<account_type>[A-Z]\d)\s+"
    r"(?P<quantity>\d[\d,]*)\s+"
    r"(?P<contract_month>[A-Z]{3})\s+"
    r"(?P<contract_year>\d{2})\s+"
    r"(?P<contract_description>.+?)\s+"
    r"(?P<price>\d+(?:\.\d+)?)\s+"
    r"(?P<currency>[A-Z]{2})(?:\s+(?P<amount>[\d,]+\.\d{2})(?P<drcr>DR|CR))?\s*$"
)

RD_LINE_RE = re.compile(
    r"^(?P<trade_date>\d{1,2}/\d{2}/\d)\s+(?:(?P<card>[A-Z0-9]+)\s+)?(?P<account_type>U\d)\s+"
    r"(?P<quantity>[\d,]+)\s+"
    r"(?P<description>USTB\s+DUE\s+\d{1,2}/\d{1,2}/\d{4}(?:\s+U)?)\s+"
    r"(?P<legend>PURCHASE|REDEMPTION)\s+"
    r"(?P<currency>[A-Z]{2})\s+(?P<amount>[\d,]+\.\d{2})(?P<drcr>DR|CR)\s*$"
)

JOURNAL_RE = re.compile(
    r"^(?P<trade_date>\d{1,2}/\d{2}/\d)\s+(?P<account_type>U\d)\s+(?P<description>.+?)\s+"
    r"(?P<currency>[A-Z]{2})\s+(?P<amount>[\d,]+\.\d{2})(?P<drcr>DR|CR)\s*$"
)

OPEN_POSITION_LINE_RE = re.compile(
    r"^(?P<trade_date>\d{1,2}/\d{2}/\d)\s+(?:(?P<card>[A-Z0-9]+)\s+)?(?P<account_type>U\d)\s+"
    r"(?P<quantity>[\d,]+)\s+"
    r"(?P<contract_month>[A-Z]{3})\s+(?P<contract_year>\d{2})\s+"
    r"(?P<contract_description>.+?)\s+(?P<price>\d+(?:\.\d+)?)\s+"
    r"(?P<currency>[A-Z]{2})\s+(?P<market_value>[\d,]+\.\d{2})(?P<drcr>DR|CR)\s*$"
)

FEE_OR_AVG_KEYWORDS = ["COMMISSION", "CLEARING FEE", "AVG LONG", "AVG SHORT", "GROSS PROFIT OR LOSS"]

# Purchase & Sale close summary rows are the realized/closed-position details.
# Format A (all on one line):
#   A2 55* 55* LTD- 6/18/26 GROSS PROFIT OR LOSS AD 14,125.00CR
#   U2 241* 241* LTD- 6/30/26 GROSS PROFIT OR LOSS US 155,735.00CR
CLOSED_POSITION_GROSS_RE = re.compile(
    r"^(?P<account_type>[A-Z]\d)\s+"
    r"(?:(?P<long>[\d,]+)\*\s+)?"
    r"(?:(?P<short>[\d,]+)\*\s+)?"
    r"(?P<close_status>.+?)\s+GROSS\s+PROFIT\s+OR\s+LOSS\s+"
    r"(?P<currency>[A-Z]{2})\s+(?P<amount>[\d,]+(?:\.\d{2})?)(?P<drcr>DR|CR)\s*$",
    re.I,
)
# Format B (standalone line — account type / quantities are on a preceding LTD- line):
#   GROSS PROFIT OR LOSS US 30,237.50DR
#   U1 GROSS PROFIT OR LOSS US 30,237.50DR  (account type prefix present but no quantities)
STANDALONE_GROSS_RE = re.compile(
    r"^(?:[A-Z]\d\s+)?GROSS\s+PROFIT\s+OR\s+LOSS\s+"
    r"(?P<currency>[A-Z]{2})\s+(?P<amount>[\d,]+(?:\.\d{2})?)(?P<drcr>DR|CR)\s*$",
    re.I,
)
# LTD- (or EX-) summary line that precedes the standalone GROSS PROFIT OR LOSS line:
#   U1 20* 20* LTD- 7/14/26
PS_LTD_SUMMARY_RE = re.compile(
    r"^(?P<account_type>[A-Z]\d)\s+"
    r"(?:(?P<long>[\d,]+)\*\s+)?"
    r"(?:(?P<short>[\d,]+)\*\s+)?"
    r"(?P<close_status>(?:LTD|EX)-\s*\S+.*?)$",
    re.I,
)

FRACTION_RE = re.compile(r"^\d+/\d+$")

def _parse_price_token(tokens: list[str]) -> float | None:
    """Parse StoneX prices like 4.31 1/2, .43 1/4, 5.0275, or 4.75."""
    if not tokens:
        return None
    try:
        base = float(tokens[0])
    except ValueError:
        return None
    if len(tokens) > 1 and FRACTION_RE.match(tokens[1]):
        num, den = tokens[1].split('/')
        try:
            base += float(num) / float(den) / 100.0
        except Exception:
            pass
    return base


def _split_amount_drcr(token: str | None) -> tuple[str | None, str | None]:
    if token is None:
        return None, None
    t = token.strip()
    m = re.match(r"^(?P<amount>(?:[\d,]+)?\.\d{2}|[\d,]+(?:\.\d+)?|0)(?P<drcr>DR|CR)?$", t)
    if not m:
        return None, None
    return m.group('amount'), m.group('drcr')



def _parse_closed_position_gross_line(
    line: str,
    last_contract: dict | None = None,
    pending_ltd_summary: dict | None = None,
) -> dict | None:
    """Parse P&S GROSS PROFIT OR LOSS lines into closed-position detail rows.

    Handles two statement formats:
    - Format A (all on one line): U1 20* 20* LTD- 7/14/26 GROSS PROFIT OR LOSS US 30,237.50DR
    - Format B (split lines): standalone GROSS PROFIT OR LOSS US 30,237.50DR where the
      account type / quantities appeared on a preceding LTD- summary line captured in
      pending_ltd_summary.
    """
    m = CLOSED_POSITION_GROSS_RE.match(line)
    if not m:
        # Format B: standalone GROSS PROFIT OR LOSS line
        m2 = STANDALONE_GROSS_RE.match(line)
        if not m2:
            return None
        row: dict = m2.groupdict()
        row["account_type"] = None
        row["close_status"] = None
        # Absorb quantities and close-date from the preceding LTD- line if available
        if pending_ltd_summary:
            row["account_type"] = pending_ltd_summary.get("account_type")
            row["long"]         = _num_any(pending_ltd_summary.get("long"))
            row["short"]        = _num_any(pending_ltd_summary.get("short"))
            row["close_status"] = pending_ltd_summary.get("close_status")
        else:
            row["long"] = None
            row["short"] = None
    else:
        row = m.groupdict()
        row["long"]  = _num_any(row.get("long"))
        row["short"] = _num_any(row.get("short"))

    row["quantity"]      = (row.get("long") or 0) - (row.get("short") or 0)
    row["realized_pnl"]  = _signed(row.get("amount"), row.get("drcr"))
    row["amount_signed"] = row["realized_pnl"]
    row["source_section"] = "Closed Positions"
    row["pnl_view"]       = "Closed Position Detail"

    status = str(row.get("close_status") or "")
    mdate  = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", status)
    row["close_date"]  = mdate.group(1) if mdate else None
    mclose = re.search(r"\bCLOSE\s+(?P<close_price>[-+]?\d+(?:\.\d+)?)", status)
    row["close_price"] = _num_any(mclose.group("close_price")) if mclose else None

    if last_contract:
        for k in [
            "contract_description", "contract_month", "contract_year", "ref_month",
            "trade_id", "card", "price", "trade_price", "product", "exchange",
            "option_type", "strike",
        ]:
            if k in last_contract and row.get(k) in (None, ""):
                row[k] = last_contract.get(k)
    return row


def _parse_daily_trade_or_ps_line(line: str) -> dict | None:
    """Parse classic daily confirmation/P&S rows that may not have debit/credit amount."""
    toks = line.split()
    if len(toks) < 7 or not re.match(r"^\d{1,2}/\d{2}/\d$", toks[0]):
        return None
    acct_idx = None
    for i in range(1, min(len(toks), 5)):
        if re.match(r"^[A-Z]\d$", toks[i]):
            acct_idx = i
            break
    if acct_idx is None or acct_idx + 2 >= len(toks):
        return None
    qty_tok = toks[acct_idx + 1]
    if not re.match(r"^[\d,]+$", qty_tok):
        return None

    currency_idx = None
    for i in range(len(toks)-1, acct_idx+2, -1):
        if re.match(r"^[A-Z]{2}$", toks[i]):
            currency_idx = i
            break
    if currency_idx is None:
        return None

    amount = drcr = None
    if currency_idx + 1 < len(toks):
        amount, drcr = _split_amount_drcr(toks[currency_idx + 1])

    if currency_idx >= 2 and FRACTION_RE.match(toks[currency_idx - 1]):
        price_tokens = toks[currency_idx-2:currency_idx]
        price_start = currency_idx - 2
    else:
        price_tokens = [toks[currency_idx - 1]]
        price_start = currency_idx - 1
    price = _parse_price_token(price_tokens)
    if price is None:
        return None

    # Optional status token immediately before price, e.g. "r".
    pre_price = toks[price_start - 1] if price_start - 1 > acct_idx else None
    if pre_price and re.match(r"^[A-Za-z]$", pre_price) and pre_price.upper() not in MONTHS:
        st = pre_price
        desc_tokens = toks[acct_idx + 2:price_start-1]
    else:
        st = None
        desc_tokens = toks[acct_idx + 2:price_start]
    if not desc_tokens:
        return None

    contract_month = contract_year = option_type = None
    # Options may use C/P or CALL/PUT before the month.
    if desc_tokens[0] in {"CALL", "PUT", "C", "P"} and len(desc_tokens) >= 4:
        option_type = "CALL" if desc_tokens[0] in {"CALL", "C"} else "PUT"
        contract_month = desc_tokens[1]
        contract_year = desc_tokens[2]
        contract_description = " ".join(desc_tokens)
    elif len(desc_tokens) >= 2 and re.match(r"^[A-Z]{3}$", desc_tokens[0]) and re.match(r"^\d{2}$", desc_tokens[1]):
        contract_month = desc_tokens[0]
        contract_year = desc_tokens[1]
        contract_description = " ".join(desc_tokens)
    else:
        contract_description = " ".join(desc_tokens)

    return {
        "trade_date": toks[0],
        "card": " ".join(toks[1:acct_idx]) or None,
        "account_type": toks[acct_idx],
        "quantity": _to_num(qty_tok),
        "contract_month": contract_month,
        "contract_year": contract_year,
        "contract_description": contract_description,
        "option_type_raw": option_type,
        "st": st,
        "price": price,
        "price_text": " ".join(price_tokens),
        "currency": toks[currency_idx],
        "amount": amount,
        "drcr": drcr,
    }

def _parse_daily_open_position_line(line: str, raw_line: str | None = None) -> dict | None:
    toks = line.split()
    if len(toks) < 9 or not re.match(r"^\d{1,2}/\d{2}/\d$", toks[0]):
        return None
    acct_idx = None
    for i in range(1, min(len(toks), 5)):
        if re.match(r"^[A-Z]\d$", toks[i]):
            acct_idx = i
            break
    if acct_idx is None or acct_idx + 2 >= len(toks):
        return None
    qty_tok = toks[acct_idx + 1]
    if not re.match(r"^[\d,]+$", qty_tok):
        return None

    # Determine whether the quantity is in the LONG or SHORT printed column.
    # The normalized text loses fixed-width spacing, so use raw_line when available.
    # Classic StoneX daily header columns are approximately:
    # LONG starts near col 25, SHORT near col 39, CONTRACT near col 50.
    side = None
    long_qty = None
    short_qty = None
    if raw_line:
        qty_match = re.search(r"^\s*\d{1,2}/\d{2}/\d\s+(?:[A-Z0-9]+\s+)?[A-Z]\d\s+(?P<qty>\d[\d,]*)\s+", raw_line)
        if qty_match:
            qty_start = qty_match.start("qty")
            # If the quantity is right-aligned in the SHORT column, its start is usually
            # well past the midpoint between LONG and SHORT labels.
            if qty_start >= 34:
                side = "Short"
                short_qty = _to_num(qty_tok)
            else:
                side = "Long"
                long_qty = _to_num(qty_tok)
    if side is None:
        # Fallback keeps backward compatibility for normalized-only lines.
        side = None
        long_qty = None
        short_qty = None
    currency_idx = None
    for i in range(len(toks)-2, acct_idx+2, -1):
        if re.match(r"^[A-Z]{2}$", toks[i]):
            currency_idx = i
            break
    if currency_idx is None or currency_idx + 1 >= len(toks):
        return None
    amount, drcr = _split_amount_drcr(toks[currency_idx + 1])
    if amount is None:
        return None
    if currency_idx >= 2 and FRACTION_RE.match(toks[currency_idx - 1]):
        price_tokens = toks[currency_idx-2:currency_idx]
        price_start = currency_idx - 2
    else:
        price_tokens = [toks[currency_idx - 1]]
        price_start = currency_idx - 1
    price = _parse_price_token(price_tokens)
    if price is None:
        return None
    # Some statements have a status token immediately before price (e.g., SE), but
    # others put the product symbol there (e.g., SCM TSR20RUBBR 210.0 US 1,020.00DR).
    # Treat the pre-price token as status only when it looks like a short code.
    pre_price = toks[price_start - 1] if price_start - 1 > acct_idx else None
    if pre_price and re.match(r"^[A-Za-z]$", pre_price):
        st = pre_price
        desc_tokens = toks[acct_idx + 2:price_start-1]
    else:
        st = None
        desc_tokens = toks[acct_idx + 2:price_start]
    if not desc_tokens:
        return None
    contract_month = contract_year = option_type = None
    if desc_tokens[0] in {"CALL", "PUT"} and len(desc_tokens) >= 4:
        option_type = desc_tokens[0]
        contract_month = desc_tokens[1]
        contract_year = desc_tokens[2]
        contract_description = " ".join(desc_tokens)
    elif len(desc_tokens) >= 3 and re.match(r"^[A-Z]{3}$", desc_tokens[0]) and re.match(r"^\d{2}$", desc_tokens[1]):
        contract_month = desc_tokens[0]
        contract_year = desc_tokens[1]
        contract_description = " ".join(desc_tokens[2:])
    else:
        contract_description = " ".join(desc_tokens)
    return {
        "trade_date": toks[0], "card": " ".join(toks[1:acct_idx]) or None,
        "account_type": toks[acct_idx], "quantity": _to_num(qty_tok),
        "long": long_qty, "short": short_qty, "side": side,
        "contract_month": contract_month, "contract_year": contract_year,
        "contract_description": contract_description, "option_type_raw": option_type,
        "st": st, "price": price, "price_text": " ".join(price_tokens),
        "currency": toks[currency_idx], "market_value": amount, "drcr": drcr,
    }


def _to_num(value: str | float | int | None) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).replace(",", ""))


def _signed(amount: str | None, drcr: str | None) -> float | None:
    value = _to_num(amount)
    if value is None:
        return None
    return -value if drcr == "DR" else value


def _statement_date(page_text: str) -> str | None:
    match = STATEMENT_DATE_RE.search(page_text)
    if match:
        return f"{int(match.group('y')):04d}-{MONTHS[match.group('m').upper()]:02d}-{int(match.group('d')):02d}"
    match = STATEMENT_DATE_DDMMM_RE.search(page_text)
    if match:
        return f"{int(match.group('y')):04d}-{MONTHS[match.group('m').upper()]:02d}-{int(match.group('d')):02d}"
    return None


def _account_number(page_text: str) -> str | None:
    match = ACCOUNT_RE.search(page_text)
    return match.group("account") if match else None


def _broker_code(page_text: str) -> str | None:
    match = SALESMAN_RE.search(page_text)
    return match.group("broker") if match else None


def _normalize_trade_date(short_date: str, statement_date: str | None) -> str:
    mm, dd, y1 = short_date.split("/")
    year = int(statement_date[:3] + y1) if statement_date else 2020 + int(y1)
    return f"{year:04d}-{int(mm):02d}-{int(dd):02d}"




def _normalize_any_date(date_text: str | None, statement_date: str | None = None) -> str | None:
    """Normalize M/DD/Y, DD-Mon-YYYY, DD-Mon-YY, or ISO date text to YYYY-MM-DD."""
    if not date_text:
        return None
    s = str(date_text).strip()
    if not s or s.lower() in {"nan", "none", "nat"}:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    if "/" in s:
        try:
            return _normalize_trade_date(s, statement_date)
        except Exception:
            return None
    m = re.match(r"^(?P<d>\d{1,2})-(?P<m>[A-Za-z]{3})-(?P<y>\d{4})$", s)
    if m:
        return f"{int(m.group('y')):04d}-{MONTHS[m.group('m').upper()]:02d}-{int(m.group('d')):02d}"
    m = re.match(r"^(?P<d>\d{1,2})-(?P<m>[A-Za-z]{3})-(?P<y>\d{2})$", s)
    if m:
        yy = int(m.group('y'))
        year = 2000 + yy if yy < 70 else 1900 + yy
        return f"{year:04d}-{MONTHS[m.group('m').upper()]:02d}-{int(m.group('d')):02d}"
    return None


def _ref_month_from_date(date_text: str | None) -> str | None:
    iso = _normalize_any_date(date_text)
    if not iso:
        return None
    y, m, _d = iso.split('-')
    mon = list(MONTHS.keys())[int(m) - 1]
    return f"{mon}-{y[-2:]}"


def _split_date_and_ref_month(value: str | None) -> tuple[str | None, str | None]:
    """Extract a date and Ref Month from combined table text.

    Some PDFs extract the End Date and Ref Month columns as one string, e.g.
    ``18-May-2026 MAY-26``.  This helper returns the actual date text and the
    Ref Month separately so Open Positions can show expiryDate and Contract
    Month/Year correctly.

    Important: do not treat the month embedded in the date itself
    (for example ``12-Jun-2026``) as the Ref Month.  The Ref Month is the
    separate month-year token after the end/expiry date, such as ``JUL-26``.
    """
    s = " ".join(str(value or "").strip().split())
    if not s or s.lower() in {"none", "nan", "nat"}:
        return None, None
    date_match = re.search(r"\b\d{1,2}-[A-Za-z]{3}-\d{2,4}\b", s)
    date_text = date_match.group(0) if date_match else None

    ref_search_text = s
    if date_match:
        # Search only outside the date token so ``12-Jun-2026`` does not produce
        # a false ``JUN-26`` contract month/year.
        ref_search_text = (s[:date_match.start()] + " " + s[date_match.end():]).strip()

    ref_match = re.search(r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{2,4}\b", ref_search_text, re.I)
    ref_month = ref_match.group(0).upper() if ref_match else None
    if ref_month:
        mon, yy = ref_month.split("-", 1)
        ref_month = f"{mon}-{yy[-2:]}"
    return date_text, ref_month


def _normalize_ref_month(value: str | None) -> str | None:
    s = " ".join(str(value or "").strip().split()).upper()
    m = re.search(r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-?(\d{2,4})\b", s)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)[-2:]}"


LME_AMOUNT_RE = r"\(?[\d,]+(?:\.\d+)?\)?"
LME_AVG_OPEN_RE = re.compile(
    rf"^(?P<trade_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{4}})\s+"
    rf"(?P<quantity>[\d,]+)\s+"
    rf"(?P<price>ASP|3MS|(?:ASP|3MS)[+-]\d+(?:\.\d+)?)\s+"
    rf"(?P<start_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{4}})\s+"
    rf"(?P<end_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{4}})\s+"
    rf"(?P<settlement_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{4}})\s+"
    rf"(?P<currency>[A-Z]{{3}})\s+(?P<market_value>{LME_AMOUNT_RE})$"
)
LME_PRODUCT_RE = re.compile(r"^(?P<delivery_date>\d{1,2}-[A-Za-z]{3}-\d{4})\s+(?P<product>LME\s+.+)$")
LME_FUT_OPT_RE = re.compile(
    rf"^(?P<trade_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{4}})\s+"
    rf"(?P<quantity>[\d,]+)\s+"
    rf"(?P<delivery_date>\d{{1,2}}-[A-Za-z]{{3}}-\d{{2}})\s+"
    rf"(?P<product>LME\s+.+?)\s+"
    rf"(?P<trade_price>\d+(?:\.\d+)?)\s+"
    rf"(?P<price_type>TradeWhite|Trade|White)?\s*"
    rf"(?P<currency>[A-Z]{{3}})\s+(?P<market_value>{LME_AMOUNT_RE})$"
)


def _parse_lme_average_row(line: str) -> dict | None:
    m = LME_AVG_OPEN_RE.match(line)
    if not m:
        return None
    row = m.groupdict()
    row["source_section"] = "LME Average Open Positions"
    row["exchange"] = "LME"
    row["contract_description"] = None
    row["contract_month"] = None
    row["contract_year"] = None
    row["ref_month"] = _ref_month_from_date(row.get("settlement_date"))
    row["market_value_signed"] = _num_any(row.get("market_value"))
    row["trade_date_iso"] = _normalize_any_date(row.get("trade_date"))
    return row


def _parse_lme_fut_opt_row(line: str) -> dict | None:
    m = LME_FUT_OPT_RE.match(line)
    if not m:
        return None
    row = m.groupdict()
    row["source_section"] = "Futures / Options Open Positions"
    row["exchange"] = "LME"
    row["contract_description"] = row.pop("product")
    row["contract_date"] = row.get("delivery_date")
    row["ref_month"] = _ref_month_from_date(row.get("delivery_date"))
    row["trade_date_iso"] = _normalize_any_date(row.get("trade_date"))
    row["market_value_signed"] = _num_any(row.get("market_value"))
    return row

def _side_from_section(section: str | None, quantity: float | None) -> str | None:
    # Statement layout does not print BUY/SELL text directly for each row; the row appears under BUY or SELL columns.
    # The parser stores source_section and leaves side blank unless a future layout rule identifies it reliably.
    return None


def pdf_text(pdf_bytes: bytes) -> List[Tuple[int, str]]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    return [(i + 1, page.get_text("text")) for i, page in enumerate(doc)]


def _base_row(groupdict: dict, stmt_date: str | None, account: str | None, page_no: int, raw_line: str, broker: str | None = None) -> dict:
    row = dict(groupdict)
    row["statement_date"] = stmt_date
    row["account_number"] = account
    row["broker_code"] = broker
    row["trade_date_iso"] = _normalize_any_date(row.get("trade_date"), stmt_date)
    row["quantity"] = _to_num(row.get("quantity"))
    if "price" in row:
        row["price"] = _to_num(row.get("price"))
    row["amount_signed"] = _signed(row.get("amount"), row.get("drcr"))
    row["page"] = page_no
    row["source_line"] = raw_line
    return row


# ---------------- Monthly statement parser (landscape Activity format) ----------------
DATE_DDMMMYYYY_RE = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
MONEY_RE = re.compile(r"^\(?\$?[\d,]+(?:\.\d+)?\)?$|^\$?\d+(?:\.\d+)?$|^\([\d,]+(?:\.\d+)?\)$")


def _num_any(value: str | None) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    neg = (s.startswith('(') and s.endswith(')')) or s.endswith('-') or s.startswith('-')
    s = s.replace('$','').replace(',','').replace('(','').replace(')','').strip()
    if s.endswith('-'):
        s = s[:-1].strip()
    if s.startswith('-'):
        s = s[1:].strip()
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def _murex_num(value: str | None) -> float | None:
    """Parse Murex-format numeric values with BRL (R$) and USD ($) currency prefixes.

    BRL format uses comma as decimal separator and period as thousands separator,
    e.g. ``R$ 6.065,98`` -> 6065.98, ``R$ 69,80`` -> 69.80.
    USD format uses the standard comma-as-thousands convention,
    e.g. ``$334.2212`` -> 334.2212, ``($250,000.00)`` -> -250000.0.
    Values with no currency prefix but a single comma and no period (e.g. ``66,79``)
    are treated as BRL-style decimals.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {'none', 'nan', 'nat', ''}:
        return None

    neg = (s.startswith('(') and s.endswith(')')) or s.startswith('-')
    s = s.replace('(', '').replace(')', '').lstrip('-').strip()

    # Detect BRL prefix (R$ or R followed by space)
    is_brl = bool(re.match(r'^R\$?\s', s, re.I))
    # Strip any leading currency symbols/letters up to first digit
    s = re.sub(r'^[A-Z€£¥]*\$?\s*', '', s).strip()

    if not s:
        return None

    if is_brl:
        # BRL: period = thousands separator, comma = decimal point
        s = s.replace('.', '').replace(',', '.')
    elif '.' in s and ',' in s:
        # Both separators present — determine which is decimal by position
        if s.rfind(',') > s.rfind('.'):
            # "1.234,56" style — comma is decimal
            s = s.replace('.', '').replace(',', '.')
        else:
            # "1,234.56" style — period is decimal
            s = s.replace(',', '')
    elif ',' in s and '.' not in s:
        # Comma only — treat as decimal (BRL-style), e.g. "66,79" or "3,007993"
        s = s.replace(',', '.')
    else:
        # Period only or no separator — standard numeric
        s = s.replace(',', '')

    try:
        v = float(s)
        return -v if neg else v
    except (ValueError, TypeError):
        return None


def _page_header_info(text: str) -> tuple[str | None, str | None]:
    stmt = None
    acct = None
    # Monthly/landscape commodity statements can print either "Monthly Statement"
    # or "Daily Statement" in the page header.  Prefer explicit header dates over
    # the first trade date on the page.
    m = re.search(r"Monthly Statement\s+(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if not m:
        m = re.search(r"Daily Statement\s+(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if not m:
        m = re.search(r"Statement Date:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if not m:
        m = re.search(r"Statement Date:\s*\n\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if m:
        stmt = _normalize_any_date(m.group(1))
    else:
        m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", text)
        if m:
            stmt = _normalize_any_date(m.group(1))
    m = re.search(r"Account Number:\s*(\S+)", text)
    if m:
        acct = m.group(1)
    return stmt, acct


def _rows_from_words(page):
    # This statement is landscape/rotated: x0 is the visual row position and y0 is the visual column position.
    words = page.get_text('words')
    buckets = {}
    for x0, y0, x1, y1, w, *_ in words:
        key = round(x0 / 3.0) * 3.0
        buckets.setdefault(key, []).append((y0, w))
    out = []
    for key in sorted(buckets):
        items = sorted(buckets[key], key=lambda t: -t[0])  # visual left -> right
        out.append((key, items, ' '.join(w for _, w in items)))
    return out


def _rows_from_words_standard(page):
    """Group words into rows for a standard (non-rotated) landscape PDF.

    In PyMuPDF's coordinate system, x0 increases left→right and y0 increases
    top→bottom.  We group by y0 so each bucket is a visual row; within a row
    words are sorted left→right by x0.  Returns the same
    ``[(key, [(x0, word), ...], line_str), ...]`` shape as ``_rows_from_words``.
    """
    words = page.get_text('words')
    buckets: Dict[float, list] = {}
    for x0, y0, x1, y1, w, *_ in words:
        key = round(y0 / 3.0) * 3.0          # bucket by visual row (y0 ascending = top→bottom)
        buckets.setdefault(key, []).append((x0, w))
    out = []
    for key in sorted(buckets):               # top → bottom
        items = sorted(buckets[key])          # left → right by x0
        out.append((key, items, ' '.join(w for _, w in items)))
    return out


def _col(items, lo, hi):
    vals = [w for y, w in items if lo <= y <= hi]
    return ' '.join(vals).strip() or None


def _desc(items):
    vals = [w for y, w in items if 385 <= y <= 545]
    return ' '.join(vals).replace(' - ', ' - ').strip() or None


def _first_numeric_from_sources(pattern: str, *sources) -> float | None:
    """Return the first numeric capture matched in any source string."""
    for source in sources:
        s = " ".join(str(source or "").strip().split())
        if not s or s.lower() in {"none", "nan", "nat", "unknown", "other"}:
            continue
        m = re.search(pattern, s, re.I)
        if m:
            try:
                return _num_any(m.group(1))
            except Exception:
                return None
    return None


def _normalize_trigger_barrier(value: str | None) -> float | None:
    """Normalize Trigger/Barrier to the trigger/barrier price only.

    Older versions stored descriptive text such as
    ``Trigger: 0.7104; Daily Cons; BP: 69.36; OQ: 57`` in the
    Trigger/Barrier field. The user-facing column should contain only the
    trigger/barrier price (0.7104 in that example). BP and OQ are stored in
    separate Ref Price and Original Quantity fields.
    """
    return _first_numeric_from_sources(r"/?\b(?:Trigger|Barrier)\s*[:=]\s*([-+]?\d+(?:\.\d+)?)", value)


def _extract_trigger_barrier(explicit_value: str | None, description: str | None) -> float | None:
    """Extract only the Trigger/Barrier price from the explicit column or description."""
    return _first_numeric_from_sources(
        r"/?\b(?:Trigger|Barrier)\s*[:=]\s*([-+]?\d+(?:\.\d+)?)",
        explicit_value,
        description,
    )


def _extract_ref_price(explicit_value: str | None, description: str | None) -> float | None:
    """Extract OTC reference/base price (BP) from descriptions like BP: 69.36 or BP=86.58."""
    return _first_numeric_from_sources(r"\bBP\s*[:=]?\s*([-+]?\d+(?:\.\d+)?)", explicit_value, description)


def _extract_original_quantity(explicit_value: str | None, description: str | None) -> float | None:
    """Extract original quantity (OQ) from descriptions like OQ 57, OQ=2, or OQ109."""
    return _first_numeric_from_sources(r"\bOQ\s*[:=]?\s*([-+]?\d+(?:\.\d+)?)", explicit_value, description)


def _otc_strike_number(token: str | None) -> float | None:
    """Convert OTC accumulator strike tokens to a numeric price.

    ``0.8456`` is returned as 0.8456.  Some text extraction/OCR variants can
    drop the decimal point and produce ``08456``; for cotton/coffee style prices
    that should be interpreted as 0.8456.
    """
    t = str(token or "").strip()
    if not t:
        return None
    if "." in t:
        return _num_any(t)
    if re.match(r"^0\d{4,5}$", t):
        return int(t) / (10 ** (len(t) - 1))
    return _num_any(t)


def _extract_otc_strike(description: str | None) -> float | None:
    """Extract the base/strike price from OTC accumulator descriptions.

    Examples:
      ICE Cotton LVL1 0.6720 /Trigger:0.7104 ... -> 0.6720
      ICE Cotton 0.8456 Daily Consumer Accum No KO ... -> 0.8456
    """
    s = " ".join(str(description or "").strip().split())
    if not s:
        return None
    if not re.search(r"\b(ACCUM|TRIGGER|BARRIER|LVL\d+|DAILY\s+(?:CONS|PROD|CONSUMER|PRODUCER)|NO\s+KO)\b|/TRIGGER\s*:", s, re.I):
        return None
    patterns = [
        r"\bLVL\d+\s+(\d+(?:\.\d+)?)\s*(?=/Trigger|\bDaily\b|\bRange\b|\bAccum\b|\bNo\s+KO\b)",
        r"\b(?:ICE|BMF|CBOT|CME|NYMEX|NYME|IFUS|MATF|LME)?\s*[A-Za-z][A-Za-z0-9/&'\-]*(?:\s+[A-Za-z][A-Za-z0-9/&'\-]*){0,3}\s+(0?\d+(?:\.\d+)?)\s+(?=/Trigger|\bDaily\b|\bConsumer\b|\bProducer\b|\bRange\b|\bAccum\b|\bNo\s+KO\b)",
    ]
    for pat in patterns:
        m = re.search(pat, s, re.I)
        if m:
            return _otc_strike_number(m.group(1))
    return None


def _ensure_otc_position_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize OTC trigger/barrier, ref price, original quantity, and strike fields."""
    if df is None or df.empty:
        return df
    out = df.copy()
    desc = _coalesce_column(out, ["contract_description", "Contract Description"], default=None)
    explicit = _coalesce_column(out, ["trigger_barrier", "Trigger/Barrier", "triggerBarrier"], default=None)

    parsed_trigger = pd.Series([_extract_trigger_barrier(e, d) for e, d in zip(explicit, desc)], index=out.index)
    parsed_ref = pd.Series([_extract_ref_price(e, d) for e, d in zip(explicit, desc)], index=out.index)
    parsed_oq = pd.Series([_extract_original_quantity(e, d) for e, d in zip(explicit, desc)], index=out.index)
    parsed_strike = pd.Series([_extract_otc_strike(d) for d in desc], index=out.index)

    # Trigger/Barrier should be the numeric trigger/barrier price only.  If the
    # existing value was an old descriptive string and no trigger price can be
    # found, blank it rather than showing BP/OQ/Daily Cons text in this field.
    if "trigger_barrier" not in out.columns:
        out["trigger_barrier"] = parsed_trigger
    else:
        old = out["trigger_barrier"]
        old_text = old.astype(str).str.upper()
        old_looked_like_otc_text = old_text.str.contains(r"TRIGGER|BARRIER|BP|OQ|DAILY|KO", regex=True, na=False)
        out["trigger_barrier"] = old.where(~old_looked_like_otc_text, None)
        out["trigger_barrier"] = parsed_trigger.where(parsed_trigger.notna(), out["trigger_barrier"])

    for target, parsed in [("ref_price", parsed_ref), ("original_quantity", parsed_oq), ("strike", parsed_strike)]:
        if target not in out.columns:
            out[target] = parsed
        else:
            current = out[target]
            missing = current.isna() | current.astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "unknown", "other", "multiple"])
            out[target] = current.where(~missing, parsed)

    return out


def _nearby_monthly_description(rows, idx: int, current_desc: str | None) -> str | None:
    """Attach wrapped OTC description lines around a StoneX Markets row.

    Some Commodity Open Positions rows with OTC/accumulator details are extracted as:
      row N-1: contract description text containing /Trigger:...
      row N:   trade date, IDs, quantity, dates, prices, value
      row N+1: continuation description text such as Range w/Daily DU...

    The row with the actual date may have a blank or partial description column,
    so this stitches the nearby description bands back onto the dated row.
    """
    if idx < 0 or idx >= len(rows):
        return current_desc

    this_key = float(rows[idx][0])

    def clean_text(v):
        return " ".join(str(v or "").split()).strip()

    def desc_from_row(j: int) -> str | None:
        if j < 0 or j >= len(rows):
            return None
        key, items, line = rows[j]
        if abs(float(key) - this_key) > 10.0:
            return None
        # Do not pull text from another dated position/subtotal row.
        date_cell = _col(items, 735, 778)
        if date_cell and DATE_DDMMMYYYY_RE.match(date_cell):
            return None
        line_clean = clean_text(line)
        if line_clean.startswith(("Long Avg", "Short Avg", "Total", "Grand Total", "Net ", "Page ")):
            return None
        if "Trade Date" in line_clean or "MarketValue" in line_clean or "Open Positions" in line_clean:
            return None
        d = clean_text(_desc(items))
        if not d:
            return None
        # Only attach neighboring lines that look like wrapped OTC detail. This
        # prevents normal futures rows from borrowing unrelated text.
        if not re.search(r"\b(LVL|TRIGGER|BARRIER|RANGE|ACCUM|KO|BP|OQ|DAILY\s+CONS|DAILY\s+PROD|CONSUMER\s+ACCUM)\b|/TRIGGER\s*:", d, re.I):
            return None
        return d

    pieces: list[str] = []
    for j in [idx - 2, idx - 1]:
        d = desc_from_row(j)
        if d and d not in pieces:
            pieces.append(d)

    current_clean = clean_text(current_desc)
    if current_clean and current_clean not in pieces:
        pieces.append(current_clean)

    for j in [idx + 1, idx + 2]:
        d = desc_from_row(j)
        if d and d not in pieces:
            pieces.append(d)

    return " ".join(pieces) if pieces else current_desc

def _section_limits(rows, names):
    out = {}
    for key, items, line in rows:
        line_ns = line.replace(' ', '')
        for name in names:
            if name not in out:
                if name in line or name.replace(' ', '') in line_ns:
                    out[name] = key
    return out



def _is_pct_line(s: str | None) -> bool:
    return bool(re.match(r"^[-+]?\d+(?:\.\d+)?%$", str(s or "").strip()))


def _pct_to_num(s: str | None) -> float | None:
    if s is None:
        return None
    return _num_any(str(s).strip().replace("%", ""))


def _parse_monthly_fx_option_ndo_lines(text: str, stmt_date: str | None, account: str | None, page_no: int) -> list[dict]:
    """Parse Markets LLC FX Option Open Positions - Non Deliverables.

    Example layout:
      Trade Date, Trade Id, Type=NDO, Curr Pair, Buy/Sell, Put/Call,
      CCY1, CCY1 Buy/(Sell), CCY2, CCY2 Buy/(Sell), Strike Price,
      Barrier, Expiry Date, Value Date, Prem VD, % Premium, Market Price, Market Value.

    PyMuPDF emits each cell on a separate line and omits blank Barrier/Prem VD
    cells, so the parser consumes optional Barrier and Prem VD only when present.
    """
    rows: list[dict] = []
    lines = [" ".join(x.strip().split()) for x in text.splitlines() if x.strip()]
    if "FX Option Open Positions" not in text:
        # Continuation pages often repeat only the column header. Do not confuse
        # this with FX Spot/Forward; NDO option pages have Put/Call and Strike Price.
        if not ("Put/Call" in lines and "Strike Price" in lines and "NDO" in lines):
            return rows
    try:
        start = next(i for i, line in enumerate(lines) if "FX Option Open Positions" in line)
    except StopIteration:
        try:
            start = next(i for i, line in enumerate(lines) if line == "Trade Date")
        except StopIteration:
            return rows

    # Stop before page footer/account summary if present.
    stop = len(lines)
    for j in range(start + 1, len(lines)):
        if lines[j].startswith("Account Information") or lines[j].startswith("Disclaimers") or lines[j].startswith("Page "):
            stop = j
            break

    def is_next_row(idx: int) -> bool:
        return idx < stop and _is_dd_mmm_yyyy(lines[idx])

    i = start + 1
    while i < stop:
        line = lines[i]
        if not _is_dd_mmm_yyyy(line):
            i += 1
            continue
        try:
            if i + 12 >= stop:
                break
            trade_date = lines[i]
            trade_id = lines[i + 1].strip()
            trade_type = lines[i + 2].strip().upper()
            curr_pair = lines[i + 3].strip().upper()
            buy_sell = lines[i + 4].strip().title()
            put_call_raw = lines[i + 5].strip()
            ccy1 = lines[i + 6].strip().upper()
            ccy1_amount = _num_any(lines[i + 7])
            ccy2 = lines[i + 8].strip().upper()
            ccy2_amount = _num_any(lines[i + 9])
            strike = _num_any(lines[i + 10])
            j = i + 11

            if trade_type != "NDO":
                i += 1
                continue
            if not re.match(r"^[A-Z]{3}/[A-Z]{3}$", curr_pair):
                i += 1
                continue
            if not (_is_ccy_line(ccy1) and _is_ccy_line(ccy2)):
                i += 1
                continue
            if ccy1_amount is None or ccy2_amount is None or strike is None:
                i += 1
                continue

            barrier = None
            # Barrier is optional. When blank, the next cell is Expiry Date.
            if j < stop and not _is_dd_mmm_yyyy(lines[j]):
                barrier = _num_any(lines[j])
                if barrier is None and not _is_pct_line(lines[j]):
                    barrier = lines[j]
                j += 1

            expiry_date = lines[j] if j < stop and _is_dd_mmm_yyyy(lines[j]) else None
            if expiry_date:
                j += 1
            value_date = lines[j] if j < stop and _is_dd_mmm_yyyy(lines[j]) else None
            if value_date:
                j += 1

            prem_vd = None
            if j < stop and _is_dd_mmm_yyyy(lines[j]):
                prem_vd = lines[j]
                j += 1

            premium_pct = None
            if j < stop and _is_pct_line(lines[j]):
                premium_pct = _pct_to_num(lines[j])
                j += 1

            market_price = _num_any(lines[j]) if j < stop else None
            if j < stop:
                j += 1
            market_value = _num_any(lines[j]) if j < stop else None
            if j < stop:
                j += 1

            if market_price is None or market_value is None:
                i += 1
                continue

            # Option direction comes from the Buy/Sell column, not the sign of CCY1 amount.
            qty_abs = abs(ccy1_amount)
            is_buy = buy_sell.upper().startswith("BUY")
            long_qty = qty_abs if is_buy else None
            short_qty = qty_abs if buy_sell.upper().startswith("SELL") else None
            option_type = "Call" if "CALL" in put_call_raw.upper() else ("Put" if "PUT" in put_call_raw.upper() else None)
            product = f"FX {curr_pair}"
            expiry_iso = _normalize_any_date(expiry_date)
            value_iso = _normalize_any_date(value_date)
            ref_month = _ref_month_from_date(expiry_date) if expiry_date else (_ref_month_from_date(value_date) if value_date else None)
            contract_desc = f"NDO {curr_pair} {put_call_raw} Strike {strike:g}" if strike is not None else f"NDO {curr_pair} {put_call_raw}"
            raw_line = " ".join(lines[i:j])

            row = {
                "trade_date": trade_date,
                "trade_date_iso": _normalize_any_date(trade_date, stmt_date),
                "trade_id": trade_id,
                "type": "NDO",
                "trade_type": "NDO",
                "curr_pair": curr_pair,
                "buy_sell": buy_sell,
                "option_type": option_type,
                "option_type_raw": put_call_raw,
                "quantity": qty_abs,
                "long": long_qty,
                "short": short_qty,
                "side": buy_sell,
                "expiry_date": expiry_iso,
                "expiration_date": expiry_iso,
                "value_date": value_date,
                "delivery_date": value_date,
                "contract_date": value_date,
                "settlement_date": value_date,
                "ref_month": ref_month,
                "contract_month": ref_month.split("-", 1)[0] if ref_month and "-" in ref_month else None,
                "contract_year": ref_month.split("-", 1)[1] if ref_month and "-" in ref_month else None,
                "contract_description": contract_desc,
                "product": product,
                "product_name": product,
                "exchange": "FX",
                "currency": ccy2,
                "primary_amount": ccy1_amount,
                "primary_currency": ccy1,
                "ccy_1": ccy1,
                "ccy_1_amount": ccy1_amount,
                "secondary_amount": ccy2_amount,
                "secondary_currency": ccy2,
                "ccy_2": ccy2,
                "ccy_2_amount": ccy2_amount,
                "strike": strike,
                "strike_price": strike,
                "trigger_barrier": barrier,
                "barrier": barrier,
                "premium_percent": premium_pct,
                "premium_pct": premium_pct,
                "market_price": market_price,
                "market_value": market_value,
                "market_value_signed": market_value,
                "nov": market_value,
                "statement_date": stmt_date,
                "account_number": account,
                "page": page_no,
                "source_section": "FX Option Open Positions - Non Deliverables",
                "source_system": "StoneX FX Option NDO",
                "source_line": raw_line,
            }
            rows.append(row)
            i = j
            continue
        except Exception:
            i += 1
            continue
    return rows


def _parse_monthly_fx_spot_forward_lines(text: str, stmt_date: str | None, account: str | None, page_no: int) -> list[dict]:
    """Parse StoneX Markets LLC FX Spot/Forward Open Positions rows from text lines.

    The Markets LLC layout usually extracts each FX row as one full line, e.g.
    ``12-Mar-2026 47023370 0 FX FWD AUD/USD AUD 270,000.00 0.7083 USD
    (191,241.00) 10-Jul-2026 0.721212 USD 3,486.12 $3,486.12``.
    Older IFL FX parser versions expected one cell per line, so this handles the
    one-line variant used in the uploaded statement.
    """
    rows: list[dict] = []
    in_fx = False
    for raw in text.splitlines():
        line = " ".join(raw.strip().split())
        if not line:
            continue
        if line == "FX Spot/Forward Open Positions":
            in_fx = True
            continue
        if in_fx and (line.startswith("Page ") or line in {"Disclaimers", "Account Information"}):
            in_fx = False
            continue
        if not in_fx or not _is_dd_mmm_yyyy(line.split()[0] if line.split() else ""):
            continue

        toks = line.split()
        try:
            # Required leading cells.
            trade_date = toks[0]
            trade_id = toks[1]
            global_id = toks[2]
            if len(toks) < 14 or toks[3].upper() != "FX":
                continue
            # Type can be "FX FWD", "FX SPOT", etc.; consume two tokens.
            trade_type = " ".join(toks[3:5])
            curr_pair = toks[5].upper()
            ccy1 = toks[6].upper()
            if not re.match(r"^[A-Z]{3}/[A-Z]{3}$", curr_pair):
                continue
            if not _is_ccy_line(ccy1):
                continue

            ccy1_amount = _num_any(toks[7])
            trade_price = _num_any(toks[8])
            ccy2 = toks[9].upper()
            ccy2_amount = _num_any(toks[10])
            j = 11
            fixing_date = None
            value_date = None
            if j < len(toks) and _is_dd_mmm_yyyy(toks[j]):
                first_date = toks[j]
                j += 1
                if j < len(toks) and _is_dd_mmm_yyyy(toks[j]):
                    fixing_date = first_date
                    value_date = toks[j]
                    j += 1
                else:
                    value_date = first_date
            if j + 3 >= len(toks):
                continue
            market_price = _num_any(toks[j])
            pnl_ccy = toks[j + 1].upper()
            native_pnl = _num_any(toks[j + 2])
            market_value = _num_any(toks[j + 3])

            if ccy1_amount is None or trade_price is None or not _is_ccy_line(ccy2) or ccy2_amount is None:
                continue
            if market_price is None or not _is_ccy_line(pnl_ccy) or market_value is None:
                continue

            product = f"FX {ccy1}/{ccy2}"
            ref_month = _ref_month_from_date(value_date) if value_date else None
            row = {
                "trade_date": trade_date,
                "trade_date_iso": _normalize_any_date(trade_date, stmt_date),
                "trade_id": trade_id,
                "global_id": global_id,
                "type": trade_type,
                "curr_pair": curr_pair,
                "quantity": abs(ccy1_amount),
                "long": ccy1_amount if ccy1_amount > 0 else None,
                "short": abs(ccy1_amount) if ccy1_amount < 0 else None,
                "side": "Long" if ccy1_amount > 0 else ("Short" if ccy1_amount < 0 else None),
                "value_date": value_date,
                "fixing_date": fixing_date,
                "delivery_date": value_date,
                "contract_date": value_date,
                "settlement_date": value_date,
                "expiry_date": _normalize_any_date(value_date),
                "ref_month": ref_month,
                "contract_month": ref_month.split("-", 1)[0] if ref_month and "-" in ref_month else None,
                "contract_year": ref_month.split("-", 1)[1] if ref_month and "-" in ref_month else None,
                "contract_description": f"{product} {value_date or ''}".strip(),
                "product": product,
                "product_name": product,
                "exchange": "FX",
                "currency": ccy2,
                "primary_amount": ccy1_amount,
                "primary_currency": ccy1,
                "ccy_1": ccy1,
                "ccy_1_amount": ccy1_amount,
                "secondary_amount": ccy2_amount,
                "secondary_currency": ccy2,
                "ccy_2": ccy2,
                "ccy_2_amount": ccy2_amount,
                "rate": trade_price,
                "trade_price": trade_price,
                "market_price": market_price,
                "pnl_ccy": pnl_ccy,
                "native_pnl": native_pnl,
                "market_value": market_value,
                "market_value_signed": market_value,
                "statement_date": stmt_date,
                "account_number": account,
                "page": page_no,
                "source_section": "FX Spot/Forward Open Positions",
                "source_system": "StoneX FX Spot/Forward",
                "source_line": line,
            }
            rows.append(row)
        except Exception:
            continue
    return rows


# ---------------------------------------------------------------------------
# Murex column-position helpers  (standard landscape coordinate system)
# ---------------------------------------------------------------------------
# The Murex PDF uses a standard landscape page (1682 × 1189 pts in PyMuPDF).
# x0 increases left→right (column identity) and y0 increases top→bottom (row
# identity).  _rows_from_words_standard groups words by y0 so each bucket is
# a visual row; items within a row are (x0, word) sorted left→right.
# Column boundaries are therefore x0-based, not y0-based.

# Two-word column header pairs in left→right order (ascending x0)
_MUREX_TWO_WORD_COLS: Dict[tuple, str] = {
    ('Trade', 'Date'): 'Trade Date',
    ('Trade', 'Id'): 'Trade Id',
    ('Global', 'Id'): 'Global Id',
    ('Contract', 'Description'): 'Contract Description',
    ('Start', 'Date'): 'Start Date',
    ('End', 'Date'): 'End Date',
    ('Ref', 'Month'): 'Ref Month',
    ('Native', 'MV'): 'Native MV',
    ('Market', 'Price'): 'Market Price',
    ('Trade', 'Price'): 'Trade Price',
    ('Ccy', '1'): 'Ccy 1',
    ('Ccy', '2'): 'Ccy 2',
    ('CCY', '1'): 'Ccy 1',
    ('CCY', '2'): 'Ccy 2',
    ('CCY1', 'Buy/(Sell)'): 'Ccy 1 Amt',
    ('CCY2', 'Buy/(Sell)'): 'Ccy 2 Amt',
    ('Ccy1', 'Buy/(Sell)'): 'Ccy 1 Amt',
    ('Ccy2', 'Buy/(Sell)'): 'Ccy 2 Amt',
}

# Single-word column header tokens
_MUREX_ONE_WORD_COLS: set = {
    'Long', 'Short', 'Trigger/Barrier', 'MarketValue', 'Commission', 'Premium',
    'Buy/(Sell)', 'Sell/(Buy)',  # FX CCY amount sub-columns — needed to create boundary
    'Type', 'Notional',
}

# Concatenated two-word headers (PyMuPDF may merge adjacent words with no space char)
_MUREX_CONCAT_COLS: Dict[str, str] = {
    'TradeDate': 'Trade Date',
    'TradeId': 'Trade Id',
    'GlobalId': 'Global Id',
    'ContractDescription': 'Contract Description',
    'StartDate': 'Start Date',
    'EndDate': 'End Date',
    'RefMonth': 'Ref Month',
    'NativeMV': 'Native MV',
    'MarketPrice': 'Market Price',
    'TradePrice': 'Trade Price',
    'Ccy1': 'Ccy 1',
    'Ccy2': 'Ccy 2',
    'CCY1': 'Ccy 1',
    'CCY2': 'Ccy 2',
}


def _murex_build_col_ranges(items: list) -> Dict[str, tuple]:
    """Build column-name → (x_lo, x_hi) ranges from a Murex header row.

    ``items`` is ``[(x0_float, word_str), ...]`` sorted left→right by x0
    (as returned by ``_rows_from_words_standard``).

    Adjacent column centres are used to compute midpoint boundaries so that
    every x-pixel belongs to exactly one column.
    """
    col_centers: list[tuple[str, float]] = []
    # items already sorted left→right (ascending x0)
    i = 0
    while i < len(items):
        x0, w = items[i]
        # Try two-word column header (consecutive words)
        if i + 1 < len(items):
            x1, w1 = items[i + 1]
            pair = (w, w1)
            if pair in _MUREX_TWO_WORD_COLS:
                col_centers.append((_MUREX_TWO_WORD_COLS[pair], (x0 + x1) / 2.0))
                i += 2
                continue
        # Try concatenated two-word header (PyMuPDF merged adjacent words)
        if w in _MUREX_CONCAT_COLS:
            col_centers.append((_MUREX_CONCAT_COLS[w], float(x0)))
            i += 1
            continue
        # Try single-word column header
        if w in _MUREX_ONE_WORD_COLS:
            col_centers.append((w, float(x0)))
        i += 1

    if len(col_centers) < 5:
        return {}

    # Sort by ascending x0 (left → right) and compute midpoint boundaries
    col_centers.sort(key=lambda t: t[1])
    result: Dict[str, tuple] = {}
    for idx, (name, cx) in enumerate(col_centers):
        x_lo = (cx + col_centers[idx - 1][1]) / 2.0 if idx > 0 else cx - 80.0
        x_hi = (cx + col_centers[idx + 1][1]) / 2.0 if idx < len(col_centers) - 1 else cx + 80.0
        result[name] = (x_lo, x_hi)
    return result


def _murex_gcol(items: list, col_ranges: Dict[str, tuple], col_name: str) -> str | None:
    """Return concatenated words within *col_name*'s x-range from *items*."""
    rang = col_ranges.get(col_name)
    if not rang:
        return None
    x_lo, x_hi = rang
    vals = [w for x, w in items if x_lo <= x <= x_hi]
    return ' '.join(vals).strip() or None


def extract_murex_statement(pdf_bytes: bytes, include_open_positions: bool = True) -> Dict[str, pd.DataFrame]:
    """Parse Murex-format StoneX daily statements.

    Uses the same ``_rows_from_words`` word-position approach as ``extract_monthly``
    (rotated-landscape coordinate system: x0 = visual row, y0 = visual column).
    Column positions are discovered dynamically from the header row on each page,
    so the extractor is robust to minor layout shifts and does not rely on hard-coded
    y-coordinate ranges.

    Handles both ``Commodity New Trades`` (Commission / Premium columns) and
    ``Commodity Open Positions`` (Market Price / MarketValue columns) on the same
    page.  Numeric values with BRL (``R$``) or USD (``$``) prefixes are parsed by
    ``_murex_num``.
    """
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    trades_list: List[dict] = []
    open_pos_list: List[dict] = []

    for pno, page in enumerate(doc, start=1):
        text = page.get_text('text')
        stmt_date, account = _page_header_info(text)
        # Choose row-grouping axis based on page rotation.
        # Standard landscape (rotation=0, e.g. G8669 1682×1189): group by y0.
        # Rotated landscape (rotation=90/270, e.g. daily statements): group by x0.
        if page.rotation in (90, 270):
            rows = _rows_from_words(page)
        else:
            rows = _rows_from_words_standard(page)

        # ── Detect column layouts for each table section on this page ────────
        # A page can have two header rows: one for New Trades, one for Open Positions.
        col_ranges_new: Dict[str, tuple] = {}
        col_ranges_open: Dict[str, tuple] = {}
        header_keys: list[float] = []

        # Header row of the FX Option NDO section, when present. The dedicated
        # NDO parser handles those rows; the commodity column extractor would
        # mis-read them, so we use this key as an upper bound for the commodity
        # open-positions section instead.
        fx_option_header_key: float | None = None

        for key, items, line in rows:
            lc = ' '.join(line.split())
            lc_ns = lc.replace(' ', '')
            has_td = 'Trade Date' in lc or 'TradeDate' in lc_ns
            has_cd = ('Contract Description' in lc or 'ContractDescription' in lc_ns
                      or 'CCY1' in lc or 'Ccy1' in lc_ns or 'Ccy 1' in lc)
            if not has_td or not has_cd:
                continue
            # FX Option header signature — leave it for the NDO parser.
            if (
                ('Curr Pair' in lc or 'Put/Call' in lc or 'Buy/Sell' in lc)
                and 'Contract Description' not in lc and 'ContractDescription' not in lc_ns
            ):
                if fx_option_header_key is None or key < fx_option_header_key:
                    fx_option_header_key = key
                continue
            cr = _murex_build_col_ranges(items)
            if not cr:
                continue
            header_keys.append(key)
            if 'Commission' in cr or 'Premium' in cr:
                col_ranges_new = cr
            if 'MarketValue' in cr or 'Market Price' in cr:
                col_ranges_open = cr

        if not col_ranges_new and not col_ranges_open:
            # Page has only an FX Option header; commodity parser has nothing
            # to do. The FX NDO parser is invoked unconditionally below.
            if fx_option_header_key is None:
                continue

        # ── Detect section boundaries ─────────────────────────────────────────
        limits = _section_limits(rows, [
            'Commodity New Trades', 'Commodity Open Positions',
            'Open Positions and Market Values', 'Itemized Cash',
            'FX Option Open Positions',
            'Account Information', 'Disclaimers',
        ])
        new_start = limits.get('Commodity New Trades', -1)
        pos_start = limits.get('Commodity Open Positions', -1)
        # Continuation pages repeat only the column header row — treat that row
        # as the start of the open-positions section if no explicit section marker.
        if pos_start == -1 and col_ranges_open and header_keys:
            pos_start = min(header_keys)
        account_end = limits.get('Account Information', 10 ** 9)
        disclaim_end = limits.get('Disclaimers', 10 ** 9)
        itemized_end = limits.get('Itemized Cash', 10 ** 9)
        # An FX Option header on a continuation page (no title text) still bounds
        # the open-positions section — those rows must go to the NDO parser only.
        fx_option_start = min(
            limits.get('FX Option Open Positions', 10 ** 9),
            fx_option_header_key if fx_option_header_key is not None else 10 ** 9,
        )

        # ── Parse data rows ───────────────────────────────────────────────────
        min_header_key = min(header_keys) if header_keys else -1

        def _section_for(k: float) -> str | None:
            sec: str | None = None
            if new_start != -1 and k > new_start:
                upper = min(
                    pos_start if pos_start != -1 else 10 ** 9,
                    itemized_end, account_end, disclaim_end, fx_option_start,
                )
                if k < upper:
                    sec = 'new_trades'
            if include_open_positions and pos_start != -1 and k > pos_start:
                if k < min(account_end, disclaim_end, fx_option_start):
                    sec = 'open_positions'
            if sec is None and col_ranges_open and new_start == -1 and pos_start != -1:
                if k > pos_start and k < min(account_end, disclaim_end, fx_option_start):
                    sec = 'open_positions'
            return sec

        # Identify data-row y-keys (rows with a valid Trade Date) so we can
        # bound multi-line Contract Description lookups. Some Murex statements
        # wrap descriptions like "BMF Corn 70.0000 Euro Option Put BRL/BAG - /
        # Cash Settled" onto y-buckets above and below the data row's y.
        data_row_keys: list[float] = []
        for k, its, _ln in rows:
            if k in header_keys:
                continue
            sec = _section_for(k)
            if sec is None:
                continue
            cr = col_ranges_new if sec == 'new_trades' else col_ranges_open
            if not cr:
                continue
            dt = _murex_gcol(its, cr, 'Trade Date')
            if dt and DATE_DDMMMYYYY_RE.match(dt):
                data_row_keys.append(k)
        data_row_keys.sort()

        # Rotation drives in-bucket reading order. Standard landscape pages
        # read left-to-right (ascending x within a y-bucket); rotated pages
        # (90/270) read top-to-bottom visually, which is descending y within
        # an x-bucket because rotation flips the visual y-axis.
        is_rotated_page = page.rotation in (90, 270)
        data_row_keys_set = set(data_row_keys)

        def _description_spanning(current_key: float, col_ranges: Dict[str, tuple]) -> str:
            """Reassemble a wrapped Contract Description.

            For both rotated and standard layouts, the wrap text lives in
            buckets adjacent to the data row's key. A midpoint cutoff between
            this data row and its neighbors keeps each row's wrapped text with
            its own row.

            Words from the data row's OWN bucket are filtered to the Contract
            Description column range, since that bucket also holds dates,
            prices, and quantities. Words from neighbor "wrap" buckets — those
            that aren't themselves data rows — are taken in full, because
            wrapped description text routinely overflows the column header's
            bounding box on rotated pages.
            """
            rng = col_ranges.get('Contract Description')
            if not rng:
                return ''
            x_lo, x_hi = rng
            idx = data_row_keys.index(current_key) if current_key in data_row_keys else -1
            prev_key = data_row_keys[idx - 1] if idx > 0 else (current_key - 24.0)
            next_key = data_row_keys[idx + 1] if 0 <= idx < len(data_row_keys) - 1 else (current_key + 24.0)
            lower_bound = (prev_key + current_key) / 2.0
            upper_bound = (current_key + next_key) / 2.0
            collected: list[tuple[float, float, str]] = []
            for k2, its2, _ in rows:
                if k2 in header_keys:
                    continue
                if not (lower_bound < k2 < upper_bound):
                    continue
                is_data_bucket = k2 in data_row_keys_set
                if is_data_bucket:
                    # The data row's own bucket holds dates/prices/quantities
                    # alongside any description text. Keep only words inside
                    # the Contract Description column range.
                    for x, w in its2:
                        if x_lo <= x <= x_hi:
                            collected.append((k2, x, w))
                else:
                    # A neighbor bucket is treated as wrapped description text
                    # only if it has at least one word in the description
                    # column range. Subtotal rows (which carry "62.50" or
                    # "($558.84)" in the quantity/value columns but nothing in
                    # the description column) fail this gate and are skipped.
                    has_desc_anchor = any(x_lo <= x <= x_hi for x, _ in its2)
                    if not has_desc_anchor:
                        continue
                    # Wrap text routinely overflows the header's bounding box,
                    # so include every word from this description-only bucket.
                    for x, w in its2:
                        collected.append((k2, x, w))
            # Reading order: outer ascending k2 always. Inner order depends on
            # rotation — rotated pages read descending y in each x-bucket;
            # standard pages read ascending x in each y-bucket.
            inner_sign = -1.0 if is_rotated_page else 1.0
            collected.sort(key=lambda t: (t[0], inner_sign * t[1]))
            return ' '.join(w for _, _, w in collected).strip()

        for key, items, line in rows:
            # Skip header rows themselves
            if key in header_keys:
                continue

            section = _section_for(key)
            if section is None:
                continue

            col_ranges = col_ranges_new if section == 'new_trades' else col_ranges_open
            if not col_ranges:
                continue

            def gcol(name: str) -> str | None:
                return _murex_gcol(items, col_ranges, name)

            # Trade date must be present and valid
            date = gcol('Trade Date')
            if not date or not DATE_DDMMMYYYY_RE.match(date):
                continue

            # Skip Long Avg / Short Avg summary rows
            trigger_raw = gcol('Trigger/Barrier')
            if trigger_raw and re.search(r'\bAvg\b', trigger_raw, re.I):
                continue

            # Contract Description may wrap onto y-buckets above/below the data
            # row. Replace the single-bucket value with one that spans the full
            # row's y-window, so options like "BMF Corn 70.0000 Euro Option Put
            # BRL/BAG - Cash Settled" are captured even when wrapped.
            desc = _description_spanning(key, col_ranges) or (gcol('Contract Description') or '')
            long_qty = _murex_num(gcol('Long'))
            short_qty = _murex_num(gcol('Short'))
            if long_qty is not None and short_qty is None:
                qty = long_qty
            elif short_qty is not None and long_qty is None:
                qty = short_qty
            elif long_qty is not None and short_qty is not None:
                qty = long_qty - short_qty
            else:
                qty = None

            end_date_raw = gcol('End Date')
            ref_month_raw = gcol('Ref Month')
            ref_month = _normalize_ref_month(ref_month_raw)
            end_date_iso = _normalize_any_date(end_date_raw)

            desc_upper = desc.upper()
            # Accumulator detection — checked before SWAP because descriptions
            # like "ICE Cotton LVL1 ... Daily Cons Range w/Daily DU" don't
            # contain the word "Accum" but are still accumulators. Markers:
            # ACCUM/ACCUMULATOR, LVL\d+ tier, /Trigger:, Daily Cons/Prod,
            # Daily Consumer/Producer, No KO, BP=/OQ= accumulator footers.
            is_accumulator = bool(
                re.search(r'\b(ACCUMULATOR|ACCUM|NO\s+KO)\b', desc_upper)
                or re.search(r'\bLVL\d+\b', desc_upper)
                or re.search(r'/TRIGGER\s*:', desc_upper)
                or re.search(r'\bDAILY\s+(?:CONS|PROD|CONSUMER|PRODUCER)\b', desc_upper)
                or re.search(r'\b(?:BP|OQ)\s*[:=]', desc_upper)
            )
            if is_accumulator:
                instr_type = 'OTC Accumulator'
            elif re.search(r'\bSWAP\b', desc_upper):
                instr_type = 'OTC SWAP'
            elif re.search(r'\bOPTION\b', desc_upper):
                instr_type = 'OTC CALL' if 'CALL' in desc_upper else ('OTC PUT' if 'PUT' in desc_upper else 'OTC OPTION')
            else:
                instr_type = 'OTC'

            m_prod = re.match(r'^([A-Z]+(?:\s+[A-Za-z][A-Za-z\s]*?)?)\s+(?:[\d.,]+\s+)?Euro\s+', desc, re.I)
            # Normalize through the central product-name mapping so business
            # names match across all parsers (e.g. "CBOT Soybean Oil" → "Soybean
            # Oil"). Otherwise Murex rows would aggregate separately from rows
            # whose product was set by enrich_open_positions_metadata.
            product = _normalize_product_from_description(m_prod.group(1), desc) if m_prod else None

            # FX pair detection: use dedicated Ccy 1 / Ccy 2 columns if present.
            # Only accept the value if it looks like a 3-letter currency code.
            if not product:
                _ccy_re = re.compile(r'^[A-Z]{3}$')
                ccy1_val = (gcol('Ccy 1') or '').strip().upper()
                ccy2_val = (gcol('Ccy 2') or '').strip().upper()
                if _ccy_re.match(ccy1_val) and _ccy_re.match(ccy2_val):
                    product = f'{ccy1_val}/{ccy2_val}'
                elif _ccy_re.match(ccy1_val):
                    product = ccy1_val

            option_type: str | None = None
            strike: float | None = None
            if 'OPTION' in desc_upper:
                option_type = 'Call' if 'CALL' in desc_upper else ('Put' if 'PUT' in desc_upper else None)
                sm = re.search(r'\b([\d.,]+)\s+Euro\s+Option\b', desc, re.I)
                if sm:
                    strike = _murex_num(sm.group(1))

            row_dict: dict = {
                'statement_date': stmt_date,
                'account_number': account,
                'page': pno,
                'trade_date': date,
                'trade_date_iso': _normalize_any_date(date),
                'trade_id': gcol('Trade Id'),
                'global_id': gcol('Global Id'),
                'long': long_qty,
                'short': short_qty,
                'quantity': qty,
                'contract_description': desc,
                'start_date': _normalize_any_date(gcol('Start Date')),
                'expiryDate': end_date_iso,
                'end_date': end_date_iso,
                'ref_month': ref_month,
                'contract_month': ref_month.split('-')[0] if ref_month and '-' in ref_month else None,
                'contract_year': ref_month.split('-')[1] if ref_month and '-' in ref_month else None,
                'Type': instr_type,
                'position_type': instr_type,
                'product': product,
                'option_type': option_type,
                'Call/Put': option_type,
                'strike': strike,
                'strikePrice': strike,
                'source_section': 'Open Positions' if section == 'open_positions' else 'Executed Trades',
                'source_line': line,
            }

            if section == 'open_positions':
                row_dict.update({
                    'trade_price': _murex_num(gcol('Trade Price')),
                    'market_price': _murex_num(gcol('Market Price')),
                    'native_mv': _murex_num(gcol('Native MV')),
                    'market_value': _murex_num(gcol('MarketValue')),
                    'market_value_signed': _murex_num(gcol('MarketValue')),
                })
                open_pos_list.append(row_dict)
            else:
                row_dict.update({
                    'trade_price': _murex_num(gcol('Trade Price')),
                    'commission': _murex_num(gcol('Commission')),
                    'premium': _murex_num(gcol('Premium')),
                })
                trades_list.append(row_dict)

        # FX Option NDO rows have their own 18-column layout that the commodity
        # column extractor cannot read. Route them through the dedicated parser.
        # Continuation pages omit the section title — match the NDO column-header
        # signature instead (mirrors _parse_monthly_fx_option_ndo_lines).
        if include_open_positions and (
            'FX Option Open Positions' in text
            or ('Put/Call' in text and 'Strike Price' in text and 'NDO' in text)
        ):
            for ndo_row in _parse_monthly_fx_option_ndo_lines(text, stmt_date, account, pno):
                ndo_row.setdefault('Type', 'NDO')
                ndo_row.setdefault('position_type', 'NDO')
                open_pos_list.append(ndo_row)

    tables: Dict[str, pd.DataFrame] = {
        'Executed Trades': pd.DataFrame(trades_list),
        'Purchase & Sale': pd.DataFrame(),
        'Closed Positions': pd.DataFrame(),
        'Receives Delivers': pd.DataFrame(),
        'Journal Entries': pd.DataFrame(),
        'Realized Gain and Loss': pd.DataFrame(),
        'Realized PNL Summary': pd.DataFrame(),
        'Open Positions': pd.DataFrame(open_pos_list),
        'Notes': pd.DataFrame(),
        'Exceptions': pd.DataFrame(),
    }
    tables = enrich_open_positions_metadata(tables)
    tables['Summary'] = build_summary(tables)
    return tables


def extract_monthly(pdf_bytes: bytes, include_open_positions: bool = True) -> Dict[str, pd.DataFrame]:
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    trades, open_pos, cash_settlements, realized, notes, exceptions = [], [], [], [], [], []
    for pno, page in enumerate(doc, start=1):
        text = page.get_text('text')
        stmt_date, account = _page_header_info(text)
        rows = _rows_from_words(page)
        limits = _section_limits(rows, [
            'Commodity New Trades', 'Cash Settlements', 'Commodity Cash Settlements',
            'Realized Gain and Loss', 'Commodity Open Positions',
            'FX Spot/Forward Open Positions', 'FX Option Open Positions - Non Deliverables', 'Account Information', 'Disclaimers'
        ])
        new_start = limits.get('Commodity New Trades', -1)
        cash_start = limits.get('Cash Settlements', 10**9)
        realized_start = limits.get('Realized Gain and Loss', 10**9)
        pos_start = limits.get('Commodity Open Positions', -1)
        fx_spot_start = limits.get('FX Spot/Forward Open Positions', -1)
        fx_ndo_start = limits.get('FX Option Open Positions - Non Deliverables', -1)
        account_start = limits.get('Account Information', 10**9)
        disclaimer_start = limits.get('Disclaimers', 10**9)

        # Continuation pages often omit the "Commodity Open Positions" title and
        # start directly with the repeated Open Positions column header.  Treat
        # that header as the start of the open-position table; otherwise pages 2+
        # of Cotton/Coffee statements can be skipped entirely.
        if pos_start == -1:
            for header_key, _header_items, header_line in rows:
                if (
                    'Trade Date Trade Id Global Id Long Short Contract Description' in header_line
                    and 'MarketValue' in header_line
                ):
                    pos_start = header_key
                    break
        if fx_spot_start == -1:
            for header_key, _header_items, header_line in rows:
                if (
                    'Trade Date Trade Id Global Id Type Curr Pair CCY1' in header_line
                    and 'Market Value' in header_line
                ):
                    fx_spot_start = header_key
                    break

        for row_idx, (key, items, line) in enumerate(rows):
            date = _col(items, 735, 778)
            if not date or not DATE_DDMMMYYYY_RE.match(date):
                continue
            section = None
            if (new_start != -1 and key > new_start) and key < min(cash_start, realized_start, pos_start if pos_start != -1 else 10**9):
                section = 'Executed Trades'
            elif 'Trade Id Long Short Type Description' in text or (limits.get('Commodity Cash Settlements', -1) != -1 and key > limits.get('Commodity Cash Settlements', -1) and key < realized_start):
                section = 'Cash Settlements'
            elif realized_start != 10**9 and key > realized_start and (pos_start == -1 or key < pos_start):
                section = 'Realized Gain and Loss'
            elif include_open_positions and fx_spot_start != -1 and key > fx_spot_start and key < min(account_start, disclaimer_start):
                section = 'FX Spot/Forward Open Positions'
            elif include_open_positions and pos_start != -1 and key > pos_start and key < min(account_start, disclaimer_start, fx_spot_start if fx_spot_start != -1 else 10**9, fx_ndo_start if 'fx_ndo_start' in locals() and fx_ndo_start != -1 else 10**9):
                section = 'Open Positions'
            else:
                continue

            common = {
                'statement_date': stmt_date,
                'account_number': account,
                'page': pno,
                'trade_date': date,
                'trade_id': _col(items, 695, 733),
                'global_id': _col(items, 645, 682),
                'long': _num_any(_col(items, 590, 620)),
                'short': _num_any(_col(items, 545, 575)),
                'contract_description': _desc(items),
                'end_date': _col(items, 298, 340),
                'ref_month': _col(items, 258, 296),
                'trigger_barrier': _col(items, 210, 258),
                'source_line': line,
            }
            common['contract_description'] = _nearby_monthly_description(rows, row_idx, common.get('contract_description'))
            trigger_barrier_raw = common.get('trigger_barrier')
            common['trigger_barrier'] = _extract_trigger_barrier(trigger_barrier_raw, common.get('contract_description'))
            common['ref_price'] = _extract_ref_price(trigger_barrier_raw, common.get('contract_description'))
            common['original_quantity'] = _extract_original_quantity(trigger_barrier_raw, common.get('contract_description'))
            common['side'] = 'Long' if common['long'] is not None else ('Short' if common['short'] is not None else None)
            common['quantity'] = common['long'] if common['long'] is not None else common['short']

            if section == 'Executed Trades':
                common.update({
                    'trade_price': _num_any(_col(items, 138, 178)),
                    'commission': _num_any(_col(items, 78, 125)),
                    'premium': _num_any(_col(items, 20, 65)),
                    'source_section': section,
                })
                trades.append(common)
            elif section == 'Open Positions':
                common.update({
                    'trade_price': _num_any(_col(items, 178, 210)),
                    'market_price': _num_any(_col(items, 125, 160)),
                    'native_mv': _num_any(_col(items, 72, 105)),
                    'market_value': _num_any(_col(items, 15, 65)),
                    'source_section': section,
                })
                open_pos.append(common)
            elif section == 'FX Spot/Forward Open Positions':
                trade_id = _col(items, 695, 733)
                global_id = _col(items, 670, 692)
                trade_type = _col(items, 620, 650)
                curr_pair = _col(items, 565, 590)
                ccy1 = _col(items, 540, 560)
                ccy1_amount = _num_any(_col(items, 455, 505))
                trade_price = _num_any(_col(items, 410, 435))
                ccy2 = _col(items, 390, 410)
                ccy2_amount = _num_any(_col(items, 300, 350))
                value_date = _col(items, 235, 260)
                market_price = _num_any(_col(items, 175, 195))
                pnl_ccy = _col(items, 150, 170)
                native_pnl = _num_any(_col(items, 75, 105))
                market_value = _num_any(_col(items, 15, 65))
                if curr_pair and ccy1 and ccy2 and ccy1_amount is not None and ccy2_amount is not None:
                    ref_month = _ref_month_from_date(value_date)
                    product = f"FX {ccy1}/{ccy2}"
                    common.update({
                        'trade_id': trade_id,
                        'global_id': global_id,
                        'type': trade_type,
                        'curr_pair': curr_pair,
                        'quantity': abs(ccy1_amount),
                        'long': ccy1_amount if ccy1_amount > 0 else None,
                        'short': abs(ccy1_amount) if ccy1_amount < 0 else None,
                        'side': 'Long' if ccy1_amount > 0 else ('Short' if ccy1_amount < 0 else None),
                        'contract_description': f"{product} {value_date or ''}".strip(),
                        'value_date': value_date,
                        'delivery_date': value_date,
                        'contract_date': value_date,
                        'settlement_date': value_date,
                        'expiry_date': _normalize_any_date(value_date),
                        'ref_month': ref_month,
                        'contract_month': ref_month.split('-', 1)[0] if ref_month and '-' in ref_month else None,
                        'contract_year': ref_month.split('-', 1)[1] if ref_month and '-' in ref_month else None,
                        'product': product,
                        'product_name': product,
                        'exchange': 'FX',
                        'currency': ccy2,
                        'primary_amount': ccy1_amount,
                        'primary_currency': ccy1,
                        'ccy_1': ccy1,
                        'ccy_1_amount': ccy1_amount,
                        'secondary_amount': ccy2_amount,
                        'secondary_currency': ccy2,
                        'ccy_2': ccy2,
                        'ccy_2_amount': ccy2_amount,
                        'rate': trade_price,
                        'trade_price': trade_price,
                        'market_price': market_price,
                        'pnl_ccy': pnl_ccy,
                        'native_pnl': native_pnl,
                        'market_value': market_value,
                        'market_value_signed': market_value,
                        'trigger_barrier': None,
                        'source_section': section,
                        'source_system': 'StoneX FX Spot/Forward',
                    })
                    open_pos.append(common)
            elif section == 'Cash Settlements':
                common.update({'source_section': section, 'trade_price': _num_any(_col(items, 138, 178)), 'cash_amount': _num_any(_col(items, 80, 130))})
                cash_settlements.append(common)
            elif section == 'Realized Gain and Loss':
                common.update({'source_section': section, 'trade_price': _num_any(_col(items, 138, 178)), 'cash_flow': _num_any(_col(items, 20, 65))})
                realized.append(common)

        if include_open_positions and 'FX Spot/Forward Open Positions' in text:
            existing_fx = {
                (r.get('page'), str(r.get('trade_id')), str(r.get('source_section')))
                for r in open_pos
            }
            for fx_row in _parse_monthly_fx_spot_forward_lines(text, stmt_date, account, pno):
                fx_key = (fx_row.get('page'), str(fx_row.get('trade_id')), str(fx_row.get('source_section')))
                if fx_key not in existing_fx:
                    open_pos.append(fx_row)
                    existing_fx.add(fx_key)

        if include_open_positions and ('FX Option Open Positions' in text or ('Put/Call' in text and 'Strike Price' in text and 'NDO' in text)):
            existing_fx_option = {
                (r.get('page'), str(r.get('trade_id')), str(r.get('source_section')))
                for r in open_pos
            }
            for fx_row in _parse_monthly_fx_option_ndo_lines(text, stmt_date, account, pno):
                fx_key = (fx_row.get('page'), str(fx_row.get('trade_id')), str(fx_row.get('source_section')))
                if fx_key not in existing_fx_option:
                    open_pos.append(fx_row)
                    existing_fx_option.add(fx_key)

    tables: Dict[str, pd.DataFrame] = {
        'Executed Trades': pd.DataFrame(trades),
        'Purchase & Sale': pd.DataFrame(),
        'Receives Delivers': pd.DataFrame(cash_settlements),
        'Journal Entries': pd.DataFrame(),
        'Realized Gain and Loss': pd.DataFrame(realized),
        'Open Positions': pd.DataFrame(open_pos),
        'Notes': pd.DataFrame(notes),
        'Exceptions': pd.DataFrame(exceptions),
    }
    tables = enrich_open_positions_metadata(tables)
    tables['Summary'] = build_summary(tables)
    return tables


def looks_like_murex_statement(pdf_bytes: bytes) -> bool:
    """Detect Murex-format StoneX daily statements.

    The Murex layout adds a ``Start Date`` column between ``Contract Description``
    and ``End Date``, which is absent from all other monthly/daily formats.
    We also require a ``Ref Month`` column header and at least one Commodity
    section heading so we do not misfire on unrelated PDFs that happen to
    mention "Start Date".
    """
    try:
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        for page in list(doc)[:5]:
            text = page.get_text('text')
            if (
                'Start Date' in text
                and 'Ref Month' in text
                and re.search(r'Commodity\s+(?:New Trades|Open Positions)', text)
            ):
                return True
        return False
    except Exception:
        return False


def looks_like_monthly_statement(pdf_bytes: bytes) -> bool:
    """Detect StoneX Markets LLC / monthly-style statements.

    Earlier versions only checked page 1 for commodity headings.  FX-only
    Markets LLC statements can show only Account Summary / Account Information
    on the first pages, with ``Open Positions and Market Values`` and
    ``FX Spot/Forward Open Positions`` starting several pages later.  Scan a
    small prefix of pages so those files route to the monthly/Markets parser
    instead of falling through to the legacy parser.
    """
    try:
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        page_texts = []
        for page in list(doc)[:8]:
            page_texts.append(page.get_text('text'))
        text = "\n".join(page_texts)
        markers = (
            'Commodity New Trades',
            'Commodity Open Positions',
            'Open Positions and Market Values',
            'FX Spot/Forward Open Positions',
            'FX Spot/Forward New Trades',
            'FX Option Open Positions',
            'Monthly Statement',
        )
        if any(marker in text for marker in markers):
            return True
        return 'StoneX Markets LLC' in text and 'Account Number:' in text and (
            'FX Spot Forward' in text or 'Curr Pair' in text or 'Market Value of Open' in text
        )
    except Exception:
        return False




def _is_dd_mmm_yyyy(s: str) -> bool:
    return bool(re.match(r"^\d{1,2}-[A-Za-z]{3}-\d{4}$", str(s).strip()))


def _is_dd_mmm_yy(s: str) -> bool:
    return bool(re.match(r"^\d{1,2}-[A-Za-z]{3}-\d{2}$", str(s).strip()))


def _is_amount_line(s: str) -> bool:
    return bool(re.match(r"^\(?[\d,]+(?:\.\d+)?\)?$", str(s).strip()))


def _is_ccy_line(s: str) -> bool:
    return bool(re.match(r"^[A-Z]{3}$", str(s).strip()))


def _parse_lme_open_positions_from_lines(lines: list[str], stmt_date: str | None, account: str | None, page_no: int, broker: str | None = None) -> list[dict]:
    """Parse metal/LME open positions where PyMuPDF splits each visual row into multiple text lines."""
    rows: list[dict] = []
    section: str | None = None
    i = 0
    while i < len(lines):
        line = lines[i]
        if line == "FUTURES / OPTIONS OPEN POSITIONS":
            section = "fut_opt"
            i += 1
            continue
        if line == "LME AVERAGE OPEN POSITIONS":
            section = "lme_avg"
            i += 1
            continue
        if line.startswith("Diagnostic Key:") and "OPEN POSITIONS" in line:
            # wait for the actual table title
            i += 1
            continue
        if line.startswith("Diagnostic Key:") and section is not None:
            section = None
            i += 1
            continue

        if section == "fut_opt" and _is_dd_mmm_yyyy(line):
            try:
                qty = lines[i + 1]
                delivery_product = lines[i + 2]
                trade_price = lines[i + 3]
                j = i + 4
                price_type = None
                if j < len(lines) and not _is_ccy_line(lines[j]):
                    price_type = lines[j]
                    j += 1
                ccy = lines[j]
                amount = lines[j + 1]
                dm = re.match(r"^(?P<delivery_date>\d{1,2}-[A-Za-z]{3}-\d{2})\s+(?P<product>LME\s+.+)$", delivery_product)
                if dm and re.match(r"^[\d,]+$", qty) and _is_ccy_line(ccy) and _is_amount_line(amount):
                    raw_line = " ".join(lines[i:j+2])
                    row = {
                        "trade_date": line,
                        "trade_date_iso": _normalize_any_date(line, stmt_date),
                        "quantity": _num_any(qty),
                        "delivery_date": dm.group("delivery_date"),
                        "contract_date": dm.group("delivery_date"),
                        "contract_description": dm.group("product"),
                        "exchange": "LME",
                        "ref_month": _ref_month_from_date(dm.group("delivery_date")),
                        "trade_price": _num_any(trade_price),
                        "price_type": price_type,
                        "currency": ccy,
                        "market_value": _num_any(amount),
                        "market_value_signed": _num_any(amount),
                        "statement_date": stmt_date,
                        "account_number": account,
                        "broker_code": broker,
                        "page": page_no,
                        "source_section": "Futures / Options Open Positions",
                        "source_line": raw_line,
                    }
                    rows.append(row)
                    i = j + 2
                    continue
            except Exception:
                pass

        if section == "lme_avg" and _is_dd_mmm_yyyy(line):
            try:
                qty = lines[i + 1]
                price_start = lines[i + 2]
                end_date = lines[i + 3]
                settlement_date = lines[i + 4]
                if not re.match(r"^[\d,]+$", qty) or not _is_dd_mmm_yyyy(end_date) or not _is_dd_mmm_yyyy(settlement_date):
                    i += 1
                    continue
                ps = price_start.split()
                if len(ps) < 2 or not _is_dd_mmm_yyyy(ps[-1]):
                    i += 1
                    continue
                price = " ".join(ps[:-1])
                start_date = ps[-1]
                # Amount/currency are usually the next two text lines, but the order can be amount then ccy.
                amount = None
                ccy = None
                for k in range(i + 5, min(i + 9, len(lines))):
                    if amount is None and _is_amount_line(lines[k]):
                        amount = lines[k]
                    if ccy is None and _is_ccy_line(lines[k]):
                        ccy = lines[k]
                product = None
                delivery_date = None
                for k in range(i + 5, min(i + 22, len(lines))):
                    pm = LME_PRODUCT_RE.match(lines[k])
                    if pm and pm.group("delivery_date") == settlement_date:
                        delivery_date = pm.group("delivery_date")
                        product = pm.group("product")
                        break
                if amount is not None and ccy is not None:
                    raw_line = " ".join(lines[i:min(i+8, len(lines))])
                    row = {
                        "trade_date": line,
                        "trade_date_iso": _normalize_any_date(line, stmt_date),
                        "quantity": _num_any(qty),
                        "delivery_date": delivery_date or settlement_date,
                        "contract_date": delivery_date or settlement_date,
                        "contract_description": product,
                        "exchange": "LME",
                        "price": price,
                        "trade_price": price,
                        "start_date": start_date,
                        "end_date": end_date,
                        "settlement_date": settlement_date,
                        "ref_month": _ref_month_from_date(settlement_date),
                        "currency": ccy,
                        "market_value": _num_any(amount),
                        "market_value_signed": _num_any(amount),
                        "statement_date": stmt_date,
                        "account_number": account,
                        "broker_code": broker,
                        "page": page_no,
                        "source_section": "LME Average Open Positions",
                        "source_line": raw_line,
                    }
                    rows.append(row)
                    i += 5
                    continue
            except Exception:
                pass

        i += 1
    return rows




def _is_open_position_non_position_line(line: str) -> bool:
    """Skip header/footer/subtotal and non-futures cash collateral lines inside OPEN POSITIONS.

    Some StoneX daily statements print USTB/T-bill collateral lines in the OPEN
    POSITIONS section, for example:
        6/20/5 U1 300,000 USTB DUE 9/18/2025 US 296,823.75CR
    These are not futures/options position rows and should not be logged as parser
    exceptions.
    """
    s = " ".join(str(line or "").strip().split())
    if not s:
        return True
    u = s.upper()
    skip_tokens = (
        "OPEN POSITIONS",
        "CONTRACT DESCRIPTION-OPEN",
        "CONTINUED",
        "ACCOUNT TOTAL",
        "TOTAL",
        "PAGE",
        "STATEMENT",
        "BROUGHT FORWARD",
        "CARRIED FORWARD",
        "TRADE CARD",
        "TRADE AT",
        "AVG LONG",
        "AVG SHORT",
        "S.P.",
        "CLOSE",
    )
    if any(tok in u for tok in skip_tokens):
        return True
    # Treasury bill/cash collateral line, not a futures/options open position.
    if re.match(r"^\d{1,2}/\d{2}/\d\s+(?:[A-Z0-9]+\s+)?[A-Z]\d\s+[\d,]+\s+USTB\s+DUE\s+", u):
        return True
    # Do not try to parse non-position lines as position rows.
    if not re.match(r"^\d{1,2}/\d{2}/\d\b", s):
        return True
    if len(s.split()) < 8:
        return True
    return False

def _is_card_continuation_line(line: str) -> bool:
    """True for the trade/card id printed on the line after a StoneX daily row."""
    s = str(line or "").strip()
    if not s:
        return False
    if re.match(r"^\d{1,2}/\d{2}/\d\s+", s):
        return False
    if any(k in s for k in ["AVG LONG", "AVG SHORT", "CLOSE", "LTD-", "EX-", "DQ", "GROSS PROFIT"]):
        return False
    if s.startswith("*") or s.startswith("-") or s.startswith("TRADE "):
        return False
    return bool(re.match(r"^[A-Z][A-Z0-9]{3,}$", s))



def looks_like_asx_statement(pdf_bytes: bytes) -> bool:
    """Detect ASX/SFE-format statements from StoneX Financial Pty Ltd (Australian entity).

    These PDFs use the same stacked-text table layout as IFL statements but come from the
    Australian subsidiary and contain FUTURES / OPTIONS CONFIRMATIONS, OPEN POSITIONS, and
    PURCHASE & SALE sections identified by 'Diagnostic Key:' headers.
    """
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = "\n".join(page.get_text("text") for page in list(doc)[:4])
        return (
            "StoneX Financial Pty Ltd" in text
            and (
                "FUTURES / OPTIONS OPEN POSITIONS" in text
                or "FUTURES / OPTIONS CONFIRMATIONS" in text
                or "PURCHASE & SALE" in text
            )
        )
    except Exception:
        return False


def _parse_asx_trade_rows(
    lines: list[str],
    stmt_date: str | None,
    account: str | None,
    page_no: int,
    source_section: str = "Executed Trades",
) -> list[dict]:
    """Parse ASX FUTURES / OPTIONS CONFIRMATIONS rows (stacked PyMuPDF format).

    Rows have the structure:
      DD-Mon-YYYY          ← trade date
      <qty>                ← buy qty (alone on line = Buy side)
      Mon-YY EXCHANGE PRODUCT  ← delivery / product
      [Call|Put]           ← optional C/P
      [strike]             ← optional strike
      <trade_price>
      [BROKER_CODE]        ← e.g. SYSFA, SYMER  (optional)
      [CCY]                ← optional; futures often omit amount
      [amount]             ← optional (present for options premium)

    Sell rows have qty + delivery on the same line:
      DD-Mon-YYYY
      <qty> Mon-YY EXCHANGE PRODUCT
      ...
    """
    rows: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not _is_dd_mmm_yyyy(line):
            i += 1
            continue
        if i + 2 >= len(lines):
            break

        nxt = lines[i + 1]
        buy_qty = sell_qty = None
        delivery_product = None
        j = i + 2

        if re.match(r"^[\d,]+$", nxt):
            buy_qty = _num_any(nxt)
            delivery_product = lines[i + 2] if i + 2 < len(lines) else None
            j = i + 3
        else:
            m_sell = re.match(r"^(?P<qty>[\d,]+)\s+(?P<dp>.+)$", nxt)
            if m_sell:
                sell_qty = _num_any(m_sell.group("qty"))
                delivery_product = m_sell.group("dp")
                j = i + 2
            else:
                i += 1
                continue

        if not delivery_product or not re.match(
            r"^(?:\d{1,2}-[A-Za-z]{3}-\d{2,4}|[A-Za-z]{3}-\d{2,4})\s+", delivery_product
        ):
            i += 1
            continue

        # Optional Call/Put + Strike
        call_put = strike = None
        if j < len(lines) and str(lines[j]).strip().upper() in {"C", "P", "CALL", "PUT"}:
            call_put = str(lines[j]).strip().upper()
            if j + 1 < len(lines):
                strike = _num_any(lines[j + 1])
            j += 2

        if j >= len(lines):
            i += 1
            continue
        trade_price = _num_any(lines[j])
        if trade_price is None:
            i += 1
            continue
        j += 1

        # Scan forward for optional broker code, ccy, amount
        broker_code = ccy = amount = None
        k = j
        while k < len(lines) and k < j + 5:
            token = str(lines[k]).strip()
            if not token:
                k += 1
                continue
            if _is_dd_mmm_yyyy(token) or token.startswith("Diagnostic Key:") or re.match(r"^Total[:\s]", token) or token == "Total":
                break
            if _is_ccy_line(token):
                ccy = token
                k += 1
                if k < len(lines):
                    amt = _num_any(lines[k])
                    if amt is not None:
                        amount = amt
                        k += 1
                break
            # Non-numeric text → broker code
            if not _is_amount_line(token) and _num_any(token) is None:
                broker_code = token
            k += 1

        delivery_parts = _parse_delivery_product_text(delivery_product)
        desc = delivery_parts.get("contract_description")
        parsed_meta = parse_contract_product(desc)
        option_type = None
        if call_put in {"C", "CALL"}:
            option_type = "Call"
        elif call_put in {"P", "PUT"}:
            option_type = "Put"
        elif parsed_meta.get("option_type"):
            option_type = parsed_meta.get("option_type")
        if strike is None:
            strike = parsed_meta.get("strike")

        qty_abs = buy_qty if buy_qty is not None else sell_qty
        row = {
            "trade_date": line,
            "trade_date_iso": _normalize_any_date(line, stmt_date),
            "quantity": qty_abs,
            "buy": buy_qty,
            "sell": sell_qty,
            "side": "Buy" if buy_qty is not None else "Sell",
            "delivery_date": delivery_parts.get("delivery_date"),
            "contract_date": delivery_parts.get("contract_date"),
            "settlement_date": delivery_parts.get("contract_date"),
            "contract_month": delivery_parts.get("contract_month"),
            "contract_year": delivery_parts.get("contract_year"),
            "ref_month": delivery_parts.get("ref_month"),
            "contract_description": desc,
            "exchange": parsed_meta.get("exchange"),
            "product": parsed_meta.get("product"),
            "option_type": option_type,
            "option_type_raw": call_put,
            "strike": strike,
            "trade_price": trade_price,
            "price": trade_price,
            "currency": ccy,
            "amount": amount,
            "broker_code": broker_code,
            "statement_date": stmt_date,
            "account_number": account,
            "page": page_no,
            "source_section": source_section,
            "source_system": broker_code or "StoneX",
            "source_line": " ".join(str(x) for x in lines[i:k]),
        }
        rows.append(row)
        i = k
    return rows


def _scan_asx_ps_total(lines: list[str], start_idx: int) -> tuple[float | None, float | None, str | None, float | None, int]:
    """Scan from a 'Total' line to extract buy_qty, sell_qty, ccy, realized_pnl.

    Handles both single-line totals:
        "Total 163 163 P&S AUD 12,925.00"
    and stacked (PyMuPDF emits each token on its own line):
        Total / 163 / 163 / P&S / AUD / 12,925.00

    Also handles option totals:
        "Total: 10 Premium Paid AUD 0.00"

    Returns (buy_qty, sell_qty, ccy, amount, next_index).
    """
    # Join the next few lines to handle both single and stacked layouts
    chunk = " ".join(str(lines[j]).strip() for j in range(start_idx, min(start_idx + 8, len(lines))))
    # Pattern A: "Total [qty] [qty] P&S CCY amount"
    m = re.search(
        r"Total:?\s+([\d,]+)\s+([\d,]+)\s+(?:P&S)\s+([A-Z]{3})\s+([\d,]+(?:\.\d+)?)",
        chunk,
    )
    # Pattern B: "Total: [qty] Premium Paid CCY amount"
    if not m:
        m = re.search(
            r"Total:?\s+([\d,]+)\s+(?:Premium\s+Paid|P&S)\s+([A-Z]{3})\s+([\d,]+(?:\.\d+)?)",
            chunk,
        )
        if m:
            buy_qty = _num_any(m.group(1))
            ccy = m.group(2)
            amount = _num_any(m.group(3))
            # Advance index past the consumed tokens
            end_k = start_idx + 1
            while end_k < len(lines) and end_k < start_idx + 8:
                if _num_any(lines[end_k]) == amount and amount is not None:
                    end_k += 1
                    break
                end_k += 1
            return buy_qty, buy_qty, ccy, amount, end_k
    if m:
        buy_qty = _num_any(m.group(1))
        sell_qty = _num_any(m.group(2))
        ccy = m.group(3)
        amount = _num_any(m.group(4))
        # Advance past consumed tokens
        end_k = start_idx + 1
        target = m.group(4)
        for j in range(start_idx + 1, min(start_idx + 8, len(lines))):
            if str(lines[j]).strip().replace(",", "") == target.replace(",", ""):
                end_k = j + 1
                break
        return buy_qty, sell_qty, ccy, amount, end_k
    return None, None, None, None, start_idx + 1


def _parse_asx_ps_rows(
    lines: list[str],
    stmt_date: str | None,
    account: str | None,
    page_no: int,
) -> tuple[list[dict], list[dict]]:
    """Parse ASX PURCHASE & SALE section.

    Returns (trade_rows, closed_position_rows).
    - trade_rows: individual buy/sell records (no P&L, just the raw trades)
    - closed_position_rows: one row per product group with realized P&L, from Total lines
    """
    trades: list[dict] = []
    closed: list[dict] = []
    i = 0
    last_contract: dict = {}   # most recently parsed trade row (for product context on Total)

    while i < len(lines):
        line = lines[i]

        # Detect Total line: "Total" or "Total:" or "Total 163 ..."
        if re.match(r"^Total[:\s]", line) or line.strip() == "Total":
            buy_qty, sell_qty, ccy, amount, next_i = _scan_asx_ps_total(lines, i)
            if amount is not None and last_contract:
                amount_signed = amount if (buy_qty or 0) >= (sell_qty or 0) else -amount
                # Treat DR amounts as negative: look for DR marker in next few lines
                for peek in range(i, min(next_i + 2, len(lines))):
                    if "DR" in str(lines[peek]).upper():
                        amount_signed = -abs(amount)
                        break
                row = {
                    "trade_date": last_contract.get("trade_date"),
                    "trade_date_iso": last_contract.get("trade_date_iso"),
                    "statement_date": stmt_date,
                    "account_number": account,
                    "contract_description": last_contract.get("contract_description"),
                    "exchange": last_contract.get("exchange"),
                    "product": last_contract.get("product"),
                    "ref_month": last_contract.get("ref_month"),
                    "contract_month": last_contract.get("contract_month"),
                    "contract_year": last_contract.get("contract_year"),
                    "option_type": last_contract.get("option_type"),
                    "strike": last_contract.get("strike"),
                    "quantity": sell_qty or buy_qty,
                    "buy": buy_qty,
                    "sell": sell_qty,
                    "currency": ccy,
                    "realized_pnl": amount_signed,
                    "amount_signed": amount_signed,
                    "pnl_view": "Closed Position Detail",
                    "source_section": "Purchase & Sale",
                    "source_system": "StoneX",
                    "page": page_no,
                }
                closed.append(row)
            i = next_i
            continue

        if not _is_dd_mmm_yyyy(line):
            i += 1
            continue
        if i + 2 >= len(lines):
            break

        nxt = lines[i + 1]
        buy_qty_t = sell_qty_t = None
        delivery_product = None
        j = i + 2

        if re.match(r"^[\d,]+$", nxt):
            buy_qty_t = _num_any(nxt)
            delivery_product = lines[i + 2] if i + 2 < len(lines) else None
            j = i + 3
        else:
            m_sell = re.match(r"^(?P<qty>[\d,]+)\s+(?P<dp>.+)$", nxt)
            if m_sell:
                sell_qty_t = _num_any(m_sell.group("qty"))
                delivery_product = m_sell.group("dp")
                j = i + 2
            else:
                i += 1
                continue

        if not delivery_product or not re.match(
            r"^(?:\d{1,2}-[A-Za-z]{3}-\d{2,4}|[A-Za-z]{3}-\d{2,4})\s+", delivery_product
        ):
            i += 1
            continue

        call_put = strike_t = None
        if j < len(lines) and str(lines[j]).strip().upper() in {"C", "P", "CALL", "PUT"}:
            call_put = str(lines[j]).strip().upper()
            if j + 1 < len(lines):
                strike_t = _num_any(lines[j + 1])
            j += 2

        if j >= len(lines):
            i += 1
            continue
        trade_price = _num_any(lines[j])
        if trade_price is None:
            i += 1
            continue
        j += 1

        broker_code = ccy_t = amount_t = None
        k = j
        while k < len(lines) and k < j + 5:
            token = str(lines[k]).strip()
            if not token:
                k += 1
                continue
            if _is_dd_mmm_yyyy(token) or token.startswith("Diagnostic Key:") or re.match(r"^Total[:\s]", token) or token == "Total":
                break
            if _is_ccy_line(token):
                ccy_t = token
                k += 1
                if k < len(lines):
                    amt = _num_any(lines[k])
                    if amt is not None:
                        amount_t = amt
                        k += 1
                break
            if not _is_amount_line(token) and _num_any(token) is None:
                broker_code = token
            k += 1

        delivery_parts = _parse_delivery_product_text(delivery_product)
        desc = delivery_parts.get("contract_description")
        parsed_meta = parse_contract_product(desc)
        option_type = None
        if call_put in {"C", "CALL"}:
            option_type = "Call"
        elif call_put in {"P", "PUT"}:
            option_type = "Put"
        elif parsed_meta.get("option_type"):
            option_type = parsed_meta.get("option_type")
        if strike_t is None:
            strike_t = parsed_meta.get("strike")

        qty_abs = buy_qty_t if buy_qty_t is not None else sell_qty_t
        row = {
            "trade_date": line,
            "trade_date_iso": _normalize_any_date(line, stmt_date),
            "quantity": qty_abs,
            "buy": buy_qty_t,
            "sell": sell_qty_t,
            "side": "Buy" if buy_qty_t is not None else "Sell",
            "delivery_date": delivery_parts.get("delivery_date"),
            "contract_date": delivery_parts.get("contract_date"),
            "settlement_date": delivery_parts.get("contract_date"),
            "contract_month": delivery_parts.get("contract_month"),
            "contract_year": delivery_parts.get("contract_year"),
            "ref_month": delivery_parts.get("ref_month"),
            "contract_description": desc,
            "exchange": parsed_meta.get("exchange"),
            "product": parsed_meta.get("product"),
            "option_type": option_type,
            "option_type_raw": call_put,
            "strike": strike_t,
            "trade_price": trade_price,
            "price": trade_price,
            "currency": ccy_t,
            "amount": amount_t,
            "broker_code": broker_code,
            "statement_date": stmt_date,
            "account_number": account,
            "page": page_no,
            "source_section": "Purchase & Sale",
            "source_system": broker_code or "StoneX",
            "source_line": " ".join(str(x) for x in lines[i:k]),
        }
        trades.append(row)
        last_contract = row
        i = k
    return trades, closed


def extract_asx_statement(pdf_bytes: bytes, include_open_positions: bool = True) -> Dict[str, pd.DataFrame]:
    """Parse StoneX Financial Pty Ltd ASX/SFE statements.

    Same stacked-text table layout as IFL statements. Sections are delimited by
    'Diagnostic Key:' lines.  Parses:
      - FUTURES / OPTIONS OPEN POSITIONS  → Open Positions
      - FUTURES / OPTIONS CONFIRMATIONS   → Executed Trades
      - PURCHASE & SALE                   → Purchase & Sale (trade detail) + Closed Positions (P&L totals)
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    open_positions: list[dict] = []
    executed_trades: list[dict] = []
    purchase_sale: list[dict] = []
    closed_positions: list[dict] = []

    for page_no, page in enumerate(doc, start=1):
        text = page.get_text("text")
        stmt_date, account = _ifl_page_header_info(text)
        raw_lines = [" ".join(x.strip().split()) for x in text.splitlines() if x.strip()]

        lines_by_section: dict[str, list[str]] = {}
        section: str | None = None

        for line in raw_lines:
            if line == "FUTURES / OPTIONS OPEN POSITIONS":
                section = "open"
                lines_by_section.setdefault(section, [])
                continue
            if line == "FUTURES / OPTIONS CONFIRMATIONS":
                section = "confirmations"
                lines_by_section.setdefault(section, [])
                continue
            if line == "PURCHASE & SALE":
                section = "ps"
                lines_by_section.setdefault(section, [])
                continue
            if line.startswith("Diagnostic Key:") or line in {
                "RECAP OF CONFIRMATION ACTIVITY",
                "FINANCIAL SUMMARY",
                "PURCHASE & SALE_Total",
                "Disclaimers",
            }:
                section = None
                continue
            if section:
                lines_by_section.setdefault(section, []).append(line)

        if include_open_positions and "open" in lines_by_section:
            open_positions.extend(
                _parse_ifl_futures_options_rows(lines_by_section["open"], stmt_date, account, page_no)
            )
        if "confirmations" in lines_by_section:
            executed_trades.extend(
                _parse_asx_trade_rows(
                    lines_by_section["confirmations"], stmt_date, account, page_no,
                    source_section="Executed Trades",
                )
            )
        if "ps" in lines_by_section:
            ps_trades, ps_closed = _parse_asx_ps_rows(
                lines_by_section["ps"], stmt_date, account, page_no
            )
            purchase_sale.extend(ps_trades)
            closed_positions.extend(ps_closed)

    tables: Dict[str, pd.DataFrame] = {
        "Executed Trades": pd.DataFrame(executed_trades),
        "Purchase & Sale": pd.DataFrame(purchase_sale),
        "Receives Delivers": pd.DataFrame(),
        "Journal Entries": pd.DataFrame(),
        "Realized Gain and Loss": pd.DataFrame(),
        "Closed Positions": pd.DataFrame(closed_positions),
        "Realized PNL Summary": pd.DataFrame(),
        "Open Positions": pd.DataFrame(open_positions),
        "Notes": pd.DataFrame(),
        "Exceptions": pd.DataFrame(),
    }
    tables = enrich_open_positions_metadata(tables)
    tables["Summary"] = build_summary(tables)
    return tables


def looks_like_ifl_statement(pdf_bytes: bytes) -> bool:
    """Detect StoneX Financial Ltd IFL statements with line-stacked open-position tables."""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = "\n".join(page.get_text("text") for page in list(doc)[:3])
        return (
            "StoneX Financial Ltd" in text
            and (
                "FUTURES / OPTIONS OPEN POSITIONS" in text
                or "FX OPEN POSITIONS" in text
                or "FX Spot/Forward Open Positions" in text
                or "FUTURES / OPTIONS CONFIRMATIONS" in text
            )
        )
    except Exception:
        return False


def _ifl_page_header_info(text: str) -> tuple[str | None, str | None]:
    stmt = None
    acct = None

    # Some IFL PDFs extract header labels and values on separate lines:
    #   Statement Date:
    #   14-May-2026
    m = re.search(r"Statement Date:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if not m:
        m = re.search(r"Statement Date:\s*\n\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    if m:
        stmt = _normalize_any_date(m.group(1))
    if stmt is None:
        # FX Spot/Forward statements print the date as the line below "Daily Statement"
        # rather than with a "Statement Date:" label.
        m = re.search(r"Daily Statement\s*\n\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
        if m:
            stmt = _normalize_any_date(m.group(1))

    m = re.search(r"Account Number:\s*(\S+)", text, re.I)
    if not m:
        m = re.search(r"Account Number:\s*\n\s*(\S+)", text, re.I)
    if m:
        acct = m.group(1)
    return stmt, acct


def _is_rate_line(s: str) -> bool:
    return bool(re.match(r"^\d+(?:\.\d+)?$", str(s or "").strip()))


def _parse_amount_ccy(text: str | None) -> tuple[float | None, str | None]:
    s = " ".join(str(text or "").strip().split())
    m = re.match(r"^(?P<amount>\(?[-+]?\d[\d,]*(?:\.\d+)?\)?)\s+(?P<ccy>[A-Z]{3})$", s)
    if not m:
        return None, None
    return _num_any(m.group("amount")), m.group("ccy")


def _parse_delivery_product_text(text: str | None) -> dict:
    """Split Delivery / Product into date/month/product components.

    Handles examples:
      - Nov-26 IFUS CANOLA (20)
      - 13-Aug-26 LME PRIMARY ALUMINIUM USD
    """
    raw = " ".join(str(text or "").strip().split())
    out = {
        "delivery_date": None,
        "contract_date": None,
        "contract_month": None,
        "contract_year": None,
        "ref_month": None,
        "contract_description": raw or None,
    }
    if not raw:
        return out

    m_date = re.match(r"^(?P<date>\d{1,2}-[A-Za-z]{3}-\d{2,4})\s+(?P<product>.+)$", raw)
    if m_date:
        delivery = m_date.group("date")
        iso = _normalize_any_date(delivery)
        out["delivery_date"] = delivery
        out["contract_date"] = delivery
        out["ref_month"] = _ref_month_from_date(delivery)
        if out["ref_month"] and "-" in out["ref_month"]:
            out["contract_month"], out["contract_year"] = out["ref_month"].split("-", 1)
        out["contract_description"] = m_date.group("product")
        return out

    m_ref = re.match(r"^(?P<mon>[A-Za-z]{3})-(?P<yy>\d{2,4})\s+(?P<product>.+)$", raw)
    if m_ref:
        mon = m_ref.group("mon").upper()
        yy = m_ref.group("yy")[-2:]
        out["contract_month"] = mon
        out["contract_year"] = yy
        out["ref_month"] = f"{mon}-{yy}"
        out["contract_description"] = f"{mon} {yy} {m_ref.group('product')}"
        return out

    return out


def _fx_product_from_code(code: str | None) -> str | None:
    c = str(code or "").upper().strip()
    m = re.match(r"^FX([A-Z]{3})([A-Z]{3})$", c)
    if not m:
        return c or None
    return f"FX {m.group(1)}/{m.group(2)}"


def _parse_ifl_fx_rows(lines: list[str], stmt_date: str | None, account: str | None, page_no: int) -> list[dict]:
    rows: list[dict] = []
    current_code = None
    current_delivery = None
    i = 0
    while i < len(lines):
        line = lines[i]
        m_head = re.match(r"^(?P<code>FX[A-Z]{6})\s+(?P<delivery>\d{1,2}-[A-Za-z]{3}-\d{2,4})$", line)
        if m_head:
            current_code = m_head.group("code")
            current_delivery = m_head.group("delivery")
            i += 1
            continue

        if current_code and _is_dd_mmm_yyyy(line):
            try:
                primary_amount, primary_ccy = _parse_amount_ccy(lines[i + 1])
                rate = _num_any(lines[i + 2])
                secondary_amount, secondary_ccy = _parse_amount_ccy(lines[i + 3])
                mtm = _num_any(lines[i + 4])
                if primary_amount is None or not primary_ccy or rate is None or secondary_amount is None or not secondary_ccy or mtm is None:
                    i += 1
                    continue
                delivery_iso = _normalize_any_date(current_delivery)
                ref_month = _ref_month_from_date(current_delivery)
                long_qty = primary_amount if primary_amount > 0 else None
                short_qty = abs(primary_amount) if primary_amount < 0 else None
                product = _fx_product_from_code(current_code)
                raw_line = " ".join(lines[i:i+5])
                row = {
                    "trade_date": line,
                    "trade_date_iso": _normalize_any_date(line, stmt_date),
                    "quantity": abs(primary_amount),
                    "long": long_qty,
                    "short": short_qty,
                    "side": "Long" if primary_amount > 0 else ("Short" if primary_amount < 0 else None),
                    "delivery_date": current_delivery,
                    "value_date": current_delivery,
                    "contract_date": current_delivery,
                    "settlement_date": current_delivery,
                    "ref_month": ref_month,
                    "contract_month": ref_month.split("-", 1)[0] if ref_month and "-" in ref_month else None,
                    "contract_year": ref_month.split("-", 1)[1] if ref_month and "-" in ref_month else None,
                    "contract_description": f"{current_code} {current_delivery}",
                    "product": product,
                    "product_name": product,
                    "exchange": "FX",
                    "currency": secondary_ccy,
                    "primary_amount": primary_amount,
                    "primary_currency": primary_ccy,
                    "ccy_1": primary_ccy,
                    "ccy_1_amount": primary_amount,
                    "secondary_amount": secondary_amount,
                    "secondary_currency": secondary_ccy,
                    "ccy_2": secondary_ccy,
                    "ccy_2_amount": secondary_amount,
                    "rate": rate,
                    "trade_price": rate,
                    "market_value": mtm,
                    "market_value_signed": mtm,
                    "statement_date": stmt_date,
                    "account_number": account,
                    "page": page_no,
                    "source_section": "FX Open Positions",
                    "source_system": "StoneX",
                    "source_line": raw_line,
                }
                rows.append(row)
                i += 5
                continue
            except Exception:
                pass
        i += 1
    return rows



def _parse_fx_spot_forward_rows(lines: list[str], stmt_date: str | None, account: str | None, page_no: int) -> list[dict]:
    """Parse StoneX Financial Ltd FX Spot/Forward Open Positions layout.

    This layout appears on newer FX statements under the title
    "FX Spot/Forward Open Positions" with columns:
        Trade Date, Trade Id, Global Id, Type, Curr Pair, CCY1,
        CCY1 Buy/(Sell), Trade Price, CCY2, CCY2 Buy/(Sell),
        Fixing Date, Value Date, Market Price, P&L CCY, Native P&L, Market Value.

    PyMuPDF emits each cell as a separate line. In the observed PDFs the Fixing
    Date column may be blank, so only one date appears between CCY2 Buy/(Sell)
    and Market Price; that date is treated as Value Date / expiryDate.
    """
    rows: list[dict] = []
    i = 0
    while i < len(lines):
        line = str(lines[i]).strip()
        if not _is_dd_mmm_yyyy(line):
            i += 1
            continue
        try:
            if i + 13 >= len(lines):
                break
            trade_id = str(lines[i + 1]).strip()
            global_id = str(lines[i + 2]).strip()
            trade_type = str(lines[i + 3]).strip()
            curr_pair = str(lines[i + 4]).strip().upper()
            ccy1 = str(lines[i + 5]).strip().upper()
            ccy1_amount = _num_any(lines[i + 6])
            trade_price = _num_any(lines[i + 7])
            ccy2 = str(lines[i + 8]).strip().upper()
            ccy2_amount = _num_any(lines[i + 9])
            j = i + 10

            # One or two date columns may be present. If only one date is present,
            # it is the Value Date in this PDF layout.
            fixing_date = None
            value_date = None
            if j < len(lines) and _is_dd_mmm_yyyy(str(lines[j]).strip()):
                first_date = str(lines[j]).strip()
                j += 1
                if j < len(lines) and _is_dd_mmm_yyyy(str(lines[j]).strip()):
                    fixing_date = first_date
                    value_date = str(lines[j]).strip()
                    j += 1
                else:
                    value_date = first_date

            if j + 3 >= len(lines):
                i += 1
                continue
            market_price = _num_any(lines[j])
            pnl_ccy = str(lines[j + 1]).strip().upper()
            native_pnl = _num_any(lines[j + 2])
            market_value = _num_any(lines[j + 3])

            if not re.match(r"^\d+$", trade_id):
                i += 1
                continue
            if not re.match(r"^FX\b", trade_type.upper()):
                i += 1
                continue
            if not re.match(r"^[A-Z]{3}/[A-Z]{3}$", curr_pair):
                i += 1
                continue
            if not (_is_ccy_line(ccy1) and _is_ccy_line(ccy2) and _is_ccy_line(pnl_ccy)):
                i += 1
                continue
            if ccy1_amount is None or ccy2_amount is None or trade_price is None or market_price is None or market_value is None:
                i += 1
                continue

            # Use the actual amount currencies as the product identity. The table's
            # Curr Pair can be printed as USD/CAD while the economic exposure is
            # CAD vs USD and the subtotal is CAD/USD.
            product = f"FX {ccy1}/{ccy2}"
            value_date_norm = _normalize_any_date(value_date)
            ref_month = _ref_month_from_date(value_date) if value_date else None
            qty = abs(ccy1_amount) if ccy1_amount is not None else None
            raw_line = " ".join(str(x).strip() for x in lines[i:j + 4])
            row = {
                "trade_date": line,
                "trade_date_iso": _normalize_any_date(line, stmt_date),
                "trade_id": trade_id,
                "global_id": global_id,
                "type": trade_type,
                "curr_pair": curr_pair,
                "quantity": qty,
                "long": ccy1_amount if ccy1_amount is not None and ccy1_amount > 0 else None,
                "short": abs(ccy1_amount) if ccy1_amount is not None and ccy1_amount < 0 else None,
                "side": "Long" if ccy1_amount and ccy1_amount > 0 else ("Short" if ccy1_amount and ccy1_amount < 0 else None),
                "value_date": value_date,
                "fixing_date": fixing_date,
                "delivery_date": value_date,
                "contract_date": value_date,
                "settlement_date": value_date,
                "expiry_date": value_date_norm,
                "ref_month": ref_month,
                "contract_month": ref_month.split("-", 1)[0] if ref_month and "-" in ref_month else None,
                "contract_year": ref_month.split("-", 1)[1] if ref_month and "-" in ref_month else None,
                "contract_description": f"{product} {value_date or ''}".strip(),
                "product": product,
                "product_name": product,
                "exchange": "FX",
                "currency": ccy2,
                "primary_amount": ccy1_amount,
                "primary_currency": ccy1,
                "ccy_1": ccy1,
                "ccy_1_amount": ccy1_amount,
                "secondary_amount": ccy2_amount,
                "secondary_currency": ccy2,
                "ccy_2": ccy2,
                "ccy_2_amount": ccy2_amount,
                "rate": trade_price,
                "trade_price": trade_price,
                "market_price": market_price,
                "pnl_ccy": pnl_ccy,
                "native_pnl": native_pnl,
                "market_value": market_value,
                "market_value_signed": market_value,
                "statement_date": stmt_date,
                "account_number": account,
                "page": page_no,
                "source_section": "FX Spot/Forward Open Positions",
                "source_system": "StoneX FX Spot/Forward",
                "source_line": raw_line,
            }
            rows.append(row)
            i = j + 4
            continue
        except Exception:
            pass
        i += 1
    return rows

def _is_ifl_source_system_line(s: str | None) -> bool:
    """Return True for execution/source-system text between price and ccy in IFL tables.

    Examples from StoneX Financial Ltd IFL PDFs:
      InterOffice
      InterOffice GiveUpClearer
      InterOffice Transfer
      InterOffice FutAsgn
      IFC InterOffice
    """
    u = " ".join(str(s or "").strip().split()).upper()
    if not u:
        return False
    if _is_ccy_line(u) or _is_amount_line(u) or _is_dd_mmm_yyyy(u):
        return False
    source_words = ("INTEROFFICE", "GIVEUPCLEARER", "TRANSFER", "FUTASGN", "IFC", "CLEARER")
    return any(w in u for w in source_words)


def _next_ccy_amount_after_price(lines: list[str], start_idx: int) -> tuple[int | None, str | None, float | None, str | None]:
    """Find Ccy + Amount after Trade Price, skipping source-system columns.

    Returns (ccy_index, ccy, amount, source_system_text).  The IFL open-position
    table visually has Trade Price, source system text, Ccy, Amount, but PyMuPDF
    emits the source-system text as its own line. Older parser versions expected
    Ccy immediately after Trade Price, so LME11630-style statements produced zero
    open positions.
    """
    source_tokens: list[str] = []
    k = start_idx
    while k < len(lines):
        token = str(lines[k]).strip()
        if _is_ccy_line(token):
            if k + 1 < len(lines):
                amount = _num_any(lines[k + 1])
                if amount is not None:
                    return k, token, amount, " ".join(source_tokens).strip() or None
            return None, None, None, None
        if _is_dd_mmm_yyyy(token) or token.startswith("Diagnostic Key:") or token in {"Total", "Last Traded Date:"}:
            return None, None, None, None
        if _is_ifl_source_system_line(token):
            source_tokens.append(token)
            k += 1
            continue
        # Occasionally the source system is split into two adjacent text lines;
        # skip short non-numeric text between price and currency, but do not scan
        # too far because summary rows follow the actual amount.
        if len(source_tokens) < 3 and not _is_amount_line(token) and _num_any(token) is None:
            source_tokens.append(token)
            k += 1
            continue
        return None, None, None, None
    return None, None, None, None


def _parse_ifl_futures_options_rows(lines: list[str], stmt_date: str | None, account: str | None, page_no: int) -> list[dict]:
    rows: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not _is_dd_mmm_yyyy(line):
            i += 1
            continue
        if i + 4 >= len(lines):
            break

        nxt = lines[i + 1]
        long_qty = short_qty = None
        delivery_product = None
        j = i + 2

        # Long rows generally have quantity on its own line, then Delivery/Product.
        if re.match(r"^[\d,]+$", nxt) and i + 5 < len(lines):
            long_qty = _num_any(nxt)
            delivery_product = lines[i + 2]
            j = i + 3
        else:
            # Short rows generally have quantity and Delivery/Product on the same line.
            m_short = re.match(r"^(?P<qty>[\d,]+)\s+(?P<delivery_product>.+)$", nxt)
            if m_short:
                short_qty = _num_any(m_short.group("qty"))
                delivery_product = m_short.group("delivery_product")
                j = i + 2
            else:
                i += 1
                continue

        if not delivery_product or not re.match(r"^(?:\d{1,2}-[A-Za-z]{3}-\d{2,4}|[A-Za-z]{3}-\d{2,4})\s+", delivery_product):
            i += 1
            continue

        call_put = None
        strike = None
        if j < len(lines) and str(lines[j]).strip().upper() in {"C", "P", "CALL", "PUT"}:
            call_put = str(lines[j]).strip().upper()
            if j + 1 < len(lines):
                strike = _num_any(lines[j + 1])
            j += 2

        if j >= len(lines):
            i += 1
            continue
        trade_price = _num_any(lines[j])
        if trade_price is None:
            i += 1
            continue

        ccy_idx, ccy, amount, source_system = _next_ccy_amount_after_price(lines, j + 1)
        if ccy_idx is None or ccy is None or amount is None:
            i += 1
            continue

        expiry_date = None
        # Option sections print Expiration Date on a following summary line, not in
        # the main trade row. Attach it to each option row so Product + Contract
        # Month/Year can group by expiry when available.
        if call_put is not None or re.search(r"\b(?:CALL|PUT)\b", str(delivery_product), re.I):
            for k in range(ccy_idx + 2, min(len(lines), ccy_idx + 22)):
                probe = str(lines[k]).strip()
                if k > ccy_idx + 2 and _is_dd_mmm_yyyy(probe):
                    break
                if probe.startswith("Diagnostic Key:") or probe == "FUTURES / OPTIONS OPEN POSITIONS":
                    break
                m_exp = re.search(r"Expiration Date:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", probe, re.I)
                if m_exp:
                    expiry_date = _normalize_any_date(m_exp.group(1))
                    break
                if re.match(r"^Expiration Date:?$", probe, re.I) and k + 1 < len(lines):
                    next_probe = str(lines[k + 1]).strip()
                    if _is_dd_mmm_yyyy(next_probe):
                        expiry_date = _normalize_any_date(next_probe)
                        break

        delivery_parts = _parse_delivery_product_text(delivery_product)
        desc = delivery_parts.get("contract_description")
        parsed_meta = parse_contract_product(desc)
        option_type = None
        if call_put in {"C", "CALL"}:
            option_type = "Call"
        elif call_put in {"P", "PUT"}:
            option_type = "Put"
        elif parsed_meta.get("option_type"):
            option_type = parsed_meta.get("option_type")
        if strike is None:
            strike = parsed_meta.get("strike")

        qty_abs = long_qty if long_qty is not None else short_qty
        raw_line = " ".join(lines[i:ccy_idx+2])
        row = {
            "trade_date": line,
            "trade_date_iso": _normalize_any_date(line, stmt_date),
            "quantity": qty_abs,
            "long": long_qty,
            "short": short_qty,
            "side": "Long" if long_qty is not None else "Short",
            "delivery_date": delivery_parts.get("delivery_date"),
            "contract_date": delivery_parts.get("contract_date"),
            "settlement_date": delivery_parts.get("contract_date"),
            "expiry_date": expiry_date,
            "expiration_date": expiry_date,
            "contract_month": delivery_parts.get("contract_month"),
            "contract_year": delivery_parts.get("contract_year"),
            "ref_month": delivery_parts.get("ref_month"),
            "contract_description": desc,
            "option_type": option_type,
            "option_type_raw": call_put,
            "strike": strike,
            "trade_price": trade_price,
            "price": trade_price,
            "currency": ccy,
            "market_value": amount,
            "market_value_signed": amount,
            "statement_date": stmt_date,
            "account_number": account,
            "page": page_no,
            "source_section": "Futures / Options Open Positions",
            "source_system": source_system or "StoneX",
            "source_line": raw_line,
        }
        rows.append(row)
        i = ccy_idx + 2
    return rows

def extract_ifl_statement(pdf_bytes: bytes, include_open_positions: bool = True) -> Dict[str, pd.DataFrame]:
    """Parse StoneX Financial Ltd IFL statements with stacked text extraction.

    This covers futures/options open positions and FX open positions, where PyMuPDF
    often emits each table cell on its own line rather than a full visual row.
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    open_positions: list[dict] = []
    notes: list[dict] = []
    exceptions: list[dict] = []

    for page_no, page in enumerate(doc, start=1):
        text = page.get_text("text")
        stmt_date, account = _ifl_page_header_info(text)
        raw_lines = [" ".join(x.strip().split()) for x in text.splitlines() if x.strip()]
        lines_by_section: dict[str, list[str]] = {}
        section = None
        for line in raw_lines:
            if line == "FUTURES / OPTIONS OPEN POSITIONS":
                section = "fut_opt"
                lines_by_section.setdefault(section, [])
                continue
            if line == "FX OPEN POSITIONS":
                section = "fx"
                lines_by_section.setdefault(section, [])
                continue
            if line == "FX Spot/Forward Open Positions":
                section = "fx_spot_forward"
                lines_by_section.setdefault(section, [])
                continue
            if line.startswith("Diagnostic Key:"):
                section = None
                continue
            if line in {"FUTURES / OPTIONS CONFIRMATIONS", "RECAP OF CONFIRMATION ACTIVITY", "FINANCIAL SUMMARY", "Disclaimers"}:
                section = None
                continue
            if section:
                lines_by_section.setdefault(section, []).append(line)

        if include_open_positions:
            if "fut_opt" in lines_by_section:
                open_positions.extend(_parse_ifl_futures_options_rows(lines_by_section["fut_opt"], stmt_date, account, page_no))
            if "fx" in lines_by_section:
                open_positions.extend(_parse_ifl_fx_rows(lines_by_section["fx"], stmt_date, account, page_no))
            if "fx_spot_forward" in lines_by_section:
                open_positions.extend(_parse_fx_spot_forward_rows(lines_by_section["fx_spot_forward"], stmt_date, account, page_no))

    tables: Dict[str, pd.DataFrame] = {
        "Executed Trades": pd.DataFrame(),
        "Purchase & Sale": pd.DataFrame(),
        "Receives Delivers": pd.DataFrame(),
        "Journal Entries": pd.DataFrame(),
        "Realized Gain and Loss": pd.DataFrame(),
        "Closed Positions": pd.DataFrame(),
        "Realized PNL Summary": pd.DataFrame(),
        "Open Positions": pd.DataFrame(open_positions),
        "Notes": pd.DataFrame(notes),
        "Exceptions": pd.DataFrame(exceptions),
    }
    tables = enrich_open_positions_metadata(tables)
    tables["Summary"] = build_summary(tables)
    return tables

def extract(pdf_bytes: bytes, include_open_positions: bool = True) -> Dict[str, pd.DataFrame]:
    if looks_like_asx_statement(pdf_bytes):
        return extract_asx_statement(pdf_bytes, include_open_positions=include_open_positions)
    if looks_like_ifl_statement(pdf_bytes):
        return extract_ifl_statement(pdf_bytes, include_open_positions=include_open_positions)
    if looks_like_murex_statement(pdf_bytes):
        return extract_murex_statement(pdf_bytes, include_open_positions=include_open_positions)
    if looks_like_monthly_statement(pdf_bytes):
        return extract_monthly(pdf_bytes, include_open_positions=include_open_positions)

    executed: List[dict] = []
    purchase_sale: List[dict] = []
    closed_positions: List[dict] = []
    receives_delivers: List[dict] = []
    journals: List[dict] = []
    open_positions: List[dict] = []
    notes: List[dict] = []
    exceptions: List[dict] = []

    pending_lme_avg: dict | None = None
    last_purchase_sale_contract: dict | None = None
    pending_ps_ltd_summary: dict | None = None  # LTD- line preceding a standalone GROSS PROFIT OR LOSS

    for page_no, text in pdf_text(pdf_bytes):
        stmt_date = _statement_date(text)
        account = _account_number(text)
        broker = _broker_code(text)
        section: str | None = None
        saw_stmt = bool(stmt_date)
        # Do not enter the OPEN section just because the page contains the OPEN header;
        # a page can contain the tail of P&S before the OPEN header. We switch to
        # open_positions only when the header/title line itself is reached below.

        normalized_lines = [" ".join(raw.strip().split()) for raw in text.splitlines() if " ".join(raw.strip().split())]
        if include_open_positions and ("LME AVERAGE OPEN POSITIONS" in text or "FUTURES / OPTIONS OPEN POSITIONS" in text):
            open_positions.extend(_parse_lme_open_positions_from_lines(normalized_lines, stmt_date, account, page_no, broker))

        raw_lines = text.splitlines()
        i = 0
        while i < len(raw_lines):
            raw = raw_lines[i]
            line = " ".join(raw.strip().split())
            if not line:
                i += 1
                continue

            if "THE FOLLOWING TRADES HAVE BEEN MADE" in line:
                section = "executed"
                i += 1
                continue
            if "P U R C H A S E" in line and "S A L E" in line:
                section = "purchase_sale"
                i += 1
                continue
            if "CONTRACT DESCRIPTION-P&S" in line:
                section = "purchase_sale"
                i += 1
                continue
            if "THE FOLLOWING ITEMS HAVE BEEN RECEIVED/DELIVERED" in line:
                section = "receives_delivers"
                i += 1
                continue
            if "THE FOLLOWING JOURNAL ENTRIES" in line:
                section = "journal"
                i += 1
                continue
            if "FUTURES / OPTIONS OPEN POSITIONS" in line:
                section = "lme_fut_opt_open" if include_open_positions else None
                i += 1
                continue
            if "LME AVERAGE OPEN POSITIONS" in line:
                section = "lme_average_open" if include_open_positions else None
                i += 1
                continue
            if "O P E N P O S I T I O N S" in line or "CONTRACT DESCRIPTION-OPEN" in line:
                section = "open_positions" if include_open_positions else None
                i += 1
                continue
            if line.startswith("*USD") or line.startswith("**TOTAL") or line.startswith("MARGIN CALL") or line.startswith("SUMMARY"):
                section = None
                i += 1
                continue
            if line.startswith("-------") or line.startswith("TRADE CARD") or line.startswith("TRADE AT"):
                i += 1
                continue

            if section in {"executed", "purchase_sale"}:
                match = TRADE_LINE_RE.match(line)
                parsed_trade = match.groupdict() if match else _parse_daily_trade_or_ps_line(line)
                if parsed_trade:
                    row = _base_row(parsed_trade, stmt_date, account, page_no, line, broker)
                    row["source_section"] = "Executed Trades" if section == "executed" else "Purchase & Sale"
                    row["side"] = _side_from_section(section, row.get("quantity"))
                    # Classic daily statements print the actual card/trade id on the next line.
                    if i + 1 < len(raw_lines):
                        next_line = " ".join(raw_lines[i + 1].strip().split())
                        if _is_card_continuation_line(next_line):
                            row["card"] = next_line
                            row["trade_id"] = next_line
                            row["source_line"] = f"{line} | {next_line}"
                            i += 1
                    if section == "purchase_sale":
                        parsed_contract = parse_contract_product(row.get("contract_description"))
                        for meta_col, meta_val in parsed_contract.items():
                            row.setdefault(meta_col, meta_val)
                        last_purchase_sale_contract = row.copy()
                        purchase_sale.append(row)
                    else:
                        executed.append(row)
                elif section == "purchase_sale" and "GROSS PROFIT OR LOSS" in line:
                    # Format A: all on one line  e.g. "U1 20* 20* LTD- 6/30/26 GROSS PROFIT OR LOSS US 30,237.50DR"
                    # Format B: standalone line  e.g. "GROSS PROFIT OR LOSS US 30,237.50DR"
                    #   (quantities/close-date were on a preceding LTD- line stored in pending_ps_ltd_summary)
                    closed = _parse_closed_position_gross_line(line, last_purchase_sale_contract, pending_ps_ltd_summary)
                    pending_ps_ltd_summary = None  # consumed regardless
                    if closed:
                        row = _base_row(closed, stmt_date, account, page_no, line, broker)
                        row["realized_pnl"] = closed.get("realized_pnl")
                        row["amount_signed"] = closed.get("amount_signed")
                        row["pnl_view"] = "Closed Position Detail"
                        parsed_contract = parse_contract_product(row.get("contract_description"))
                        for meta_col, meta_val in parsed_contract.items():
                            if row.get(meta_col) in (None, "", "Other"):
                                row[meta_col] = meta_val
                        closed_positions.append(row)
                    notes.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "text": line})
                elif section == "purchase_sale" and PS_LTD_SUMMARY_RE.match(line) and "GROSS PROFIT OR LOSS" not in line:
                    # Format B LTD- summary line (no GROSS PROFIT on this line):
                    # store quantities + close date for the next GROSS PROFIT OR LOSS line.
                    ltd_m = PS_LTD_SUMMARY_RE.match(line)
                    pending_ps_ltd_summary = ltd_m.groupdict() if ltd_m else None
                    notes.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "text": line})
                elif any(keyword in line for keyword in FEE_OR_AVG_KEYWORDS):
                    notes.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "text": line})
                elif _is_card_continuation_line(line):
                    pass
                elif re.match(r"^\d{1,2}/\d{2}/\d\s+", line):
                    # Confirmation/P&S trade rows in daily statements are often split across
                    # multiple fixed-width lines. They are not Open Positions, so do not
                    # flood Exceptions with harmless confirmation rows.
                    notes.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "text": line})

            elif section == "receives_delivers":
                match = RD_LINE_RE.match(line)
                if match:
                    row = _base_row(match.groupdict(), stmt_date, account, page_no, line, broker)
                    receives_delivers.append(row)
                elif re.match(r"^\d{1,2}/\d{2}/\d\s+", line):
                    exceptions.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "reason": "Unmatched receive/deliver line", "source_line": line})

            elif section == "journal":
                match = JOURNAL_RE.match(line)
                if match:
                    row = _base_row(match.groupdict(), stmt_date, account, page_no, line, broker)
                    journals.append(row)
                elif re.match(r"^\d{1,2}/\d{2}/\d\s+", line):
                    exceptions.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "reason": "Unmatched journal line", "source_line": line})

            elif section == "lme_fut_opt_open" and include_open_positions:
                parsed = _parse_lme_fut_opt_row(line)
                if parsed:
                    row = _base_row(parsed, stmt_date, account, page_no, line, broker)
                    row["quantity"] = _num_any(parsed.get("quantity"))
                    row["market_value"] = _num_any(parsed.get("market_value"))
                    row["market_value_signed"] = _num_any(parsed.get("market_value"))
                    row["trade_price"] = _num_any(parsed.get("trade_price"))
                    open_positions.append(row)
                elif line.startswith("Total") or line.startswith("Last Traded Date") or line.startswith("Trade Date"):
                    pass

            elif section == "lme_average_open" and include_open_positions:
                prod_match = LME_PRODUCT_RE.match(line)
                if prod_match and pending_lme_avg is not None:
                    pending_lme_avg["delivery_date"] = prod_match.group("delivery_date")
                    pending_lme_avg["contract_date"] = prod_match.group("delivery_date")
                    pending_lme_avg["contract_description"] = prod_match.group("product")
                    row = _base_row(pending_lme_avg, stmt_date, account, page_no, pending_lme_avg.get("source_line", line), broker)
                    row["quantity"] = _num_any(pending_lme_avg.get("quantity"))
                    row["market_value"] = _num_any(pending_lme_avg.get("market_value"))
                    row["market_value_signed"] = _num_any(pending_lme_avg.get("market_value"))
                    row["trade_price"] = pending_lme_avg.get("price")
                    open_positions.append(row)
                    pending_lme_avg = None
                else:
                    parsed = _parse_lme_average_row(line)
                    if parsed:
                        parsed["source_line"] = line
                        pending_lme_avg = parsed
                    elif line.startswith("Total") or line.startswith("Priced Lots") or line.startswith("Unpriced Lots") or "P&L" in line or line.startswith("Trade Date"):
                        pass

            elif section == "open_positions" and include_open_positions:
                parsed = _parse_daily_open_position_line(line, raw)
                if parsed:
                    row = _base_row(parsed, stmt_date, account, page_no, line, broker)
                    row["market_value_signed"] = _signed(row.get("market_value"), row.get("drcr"))
                    # Classic daily statements print the actual card/trade id on the next line.
                    # Attach it to the position row and skip that continuation line.
                    if i + 1 < len(raw_lines):
                        next_line = " ".join(raw_lines[i + 1].strip().split())
                        if _is_card_continuation_line(next_line):
                            row["card"] = next_line
                            row["trade_id"] = next_line
                            row["source_line"] = f"{line} | {next_line}"
                            i += 1
                    open_positions.append(row)
                else:
                    match = OPEN_POSITION_LINE_RE.match(line)
                    if match:
                        row = _base_row(match.groupdict(), stmt_date, account, page_no, line, broker)
                        row["market_value_signed"] = _signed(row.get("market_value"), row.get("drcr"))
                        if i + 1 < len(raw_lines):
                            next_line = " ".join(raw_lines[i + 1].strip().split())
                            if _is_card_continuation_line(next_line):
                                row["card"] = next_line
                                row["trade_id"] = next_line
                                row["source_line"] = f"{line} | {next_line}"
                                i += 1
                        open_positions.append(row)
                    elif _is_open_position_non_position_line(line):
                        # Header/footer/subtotal lines and USTB/cash-collateral rows are not
                        # futures/options open positions, so keep them out of Exceptions.
                        notes.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "text": line})
                    elif _is_card_continuation_line(line):
                        # Card/trade-id continuation lines are attached to the previous parsed row.
                        pass
                    elif re.match(r"^\d{1,2}/\d{2}/\d\s+", line):
                        exceptions.append({"statement_date": stmt_date, "account_number": account, "page": page_no, "section": section, "reason": "Unmatched open position line", "source_line": line})

            i += 1

        if not saw_stmt:
            exceptions.append({"statement_date": None, "account_number": account, "page": page_no, "section": None, "reason": "Statement date not found", "source_line": ""})

    # Backstop parse ensures continuation pages with only the classic open-position
    # column header are included. Avoid duplicates by source line + page.
    backstop_positions = _daily_open_positions_backstop(pdf_bytes, include_open_positions=include_open_positions)
    if backstop_positions:
        existing_keys = {(r.get("page"), str(r.get("source_line", "")).split(" | ")[0]) for r in open_positions}
        for r in backstop_positions:
            key = (r.get("page"), str(r.get("source_line", "")).split(" | ")[0])
            if key not in existing_keys:
                open_positions.append(r)
                existing_keys.add(key)

    tables: Dict[str, pd.DataFrame] = {
        "Executed Trades": pd.DataFrame(executed),
        "Purchase & Sale": pd.DataFrame(purchase_sale),
        "Receives Delivers": pd.DataFrame(receives_delivers),
        "Journal Entries": pd.DataFrame(journals),
        "Realized Gain and Loss": pd.DataFrame(),
        "Closed Positions": pd.DataFrame(closed_positions),
        "Realized PNL Summary": _parse_realized_pnl_summary_from_text(pdf_bytes),
        "Open Positions": pd.DataFrame(open_positions),
        "Notes": pd.DataFrame(notes),
        "Exceptions": pd.DataFrame(exceptions),
    }
    tables = enrich_open_positions_metadata(tables)
    tables["Summary"] = build_summary(tables)
    return tables


def enrich_open_positions_metadata(tables: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Add exchange/product metadata to raw Open Positions without changing row count."""
    pos = tables.get("Open Positions", pd.DataFrame())
    if pos is None or pos.empty or "contract_description" not in pos.columns:
        return tables
    df = pos.copy()

    # Normalize End Date / Ref Month extraction before metadata enrichment.
    # In Open Positions and Market Values statements, PyMuPDF can combine the
    # End Date and Ref Month columns into one value like "18-May-2026 MAY-26".
    # The date should drive expiryDate; Ref Month should drive Contract Month/Year.
    if "end_date" in df.columns:
        split_values = df["end_date"].apply(lambda x: pd.Series(_split_date_and_ref_month(x), index=["_parsed_end_date", "_parsed_ref_month"]))
        parsed_end = split_values["_parsed_end_date"]
        parsed_ref = split_values["_parsed_ref_month"]
        if "expiry_date" not in df.columns:
            df["expiry_date"] = None
        expiry_blank = df["expiry_date"].isna() | df["expiry_date"].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat"])
        df.loc[expiry_blank, "expiry_date"] = parsed_end.loc[expiry_blank]
        df["end_date"] = parsed_end.where(parsed_end.notna(), df["end_date"])
        if "ref_month" not in df.columns:
            df["ref_month"] = None
        existing_ref = df["ref_month"].apply(_normalize_ref_month)
        df["ref_month"] = existing_ref.where(existing_ref.notna(), parsed_ref)

    if "ref_month" in df.columns:
        normalized_ref = df["ref_month"].apply(_normalize_ref_month)
        df["ref_month"] = normalized_ref.where(normalized_ref.notna(), df["ref_month"])
        ref_parts = df["ref_month"].astype(str).str.extract(r"^(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-(\d{2})$", expand=True)
        if "contract_month" not in df.columns:
            df["contract_month"] = None
        if "contract_year" not in df.columns:
            df["contract_year"] = None
        month_blank = df["contract_month"].isna() | df["contract_month"].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat"])
        year_blank = df["contract_year"].isna() | df["contract_year"].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat"])
        df.loc[month_blank, "contract_month"] = ref_parts[0].loc[month_blank]
        df.loc[year_blank, "contract_year"] = ref_parts[1].loc[year_blank]

    parsed = df["contract_description"].apply(lambda x: pd.Series(parse_contract_product(x)))
    for col in parsed.columns:
        if col not in df.columns:
            df[col] = parsed[col]
        else:
            current = df[col]
            missing = current.isna() | (current.astype(str).str.strip() == "") | (current.astype(str).str.lower().isin(["none", "other", "unknown"]))
            df[col] = current.where(~missing, parsed[col])
    tables["Open Positions"] = df
    return tables




def _parse_realized_pnl_summary_from_text(pdf_bytes: bytes) -> pd.DataFrame:
    """Parse summary-level realized P&L rows when no row-level realized detail exists.

    Classic daily statements can show only a summary line such as:
      REALIZED PROFIT & LOSS
       USD 431,445.00- 440,805.00-
    where values are M-T-D and Y-T-D. This function captures those rows.
    """
    rows: list[dict] = []
    amount_pat = r"(?:[\d,]+(?:\.\d+)?-?|\([\d,]+(?:\.\d+)?\))"
    for page_no, text in pdf_text(pdf_bytes):
        stmt_date = _statement_date(text)
        account = _account_number(text)
        lines = [" ".join(x.strip().split()) for x in text.splitlines() if x.strip()]
        for i, line in enumerate(lines):
            if "REALIZED PROFIT" in line.upper() and "LOSS" in line.upper():
                # Usually the next non-empty line is: USD 431,445.00- 440,805.00-
                for j in range(i + 1, min(i + 6, len(lines))):
                    nxt = lines[j]
                    m = re.match(rf"^(?P<currency>[A-Z]{{3}})\s+(?P<mtd>{amount_pat})\s+(?P<ytd>{amount_pat})\s*$", nxt)
                    if m:
                        rows.append({
                            "statement_date": stmt_date,
                            "account_number": account,
                            "currency": m.group("currency"),
                            "mtd_realized_pnl": _num_any(m.group("mtd")),
                            "ytd_realized_pnl": _num_any(m.group("ytd")),
                            "source_sheet": "Statement Summary",
                            "source_section": "REALIZED PROFIT & LOSS",
                            "page": page_no,
                            "source_line": nxt,
                        })
                        break
    return pd.DataFrame(rows)

def build_summary(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for name, df in tables.items():
        if name == "Summary":
            continue
        rows.append({"sheet": name, "rows": len(df), "columns": len(df.columns) if not df.empty else 0})

    executed = tables.get("Executed Trades", pd.DataFrame())
    if not executed.empty and "quantity" in executed.columns:
        group_cols = [c for c in ["statement_date", "account_type", "contract_month", "contract_year", "contract_description"] if c in executed.columns]
        grouped = executed.groupby(group_cols, dropna=False)["quantity"].sum().reset_index() if group_cols else pd.DataFrame()
        rows.append({"sheet": "Executed Trades total quantity", "rows": float(executed["quantity"].sum()), "columns": ""})
        rows.append({"sheet": "Executed Trades grouped rows", "rows": len(grouped), "columns": len(grouped.columns) if not grouped.empty else 0})
    return pd.DataFrame(rows)


def _normalize_product_from_description(product_text: str, full_desc: str = "") -> str:
    """Normalize product text from PDF contract description into business product name."""
    raw = " ".join(str(product_text or "").upper().replace("  ", " ").split())
    full = " ".join(str(full_desc or "").upper().replace("  ", " ").split())
    search = f"{raw} {full}"

    product_mapping = {
        "COFFEE": "Coffee",
        "COFFEE C": "Coffee",
        "COFFEE P": "Coffee",
        "COTTON": "Cotton",
        # ASX/SFE/NZFE instruments — more specific keys must precede shorter overlapping ones.
        # Equity index
        "SPI 200 INDEX": "SPI 200 Index",
        "SPI 200": "SPI 200 Index",
        "NIKKEI 225": "Nikkei 225",
        # Interest rate / bonds
        "10 YEAR T-BOND": "10-Year T-Bond",
        "10Y T-BOND": "10-Year T-Bond",
        "3-YEAR TREASURY BOND": "3-Year T-Bond",
        "3Y T-BOND": "3-Year T-Bond",
        "BANK BILLS": "Bank Bills",
        "BANK ACCEPTD BILL": "90-Day Bank Bill",
        "30 DAY INTERBANK": "30-Day Interbank",
        # NZ electricity (NZFE) — quarterly variants before base
        "BENMORE BASELOAD Q": "NZ Benmore Quarterly",
        "OTAHUHU BASE LOAD Q": "NZ Otahuhu Quarterly",
        "BENMORE BASELOAD": "NZ Benmore Baseload",
        "OTAHUHU BASE LOAD": "NZ Otahuhu Baseload",
        # AU electricity strip options (SFE) — before base-load catch-all
        "BASE LD NSW STRIP": "NSW Strip",
        "BASE LD VIC STRIP": "VIC Strip",
        "BASE LD QLD STRIP": "QLD Strip",
        "BASE LD SA STRIP": "SA Strip",
        # AU electricity base-load futures (SFE) — region-specific
        "NSW BASE LOAD ELEC": "NSW Baseload Electricity",
        "NSW BASE QLY": "NSW Quarterly Strip",
        "VIC BASE LOAD ELEC": "VIC Baseload Electricity",
        "VIC BASE QLY": "VIC Quarterly Strip",
        "QLD BASE LOAD ELEC": "QLD Baseload Electricity",
        "QLD BASE QLY": "QLD Quarterly Strip",
        "SA BASE LOAD ELEC": "SA Baseload Electricity",
        "SA BASE QLY": "SA Quarterly Strip",
        # Grains
        "EASTERN AUSTRALIA WHEAT": "Eastern Australia Wheat",
        "EAU WHEAT": "Eastern Australia Wheat",
        "WHEAT": "Wheat",
        "AUST FEED BARLEY": "Feed Barley",
        "FEED BARLEY": "Feed Barley",
        "BARLEY": "Barley",
        # Metals / dairy / rubber
        "IRON ORE F": "Iron Ore",
        "IRON ORE": "Iron Ore",
        "SKMILK": "Skim Milk",
        "SKIM MILK": "Skim Milk",
        "WHOLE MILK": "Whole Milk",
        "WHMILK": "Whole Milk",
        "TSR20RUBBR": "TSR20 Rubber",
        "RUBBR": "Rubber",
        # FX
        "USD/KRW": "USD/KRW",
        "INR/USD": "INR/USD",
        # Agricultural (CBOT / CME)
        "CORN": "Corn",
        # Multi-word variants must come before the bare "SOYBEAN" entry,
        # because the lookup uses substring matching.
        "SOYBEAN OIL": "Soybean Oil",
        "SOYBEAN MEAL": "Soybean Meal",
        "SOYBEAN": "Soybean",
        "SOYBEANS": "Soybean",
        "CANOLA": "Canola",
        "RAPESEED": "Rapeseed",
        "RAPE SEED": "Rapeseed",
    }
    for key, value in product_mapping.items():
        if key in search:
            return value

    # Remove common option markers, C/P option flags, strike prices, and suffixes such as SE
    # so examples like "COFFEE C 2750 SE" and "COFFEE C 3250 SE" normalize to Coffee.
    cleaned = re.sub(r"\b(EUROPEAN\s+OPTION|EURO\s+OPTION|OPTION|CALL|PUT|PULL|O)\b", " ", raw, flags=re.I)
    cleaned = re.sub(r"\b(C|P)\s+\d+(?:\.\d+)?\b", " ", cleaned, flags=re.I)
    # Remove option strike + status suffix from product text. Examples:
    #   GA 2850 E      -> GA
    #   COFFEE C 2750 SE -> COFFEE
    cleaned = re.sub(r"\b\d+(?:\.\d+)?\b\s+\b(?:E|SE|CE|PE|EU|US|AD)\b\s*$", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\bSE\b", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\b\d+(?:\.\d+)?\b\s*$", "", cleaned).strip()
    return " ".join(cleaned.split()) or raw or "Unknown"


def parse_contract_product(desc: str | None) -> dict:
    """Parse a StoneX contract description into grouping fields.

    The PDF contract description is the best source for product identity.
    Example: "MAY 26 ASX EAU WHEAT" maps to exchange=ASX and product=Wheat.
    """
    desc = "" if desc is None else str(desc).strip()
    upper = " ".join(desc.upper().split())

    exchanges = "LME|SCM|BMF|CBOT|CBT|NYMEX|NYME|CME|ICE|IFUS|MATF|MGEX|IMM|ASX|SFE|SGX|KFX|NZFE|NZF|ABX|TOCOM|FX"
    months = "JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC"

    option_type = "Call" if re.search(r"\b(CALL|C)\b", upper) else ("Put" if re.search(r"\b(PUT|P|PULL)\b", upper) else None)
    option_style = "European" if re.search(r"\b(?:EURO|EUROPEAN)\s+OPTION\b", upper) else None

    strike_value = None
    strike_patterns = [
        # COFFEE C 2750 SE / COFFEE P 3250 SE
        r"\b(?:C|P|CALL|PUT)\s+(\d+(?:\.\d+)?)\b",
        # LME/Euro-option style descriptions
        r"\b(\d+(?:\.\d+)?)\s+(?:Euro|European)\s+Option\b",
        # CALL JUN 26 NYME GA 2850 / PUT MAY 26 IMM WDPAUWK4 7150
        rf"^(?:CALL|PUT|C|P)\s+(?:{months})\s+\d{{2,4}}\s+.*?\b(\d+(?:\.\d+)?)\b(?:\s+[A-Z]{{1,4}})?\s*$",
        # Exchange-led option descriptions where the strike is the last numeric token.
        rf"\b(?:{exchanges})\b.+?\b(\d+(?:\.\d+)?)\b(?:\s+[A-Z]{{1,4}})?\s*$",
    ]
    for pat in strike_patterns:
        mstrike = re.search(pat, desc, re.I)
        if mstrike and option_type is not None:
            strike_value = _num_any(mstrike.group(1))
            break

    # Final safety net for options: after removing CALL/PUT and the ref month/year,
    # the strike is usually the last numeric token in the contract description.
    # This handles PDFs like "CALL JUN 26 NYME GA 2850 E" where 2850 is printed
    # as a separate option strike column before the status code.
    if strike_value is None and option_type is not None:
        option_body = re.sub(r"^(CALL|PUT|PULL|C|P)\s+", "", upper, flags=re.I).strip()
        option_body = re.sub(rf"^(?:{months})\s+\d{{2,4}}\s+", "", option_body, flags=re.I).strip()
        numeric_tokens = re.findall(r"\b\d+(?:\.\d+)?\b", option_body)
        if numeric_tokens:
            strike_value = _num_any(numeric_tokens[-1])

    # OTC accumulators/swaps are not listed options, but the base level still
    # belongs in strikePrice for position review.
    if strike_value is None:
        strike_value = _extract_otc_strike(desc)


    # Murex FX pair format: "USD 15,000,000.00 BRL ..." or "USD (15,000,000.00) BRL ..."
    # Guard: first token must not be a calendar month (e.g. "JUL 26 CBT CORN" would otherwise
    # match with ccy1=JUL, amount=26, ccy2=CBT — producing a bogus "JUL/CBT" FX product).
    _MONTH_ABBREVS = {"JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"}
    _EXCHANGE_ABBREVS = {"LME","SCM","BMF","CBOT","CBT","NYMEX","NYME","CME","ICE","IFUS",
                         "MATF","MGEX","IMM","ASX","SFE","SGX","KFX","NZFE","NZF","ABX","TOCOM"}
    murex_fx_m = re.match(r"^([A-Z]{3})\s+\(?[\d,.]+\)?\s+([A-Z]{3})\b", upper)
    if (murex_fx_m
            and murex_fx_m.group(1) not in _MONTH_ABBREVS
            and murex_fx_m.group(2) not in _MONTH_ABBREVS
            and murex_fx_m.group(1) not in _EXCHANGE_ABBREVS
            and murex_fx_m.group(2) not in _EXCHANGE_ABBREVS):
        ccy1 = murex_fx_m.group(1).upper()
        ccy2 = murex_fx_m.group(2).upper()
        pair = f"{ccy1}/{ccy2}"
        return {"product": pair, "exchange": "FX", "product_name": pair, "strike": strike_value, "option_type": option_type, "option_style": option_style, "unit": None}

    # Special handling for FX rows, e.g. FXCADUSD 13-Nov-26.
    fxm = re.match(r"^(FX[A-Z]{6})\b", upper)
    if fxm:
        code = fxm.group(1)
        product_name = _fx_product_from_code(code) if '_fx_product_from_code' in globals() else code
        return {"product": product_name, "exchange": "FX", "product_name": product_name, "strike": strike_value, "option_type": option_type, "option_style": option_style, "unit": None}

    # Special handling for LME rows.
    if upper.startswith("LME "):
        parts = desc.split()
        exchange = "LME"
        unit = parts[-1] if len(parts) > 2 and re.match(r"^[A-Z]{3}|\$$", parts[-1]) else None
        raw_product = " ".join(parts[1:-1] if unit else parts[1:])
        product_name = _normalize_product_from_description(raw_product, desc)
        return {"product": product_name, "exchange": exchange, "product_name": product_name, "strike": strike_value, "option_type": option_type, "option_style": option_style, "unit": unit}

    # Remove leading CALL/PUT then leading delivery date or month/year before detecting exchange.
    work = re.sub(r"^(CALL|PUT|PULL|C|P)\s+", "", upper, flags=re.I).strip()
    work = re.sub(r"^\d{1,2}-[A-Z]{3}-\d{2,4}\s+", "", work, flags=re.I).strip()
    work = re.sub(rf"^(?:{months})(?:\s+|-)\d{{2,4}}\s+", "", work, flags=re.I).strip()

    m = re.search(rf"\b({exchanges})\b\s+(.+)$", work, re.I)
    if m:
        exchange = m.group(1).upper()
        product_text = m.group(2).strip()
    else:
        # Fallback for older CME/CBOT descriptions where price follows product.
        m2 = re.search(rf"\b({exchanges})\b\s+(.+?)(?=\s+\d+(?:\.\d+)?(?:\s|$))", upper, re.I)
        exchange = m2.group(1).upper() if m2 else None
        product_text = m2.group(2).strip() if m2 else upper

    # When the exchange resolves to "FX", the text after "FX" is the direction/month
    # (e.g. "FWD AUG-26"), NOT a product name. The true product identity is the
    # currency pair that precedes "FX" in the description (e.g. "USD/BRL").
    # We return "FX USD/BRL" so that:
    #   (a) all trades in the same pair share one product regardless of expiry date
    #       → essential for the Product aggregation grouping to collapse correctly, and
    #   (b) the "FX " prefix keeps _fx_position_mask detection working.
    if exchange == "FX":
        ccy_pair_m = re.search(r"\b([A-Z]{3})/([A-Z]{3})\b", upper)
        if ccy_pair_m:
            ccy_pair = f"{ccy_pair_m.group(1)}/{ccy_pair_m.group(2)}"
            # Return immediately — no further normalization needed for FX pairs.
            return {
                "product": f"FX {ccy_pair}",
                "exchange": "FX",
                "product_name": f"FX {ccy_pair}",
                "strike": strike_value,
                "option_type": option_type,
                "option_style": option_style,
                "unit": None,
            }
        # If no explicit ccy_pair, retain whatever _normalize_product_from_description
        # finds — it may already have a sensible mapping for the product_text token.

    unit_match = re.search(r"\b[A-Z]{3}/([A-Z]+)\b", upper)
    product_name = _normalize_product_from_description(product_text, desc)

    return {
        "product": product_name,
        "exchange": exchange,
        "product_name": product_name,
        "strike": strike_value,
        "option_type": option_type,
        "option_style": option_style,
        "unit": unit_match.group(1) if unit_match else None,
    }


def _prepared_positions_for_grouping(tables: Dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, str | None, str | None]:
    """Normalize open positions before any grouping so app selections are truly dynamic."""
    pos = tables.get("Open Positions", pd.DataFrame())
    if pos is None or pos.empty:
        return pd.DataFrame(), None, None

    df = pos.copy()
    if "contract_description" not in df.columns:
        return pd.DataFrame(), None, None

    parsed = df["contract_description"].apply(lambda x: pd.Series(parse_contract_product(x)))
    for col in parsed.columns:
        if col not in df.columns:
            df[col] = parsed[col]
        else:
            current = df[col]
            missing = current.isna() | (current.astype(str).str.strip() == "") | (current.astype(str).str.lower().isin(["none", "other", "unknown"]))
            df[col] = current.where(~missing, parsed[col])

    if "long" in df.columns or "short" in df.columns:
        long_qty = pd.to_numeric(df.get("long", 0), errors="coerce").fillna(0)
        short_qty = pd.to_numeric(df.get("short", 0), errors="coerce").fillna(0)
        df["net_qty"] = long_qty - short_qty
        df["long_qty"] = long_qty
        df["short_qty"] = short_qty
    elif "quantity" in df.columns:
        qty = pd.to_numeric(df["quantity"], errors="coerce").fillna(0)
        if "side" in df.columns:
            side = df["side"].fillna("").astype(str).str.lower()
            df["net_qty"] = qty.where(~side.str.contains("short|sell"), -qty)
        else:
            df["net_qty"] = qty
        df["long_qty"] = df["net_qty"].where(df["net_qty"] > 0, 0)
        df["short_qty"] = (-df["net_qty"]).where(df["net_qty"] < 0, 0)
    else:
        df["net_qty"] = 0.0
        df["long_qty"] = 0.0
        df["short_qty"] = 0.0

    if "ref_month" not in df.columns:
        if "contract_month" in df.columns and "contract_year" in df.columns:
            df["ref_month"] = df["contract_month"].astype(str) + "-" + df["contract_year"].astype(str)
        elif "contract_month" in df.columns:
            df["ref_month"] = df["contract_month"].astype(str)
        else:
            df["ref_month"] = None

    if "end_date" in df.columns:
        split_values = df["end_date"].apply(lambda x: pd.Series(_split_date_and_ref_month(x), index=["_parsed_end_date", "_parsed_ref_month"]))
        parsed_end = split_values["_parsed_end_date"]
        parsed_ref = split_values["_parsed_ref_month"]
        df["ref_month"] = df["ref_month"].where(_series_nonblank(df["ref_month"]), parsed_ref)
        df["end_date"] = parsed_end.where(parsed_end.notna(), df["end_date"])
        if "expiry_date" not in df.columns:
            df["expiry_date"] = None
        expiry_blank = df["expiry_date"].isna() | df["expiry_date"].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat"])
        df.loc[expiry_blank, "expiry_date"] = parsed_end.loc[expiry_blank]

    # Normalize instrument type / FX / OTC fields before grouping.
    df = _ensure_fx_position_columns(df)
    df = _ensure_otc_position_columns(df)
    df = _ensure_position_type_column(df)

    df = _add_expiry_or_ref_group_key(df)

    # Flag options before grouping so grouped risk views can split:
    # - NOV = option OTE / net option value
    # - Unrealised PNL (OTE) = non-option OTE only
    option_signal = pd.Series(False, index=df.index)
    for opt_col in ["option_type", "option_type_raw", "call_put", "Call/Put"]:
        if opt_col in df.columns:
            option_signal = option_signal | df[opt_col].astype(str).str.strip().str.lower().isin(["c", "call", "p", "put"])
    # Some statements include option rows where CALL/PUT is embedded in the description.
    if "contract_description" in df.columns:
        option_signal = option_signal | df["contract_description"].astype(str).str.upper().str.contains(r"\b(?:CALL|PUT)\b", regex=True)
    df["_is_option"] = option_signal.fillna(False)
    df["_option_position_rows"] = df["_is_option"].astype(int)
    df["_non_option_position_rows"] = (~df["_is_option"]).astype(int)

    price_col = "trade_price" if "trade_price" in df.columns else ("price" if "price" in df.columns else None)
    # Use Market Value as the canonical open-trade value for grouping whenever
    # it is available. Other value fields are only fallback sources if Market Value
    # is missing for a row.
    df["_open_trade_value_for_risk"] = _coalesced_open_trade_value(df)
    mv_col = "_open_trade_value_for_risk" if df["_open_trade_value_for_risk"].notna().any() else None
    if price_col:
        df[price_col] = pd.to_numeric(df[price_col], errors="coerce")
    if mv_col:
        mv_values = pd.to_numeric(df[mv_col], errors="coerce")
        df["_nov_value"] = mv_values.where(df["_is_option"], 0.0)
        df["_ote_non_option_value"] = mv_values.where(~df["_is_option"], 0.0)
    else:
        df["_nov_value"] = 0.0
        df["_ote_non_option_value"] = 0.0

    return df, price_col, mv_col





def _coalesce_column(df: pd.DataFrame, candidates: list[str], default=None) -> pd.Series:
    """Return the first available column from candidates as a Series."""
    for col in candidates:
        if col in df.columns:
            return df[col]
    return pd.Series([default] * len(df), index=df.index)


def _fill_missing_from_candidates(df: pd.DataFrame, target: str, candidates: list[str], default=None) -> pd.DataFrame:
    values = _coalesce_column(df, candidates, default=default)
    if target not in df.columns:
        df[target] = values
    else:
        current = df[target]
        missing = current.isna() | current.astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "unknown", "other"])
        df[target] = current.where(~missing, values)
    return df


def _ensure_fx_position_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize FX currency/amount fields for Open Positions and grouping.

    StoneX IFL FX sections print:
        Primary Amount / Ccy / Rate / Secondary Amount / Ccy / MTM
    These map to:
        CCY 1 Amount, CCY 1, Trade Price, CCY 2 Amount, CCY 2, OTE
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    # Normalize combined End Date / Ref Month values before building expiryDate.
    if "end_date" in out.columns:
        split_values = out["end_date"].apply(lambda x: pd.Series(_split_date_and_ref_month(x), index=["_parsed_end_date", "_parsed_ref_month"]))
        parsed_end = split_values["_parsed_end_date"]
        parsed_ref = split_values["_parsed_ref_month"]
        out["end_date"] = parsed_end.where(parsed_end.notna(), out["end_date"])
        if "ref_month" not in out.columns:
            out["ref_month"] = None
        out["ref_month"] = out["ref_month"].where(_series_nonblank(out["ref_month"]), parsed_ref)
    if "ref_month" in out.columns:
        normalized_ref = out["ref_month"].apply(_normalize_ref_month)
        out["ref_month"] = normalized_ref.where(normalized_ref.notna(), out["ref_month"])
    out = _fill_missing_from_candidates(out, "ccy_1", ["ccy_1", "ccy1", "primary_currency", "primary_ccy"])
    out = _fill_missing_from_candidates(out, "ccy_2", ["ccy_2", "ccy2", "secondary_currency", "secondary_ccy"])
    out = _fill_missing_from_candidates(out, "ccy_1_amount", ["ccy_1_amount", "ccy1_amount", "primary_amount", "primary_amount_signed"])
    out = _fill_missing_from_candidates(out, "ccy_2_amount", ["ccy_2_amount", "ccy2_amount", "secondary_amount", "secondary_amount_signed"])
    out = _fill_missing_from_candidates(out, "trade_price", ["trade_price", "rate", "price"])

    # expiry_date is the user-facing date dimension for positions.
    # Priority:
    # - options: true Expiration Date from the option summary lines
    # - FX: Value Date from the FX OPEN POSITIONS header
    # - futures/forwards/LME: Delivery date from Delivery / Product
    # This lets Product + Contract Month/Year and Account Grouping use the actual
    # date when the statement provides one, while Product grouping stays product-only.
    out = _fill_missing_from_candidates(
        out,
        "expiry_date",
        ["expiry_date", "expiryDate", "expiration_date", "value_date", "delivery_date", "contract_date", "end_date"],
        default=None,
    )
    out["expiry_date"] = out["expiry_date"].astype(object)
    for date_col in ["expiration_date", "value_date", "delivery_date", "contract_date", "end_date"]:
        if date_col in out.columns:
            normalized_dates = out[date_col].apply(_normalize_any_date)
            expiry_blank = out["expiry_date"].isna() | out["expiry_date"].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat"])
            out.loc[expiry_blank, "expiry_date"] = normalized_dates.loc[expiry_blank]

    # Keep raw settlement_date in raw tables for audit/download compatibility,
    # but all user-facing position views remove Settlement Date.
    out = _fill_missing_from_candidates(out, "settlement_date", ["settlement_date", "value_date", "delivery_date", "contract_date", "end_date"], default=None)

    for col in ["ccy_1_amount", "ccy_2_amount", "trade_price"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    return out




def _series_nonblank(series: pd.Series) -> pd.Series:
    """True where display/grouping values are meaningful."""
    if series is None:
        return pd.Series(dtype=bool)
    text = series.astype(str).str.strip().str.lower()
    return series.notna() & ~text.isin(["", "none", "nan", "nat", "unknown", "other", "multiple"])


def _add_expiry_or_ref_group_key(df: pd.DataFrame) -> pd.DataFrame:
    """Use true expiry/value date as the month grouping key when available.

    Product + Contract Month/Year and Account Grouping still display
    Contract Month/Year, but rows with an actual expiryDate/value date group
    by that date so different expiries do not collapse together.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    if "ref_month" not in out.columns:
        out["ref_month"] = None
    if "expiry_date" not in out.columns:
        out["expiry_date"] = None
    expiry_norm = out["expiry_date"].apply(_normalize_any_date)
    expiry_norm = expiry_norm.where(_series_nonblank(expiry_norm), out["expiry_date"])
    out["_group_expiry_or_ref_month"] = out["ref_month"]
    has_expiry = _series_nonblank(expiry_norm)
    out.loc[has_expiry, "_group_expiry_or_ref_month"] = expiry_norm.loc[has_expiry]
    return out


def _expiry_aware_group_cols(group_cols: list[str] | None, df: pd.DataFrame) -> list[str]:
    """Replace ref_month grouping with expiry/value-date-aware key where possible."""
    cols = list(group_cols or [])
    if "ref_month" in cols and df is not None and "_group_expiry_or_ref_month" in df.columns:
        cols = ["_group_expiry_or_ref_month" if c == "ref_month" else c for c in cols]
    return cols

def _fx_position_mask(df: pd.DataFrame) -> pd.Series:
    """Identify FX open-position rows."""
    if df is None or df.empty:
        return pd.Series(dtype=bool)
    mask = pd.Series(False, index=df.index)
    if "source_section" in df.columns:
        mask = mask | df["source_section"].astype(str).str.upper().str.contains("FX OPEN", na=False)
    if "exchange" in df.columns:
        mask = mask | df["exchange"].astype(str).str.upper().eq("FX")
    if "product" in df.columns:
        mask = mask | df["product"].astype(str).str.upper().str.startswith("FX ")
    if "contract_description" in df.columns:
        mask = mask | df["contract_description"].astype(str).str.upper().str.match(r"^FX[A-Z]{6}\b", na=False)
    return mask.fillna(False)


def _fx_group_cols_for_base(group_cols: list[str] | None, mode: str = "custom") -> list[str]:
    """Return FX grouping keys for the selected grouped-position view.

    Product grouping remains a product-level summary. Product + Contract
    Month/Year and Account Grouping keep FX option/NDO economics by adding
    Call/Put and strike to the key. Plain FX forwards/spots/swaps have those
    fields blank, so they continue grouping by product/expiry only.

    ccy_1/ccy_2 are intentionally excluded — parse_contract_product now
    encodes the CCY pair directly in the product name (e.g. "FX USD/BRL"),
    so those fields are redundant grouping keys and are not shown in the UI.
    """
    base = [str(c) for c in (group_cols or [])]
    base_set = set(base)
    fx_option_keys = ["option_type", "strike"]
    if mode == "product_month":
        return ["product", "position_type", "expiry_date"] + fx_option_keys
    if "account_number" in base_set:
        return ["account_number", "product", "position_type", "expiry_date"] + fx_option_keys
    if base_set <= {"product", "exchange"}:
        return [c for c in ["product", "exchange"] if c in base]
    return ["product", "position_type", "expiry_date"] + fx_option_keys


def _group_prepared_positions_with_fx(
    df: pd.DataFrame,
    base_group_cols: list[str] | None,
    price_col: str | None,
    mv_col: str | None,
    fx_group_cols: list[str] | None = None,
) -> pd.DataFrame:
    """Group non-FX rows normally and FX rows using FX-specific keys."""
    if df is None or df.empty:
        return pd.DataFrame()
    work = _ensure_fx_position_columns(df)
    fx_mask = _fx_position_mask(work)
    frames: list[pd.DataFrame] = []

    non_fx = work.loc[~fx_mask].copy()
    if not non_fx.empty:
        frames.append(_group_prepared_positions(non_fx, base_group_cols or [], price_col, mv_col))

    fx = work.loc[fx_mask].copy()
    if not fx.empty:
        fx_cols = fx_group_cols or _fx_group_cols_for_base(base_group_cols)
        frames.append(_group_prepared_positions(fx, fx_cols, price_col, mv_col))

    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)

def _unique_or_multiple(series: pd.Series):
    vals = [v for v in series.dropna().unique().tolist() if str(v).strip() != "" and str(v).strip().lower() != "none"]
    if not vals:
        return None
    return vals[0] if len(vals) == 1 else "Multiple"



def _raw_option_position_mask(df: pd.DataFrame) -> pd.Series:
    """Identify option rows before grouping so option value can be split into NOV.

    Option rows should contribute their OTE/market value to NOV, while futures/forwards
    contribute to Unrealised PNL (OTE). This is safer than trying to infer options after
    grouping, because a Product-only group may contain both futures and options.
    """
    if df is None or df.empty:
        return pd.Series(dtype=bool)
    mask = pd.Series(False, index=df.index)
    for col in ["option_type", "option_type_raw", "call_put", "Call/Put"]:
        if col in df.columns:
            values = df[col].astype(str).str.strip().str.lower()
            mask = mask | values.isin(["c", "p", "call", "put"])
    if "contract_description" in df.columns:
        desc = df["contract_description"].astype(str).str.upper()
        mask = mask | desc.str.contains(r"\b(?:CALL|PUT)\b", regex=True)
    return mask.fillna(False)

def _group_prepared_positions(df: pd.DataFrame, group_cols: list[str], price_col: str | None, mv_col: str | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    group_cols = [c for c in group_cols if c in df.columns]
    if not group_cols:
        group_cols = [c for c in ["account_number", "product", "option_type", "ref_month"] if c in df.columns]
    if not group_cols:
        return df

    df = df.copy()
    df["_position_rows"] = 1

    agg = {
        "long_qty": "sum",
        "short_qty": "sum",
        "net_qty": "sum",
        "_position_rows": "sum",
    }
    for split_col in ["_nov_value", "_ote_non_option_value", "_option_position_rows", "_non_option_position_rows"]:
        if split_col in df.columns:
            agg[split_col] = "sum"
    # Currency amount columns are additive for FX grouped views.
    for amt_col in ["ccy_1_amount", "ccy_2_amount", "primary_amount", "secondary_amount"]:
        if amt_col in df.columns and amt_col not in group_cols and amt_col not in agg:
            agg[amt_col] = "sum"
    # Preserve useful descriptor fields on grouped rows when they are unique inside the group.
    descriptor_cols = [
        "exchange", "product_name", "contract_description", "currency", "unit", "type", "position_type", "trade_type", "trigger_barrier", "ref_price", "original_quantity",
        "statement_date", "account_number", "ref_month", "contract_month", "contract_year",
        "option_type", "option_type_raw", "strike", "source_system", "sourceSystem",
        "ccy_1", "ccy_2", "primary_currency", "secondary_currency",
        "rate", "trade_price",
        "settlement_date", "delivery_date", "end_date", "expiry_date", "closing_price", "settlement_price"
    ]
    for desc_col in descriptor_cols:
        if desc_col in df.columns and desc_col not in group_cols and desc_col not in agg:
            agg[desc_col] = _unique_or_multiple
    if price_col:
        df["_weighted_price_qty"] = df[price_col].fillna(0) * df["net_qty"].abs()
        agg["_weighted_price_qty"] = "sum"
    if mv_col:
        # Split value before grouping:
        # - futures/forwards -> Unrealised PNL (OTE)
        # - options -> NOV (Net Option Value)
        # This keeps Product-only groups accurate when they contain both futures and options.
        option_mask = _raw_option_position_mask(df)
        mv_values = df[mv_col].apply(_num_any) if hasattr(df[mv_col], "apply") else pd.to_numeric(df[mv_col], errors="coerce")
        df["_ote_value"] = mv_values.where(~option_mask, 0.0)
        df["_nov_value"] = mv_values.where(option_mask, 0.0)
        agg["_ote_value"] = "sum"
        agg["_nov_value"] = "sum"

    grouped = df.groupby(group_cols, dropna=False).agg(agg).reset_index()
    rename_map = {"_position_rows": "position_rows"}
    if mv_col:
        rename_map["_ote_value"] = "market_value"
        rename_map["_nov_value"] = "nov"
    grouped = grouped.rename(columns=rename_map)

    if price_col:
        qty_abs = df.groupby(group_cols, dropna=False)["net_qty"].apply(lambda s: s.abs().sum()).reset_index(name="_abs_qty")
        grouped = grouped.merge(qty_abs, on=group_cols, how="left")
        grouped["avg_trade_price"] = grouped.apply(lambda r: r["_weighted_price_qty"] / r["_abs_qty"] if r.get("_abs_qty", 0) else None, axis=1)
        grouped = grouped.drop(columns=["_weighted_price_qty", "_abs_qty"], errors="ignore")

    return grouped.sort_values(group_cols).reset_index(drop=True)


def grouped_positions_custom(tables: Dict[str, pd.DataFrame], group_cols: list[str] | None = None) -> pd.DataFrame:
    """Group positions using caller-selected dimensions from the sidebar.

    FX rows use purpose-built grouping keys because their risk identity is the
    currency pair, primary/secondary amounts, trade price, and expiry date rather
    than futures contract month/strike.
    """
    df, price_col, mv_col = _prepared_positions_for_grouping(tables)
    if not group_cols:
        group_cols = ["account_number", "product", "position_type", "option_type", "ref_month"]
    group_cols = _expiry_aware_group_cols(group_cols, df)
    fx_group_cols = _fx_group_cols_for_base(group_cols)
    return _group_prepared_positions_with_fx(df, group_cols, price_col, mv_col, fx_group_cols=fx_group_cols)



STANDARD_POSITION_COLUMNS = [
    "Global ID",
    "Contract Description",
    "Product",
    "Type",
    "Exchange",
    "Currency",
    "Trigger/Barrier",
    "Ref Price",
    "Original Quantity",
    "CCY 1",
    "CCY 1 Amount",
    "CCY 2",
    "CCY 2 Amount",
    "expiryDate",
    "End Date",
    "settlementPrice",
    "Trade ID",
    "Trade Date",
    "Account Number",
    "Broker Code",
    "Net Quantity",
    "Last Update",
    "Contract Month/Year",
    "Trade Price",
    "Avg Fill Price",
    "Delta",
    "Call/Put",
    "strikePrice",
    "sourceSystem",
    "NOV",
    "Unrealised PNL (OTE)",
    "Realised PNL",
    "Day PNL",
    "Market Value",
]


OPEN_POSITION_COLUMNS = [
    "Global ID",
    "Trade ID",
    "Trade Date",
    "Product",
    "Type",
    "Contract Description",
    "Exchange",
    "Currency",
    "Trigger/Barrier",
    "Ref Price",
    "Original Quantity",
    "CCY 1",
    "CCY 1 Amount",
    "CCY 2",
    "CCY 2 Amount",
    "Account Number",
    "Broker Code",
    "Quantity",
    "expiryDate",
    "Contract Month/Year",
    "Trade Price",
    "settlementPrice",
    "Call/Put",
    "strikePrice",
    "NOV",
    "Unrealised PNL (OTE)",
]

GROUPED_POSITION_COLUMNS = [
    "Product",
    "Type",
    "Exchange",
    "Contract Month/Year",
    "expiryDate",
    "Trigger/Barrier",
    "Currency",
    "CCY 1",
    "CCY 1 Amount",
    "CCY 2",
    "CCY 2 Amount",
    "Call/Put",
    "strikePrice",
    "Delta",
    "Account Number",
    "Net Quantity",
    "Avg Fill Price",
    "settlementPrice",
    "NOV",
    "Unrealised PNL (OTE)",
]


def _first_existing(df: pd.DataFrame, cols: list[str], default=None):
    for col in cols:
        if col in df.columns:
            return df[col]
    return pd.Series([default] * len(df), index=df.index)


def _combine_month_year(df: pd.DataFrame) -> pd.Series:
    if "ref_month" in df.columns:
        return df["ref_month"]
    if "contract_month" in df.columns and "contract_year" in df.columns:
        return df["contract_month"].astype(str).str.strip() + "-" + df["contract_year"].astype(str).str.strip()
    if "contract_month" in df.columns:
        return df["contract_month"]
    # LME/metal statements often provide a true delivery date instead of an
    # exchange month code. Derive Contract Month/Year from Delivery when needed.
    if "delivery_date" in df.columns:
        return df["delivery_date"].apply(_ref_month_from_date)
    if "contract_date" in df.columns:
        return df["contract_date"].apply(_ref_month_from_date)
    return pd.Series([None] * len(df), index=df.index)


def _compute_net_qty_for_view(df: pd.DataFrame) -> pd.Series:
    if "net_qty" in df.columns:
        return pd.to_numeric(df["net_qty"], errors="coerce")
    if "long_qty" in df.columns or "short_qty" in df.columns:
        long_qty = pd.to_numeric(df.get("long_qty", 0), errors="coerce").fillna(0)
        short_qty = pd.to_numeric(df.get("short_qty", 0), errors="coerce").fillna(0)
        return long_qty - short_qty
    if "long" in df.columns or "short" in df.columns:
        long_qty = pd.to_numeric(df.get("long", 0), errors="coerce").fillna(0)
        short_qty = pd.to_numeric(df.get("short", 0), errors="coerce").fillna(0)
        return long_qty - short_qty
    if "quantity" in df.columns:
        return pd.to_numeric(df["quantity"], errors="coerce")
    return pd.Series([None] * len(df), index=df.index)


def _direction_from_net_qty(net_qty: pd.Series) -> pd.Series:
    def _direction(v):
        try:
            if pd.isna(v):
                return None
            if float(v) > 0:
                return "Long"
            if float(v) < 0:
                return "Short"
            return "Flat"
        except Exception:
            return None
    return net_qty.apply(_direction)







def _normalize_position_type_value(value):
    """Normalize raw statement type text into the user-facing Type values."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "nat", "unknown", "other"}:
        return None
    upper = re.sub(r"\s+", " ", text.upper())
    if upper == "MULTIPLE":
        return "Multiple"
    if upper in {"FX FWD", "FX FORWARD", "FORWARD"}:
        return "FX Forward"
    if upper in {"FX SPOT", "SPOT"}:
        return "FX Spot"
    if upper in {"FX SWAP", "FXSWAP", "SWAP FX"}:
        return "FX Swap"
    if upper in {"FX NDF", "NDF FX", "NDF", "NON DELIVERABLE FORWARD", "NON-DELIVERABLE FORWARD", "NON DELIVERABLE FORWARDS", "NON-DELIVERABLE FORWARDS"}:
        return "FX NDF"
    if upper in {"NDO", "FX NDO", "NDO FX", "NON DELIVERABLE OPTION", "NON-DELIVERABLE OPTION", "NON DELIVERABLE OPTIONS", "NON-DELIVERABLE OPTIONS"}:
        return "NDO"
    if upper.startswith("FX FWD"):
        return "FX Forward"
    if upper.startswith("FX SPOT"):
        return "FX Spot"
    if upper.startswith("FX SWAP"):
        return "FX Swap"
    if upper.startswith("FX NDF") or upper.startswith("NDF FX") or upper.startswith("NDF "):
        return "FX NDF"
    if "ACCUMULATOR" in upper or re.search(r"\bACCUM\b", upper):
        return "OTC Accumulator"
    if "OPTION" in upper or upper in {"CALL", "PUT", "C", "P"}:
        return "Option"
    if "SWAP" in upper or "OTC" in upper:
        return "OTC Swap"
    if upper in {"FUT", "FUTURE", "FUTURES"}:
        return "Future"
    return text


def _infer_position_type_row(row) -> str | None:
    """Infer Type for Open Positions and grouped-position descriptor rows."""
    def get_any(*names):
        for name in names:
            try:
                if name in row.index:
                    value = row.get(name)
                    norm = _normalize_position_type_value(value)
                    if norm:
                        return norm
            except Exception:
                pass
        return None

    raw_type = get_any("type", "Type", "position_type", "trade_type")
    desc = ""
    for name in ["contract_description", "Contract Description", "product", "Product"]:
        try:
            if name in row.index and row.get(name) is not None and not pd.isna(row.get(name)):
                desc = str(row.get(name)).strip()
                if desc:
                    break
        except Exception:
            pass
    upper_desc = desc.upper()

    # FX rows: prefer explicit FX type column. If the older FX layout has no
    # explicit Type, open FX rows are forwards unless the description says Spot/NDF/Swap.
    is_fx = False
    for name in ["source_section", "exchange", "Exchange", "product", "Product", "contract_description"]:
        try:
            if name in row.index:
                val = str(row.get(name) or "").upper().strip()
                if "FX OPEN" in val or val == "FX" or val.startswith("FX"):
                    is_fx = True
        except Exception:
            pass
    if is_fx:
        if raw_type == "NDO":
            return "NDO"
        if raw_type and raw_type.startswith("FX"):
            return raw_type
        if "NDF" in upper_desc:
            return "FX NDF"
        if "SWAP" in upper_desc:
            return "FX Swap"
        if "SPOT" in upper_desc:
            return "FX Spot"
        return "FX Forward"

    # OTC accumulator structures should be identified distinctly from plain OTC swaps
    # and should override any generic option/swap wording in the description.
    # Examples:
    #   "ICE Cotton 0.8456 Daily Consumer Accum No KO 50% Upfront ..." (has Accum)
    #   "ICE Cotton LVL1 0.6625 /Trigger:0.7074 Daily Cons Range w/Daily DU OQ109 BP:68.62"
    #   "ICE Cotton LVL1 0.7100 /Trigger:0.6632 Daily Prod Range w/Daily DU OQ109 BP:68.62"
    # The second family carries no "Accum" word but is still an accumulator —
    # the LVL\d+ tier + /Trigger: + Daily Cons/Prod combination identifies it.
    accumulator_patterns = [
        r"\bACCUMULATOR\b", r"\bACCUM\b",
        r"\bDAILY\s+CONSUMER\s+ACCUM\b", r"\bDAILY\s+PRODUCER\s+ACCUM\b",
        r"\bNO\s+KO\b",
        r"\bLVL\d+\b",
        r"/TRIGGER\s*:",
        r"\bDAILY\s+(?:CONS|PROD|CONSUMER|PRODUCER)\b",
        r"\b(?:BP|OQ)\s*[:=]",
    ]
    if any(re.search(pat, upper_desc) for pat in accumulator_patterns):
        return "OTC Accumulator"

    # Listed options and exchange options.
    option_text = False
    for name in ["option_type", "option_type_raw", "call_put", "Call/Put", "C/P"]:
        try:
            if name in row.index:
                val = str(row.get(name) or "").strip().upper()
                if val in {"C", "P", "CALL", "PUT"}:
                    option_text = True
        except Exception:
            pass
    if option_text or re.search(r"\b(CALL|PUT|OPTION)\b", upper_desc):
        return "Option"

    # OTC/swap/structured rows. Examples from the Markets LLC PDFs include
    # Euro Swap rows and daily consumer/producer structured rows with Trigger/BP/OQ.
    otc_patterns = [
        r"\bSWAP\b", r"\bOTC\b",
        r"\bTRIGGER\b", r"\bDAILY\s+CONS\b", r"\bDAILY\s+PROD\b", r"\bLVL\d*\b",
        r"\bBP\s*[:=]", r"\bOQ\s*[:=]", r"\bOQ\d+\b",
    ]
    if any(re.search(pat, upper_desc) for pat in otc_patterns):
        return "OTC Swap"

    if raw_type:
        return raw_type
    return "Future"


def _ensure_position_type_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add/normalize raw `type` values for Open Positions before display/grouping."""
    if df is None or df.empty:
        return df
    out = df.copy()
    inferred = out.apply(_infer_position_type_row, axis=1)
    if "type" in out.columns:
        current = out["type"].apply(_normalize_position_type_value)
        missing = current.isna() | current.astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "unknown", "other"])
        out["type"] = current.where(~missing, inferred)
        # If the contract description identifies an accumulator, that is more
        # specific than a generic upstream/raw OTC Swap type.
        accumulator_mask = inferred.astype(str).str.upper().eq("OTC ACCUMULATOR")
        out["type"] = out["type"].where(~accumulator_mask, inferred)
    else:
        out["type"] = inferred
    # Keep a normalized technical grouping field as well as the user-facing raw
    # Type field.  Several grouping routines intentionally include
    # ``position_type`` so FX Spot / FX Forward / FX Swap / NDO rows do not
    # collapse together when they share the same currency pair and expiry.
    out["position_type"] = out["type"].apply(_normalize_position_type_value)
    return out

def _view_has_fx_rows(df: pd.DataFrame) -> bool:
    """Return True when a standardized view contains at least one FX row."""
    if df is None or df.empty:
        return False
    checks = []
    for col in ["Exchange", "exchange"]:
        if col in df.columns:
            checks.append(df[col].astype(str).str.upper().eq("FX"))
    for col in ["Product", "product"]:
        if col in df.columns:
            checks.append(df[col].astype(str).str.upper().str.startswith("FX"))
    for col in ["CCY 1", "CCY 2", "ccy_1", "ccy_2"]:
        if col in df.columns:
            checks.append(_series_nonblank(df[col]))
    if not checks:
        return False
    mask = checks[0]
    for chk in checks[1:]:
        mask = mask | chk
    return bool(mask.fillna(False).any())


def _view_has_option_rows(df: pd.DataFrame) -> bool:
    """Return True when a view contains real option rows.

    Some statement formats include Call/Put and strike columns even for futures,
    but those columns are blank/None for non-option rows. This helper lets the
    UI hide those option-only columns unless the uploaded PDF actually contains
    option positions.
    """
    if df is None or df.empty:
        return False

    checks = []
    for col in ["Call/Put", "option_type", "C/P", "cp"]:
        if col in df.columns:
            text = df[col].astype(str).str.strip().str.lower()
            checks.append(df[col].notna() & text.isin(["c", "p", "call", "put"]))

    # Do not treat a populated strikePrice by itself as an option signal.
    # OTC accumulators use strikePrice too, but they are not options unless
    # Call/Put or explicit option text is present.

    if "Contract Description" in df.columns:
        checks.append(df["Contract Description"].astype(str).str.upper().str.contains(r"\b(?:CALL|PUT|OPTION)\b", regex=True, na=False))
    if "contract_description" in df.columns:
        checks.append(df["contract_description"].astype(str).str.upper().str.contains(r"\b(?:CALL|PUT|OPTION)\b", regex=True, na=False))

    if not checks:
        return False
    mask = checks[0]
    for chk in checks[1:]:
        mask = mask | chk
    return bool(mask.fillna(False).any())


def _apply_conditional_position_columns(
    df: pd.DataFrame,
    drop_empty_expiry: bool = True,
    drop_empty_options: bool = True,
) -> pd.DataFrame:
    """Hide columns that should only appear when the uploaded PDFs contain data for them.

    For Open Positions, expiryDate should remain available in the Customize table
    list even when it is blank, so callers can pass drop_empty_expiry=False and
    hide it only from the default selected columns.

    Call/Put and strikePrice are option-only fields. If no option rows are present
    in the uploaded PDFs/current view, they are removed from Open Positions and
    Grouped Positions so futures-only / FX-only statements do not show empty
    option columns.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    out = out.drop(columns=["Settlement Date"], errors="ignore")
    fx_cols = ["CCY 1", "CCY 1 Amount", "CCY 2", "CCY 2 Amount"]
    if not _view_has_fx_rows(out):
        out = out.drop(columns=fx_cols, errors="ignore")
    if drop_empty_options and not _view_has_option_rows(out):
        # Hide Call/Put, Delta and NOV when there are no option rows. Keep strikePrice when it
        # contains OTC accumulator/structured-product strike levels.
        drop_cols = ["Call/Put", "Delta", "NOV"]
        if "strikePrice" in out.columns and not _series_nonblank(out["strikePrice"]).any():
            drop_cols.append("strikePrice")
        out = out.drop(columns=drop_cols, errors="ignore")
    elif "NOV" in out.columns and not _series_nonblank(out["NOV"]).any():
        out = out.drop(columns=["NOV"], errors="ignore")
    if "Trigger/Barrier" in out.columns and not _series_nonblank(out["Trigger/Barrier"]).any():
        out = out.drop(columns=["Trigger/Barrier"], errors="ignore")
    for otc_col in ["Ref Price", "Original Quantity"]:
        if otc_col in out.columns and not _series_nonblank(out[otc_col]).any():
            out = out.drop(columns=[otc_col], errors="ignore")
    if drop_empty_expiry and "expiryDate" in out.columns and not _series_nonblank(out["expiryDate"]).any():
        out = out.drop(columns=["expiryDate"], errors="ignore")
    return out

def _pdf_contract_description_for_view(df: pd.DataFrame) -> pd.Series:
    """Return the PDF contract-description text for display.

    In the StoneX PDF, the CONTRACT DESCRIPTION column includes the contract
    month/year plus product text (for example, "JUN 26 SFE SPI 200").
    Internally the parser stores month/year separately for many futures rows,
    so this reconstructs the PDF-facing value for Contract Description.
    """
    if df is None or df.empty:
        return pd.Series(dtype=object)

    desc = _first_existing(df, ["contract_description", "product", "product_name"], default=None)
    product = _first_existing(df, ["product", "product_name"], default=None)
    month = _first_existing(df, ["contract_month"], default=None)
    year = _first_existing(df, ["contract_year"], default=None)
    ref_month = _first_existing(df, ["ref_month", "Contract Month/Year"], default=None)
    exchange_series = _first_existing(df, ["exchange", "Exchange"], default=None)

    def clean(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        t = str(v).strip()
        return "" if t.lower() in {"nan", "none", "nat"} else t

    out = []
    for idx in df.index:
        d = clean(desc.loc[idx] if idx in desc.index else None)
        p = clean(product.loc[idx] if idx in product.index else None)
        if d.lower() == "multiple" or not d:
            d = p

        m = clean(month.loc[idx] if idx in month.index else None).upper()
        y = clean(year.loc[idx] if idx in year.index else None)
        r = clean(ref_month.loc[idx] if idx in ref_month.index else None).upper().replace("/", "-")
        e = clean(exchange_series.loc[idx] if idx in exchange_series.index else None).upper()

        if (not m or not y or m == "MULTIPLE" or y == "MULTIPLE") and re.match(r"^[A-Z]{3}-\d{2}$", r):
            m, y = r.split("-", 1)

        already_has_month = bool(re.match(r"^(CALL|PUT|C|P)?\s*[A-Z]{3}\s+\d{2}\b", d, re.I))
        if e == "FX" or d.upper().startswith("FX "):
            already_has_month = True
        if m and y and re.match(r"^[A-Z]{3}$", m) and re.match(r"^\d{2}$", y) and not already_has_month:
            d = f"{m} {y} {d}".strip()
        out.append(d or None)
    return pd.Series(out, index=df.index)

def _normalize_call_put_for_display(value):
    """Return Call/Put values for display; non-options should show None, not Other."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    text = str(value).strip()
    if text == "" or text.lower() in {"none", "nan", "nat", "other", "unknown"}:
        return None
    upper = text.upper()
    if upper in {"C", "CALL"}:
        return "Call"
    if upper in {"P", "PUT"}:
        return "Put"
    return text


def _normalize_call_put_series(series: pd.Series) -> pd.Series:
    return series.apply(_normalize_call_put_for_display)




def _position_type_from_row(raw_type=None, desc=None, exchange=None, product=None, option_type=None, source_section=None) -> str | None:
    """Classify a parsed open-position row into the user-facing Type column."""
    def clean(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        return " ".join(str(v).strip().upper().split())

    raw = clean(raw_type)
    desc_u = clean(desc)
    exchange_u = clean(exchange)
    product_u = clean(product)
    option_u = clean(option_type)
    source_u = clean(source_section)

    if raw in {"MULTIPLE", "MIXED"}:
        return "Multiple"
    if re.search(r"\b(ACCUMULATOR|ACCUM)\b", raw):
        return "OTC Accumulator"

    # Explicit FX table types from newer statements.
    if raw in {"FX FWD", "FWD", "FX FORWARD", "FORWARD"}:
        return "FX Forward"
    if raw in {"FX SPOT", "SPOT"}:
        return "FX Spot"
    if raw in {"FX NDF", "NDF FX", "NDF", "NON DELIVERABLE FORWARD", "NON-DELIVERABLE FORWARD", "NON DELIVERABLE FORWARDS", "NON-DELIVERABLE FORWARDS"}:
        return "FX NDF"
    if raw in {"FX SWAP", "FX SWP", "SWAP"} and (exchange_u == "FX" or product_u.startswith("FX")):
        return "FX Swap"

    # FX rows without an explicit type are forward-style open positions unless
    # the row itself names spot/swap/NDF.
    if exchange_u == "FX" or product_u.startswith("FX") or desc_u.startswith("FX") or "FX OPEN POSITIONS" in source_u:
        blob = " ".join([raw, desc_u, source_u])
        if re.search(r"\bNDF\b", blob):
            return "FX NDF"
        if re.search(r"\bSPOT\b", blob):
            return "FX Spot"
        if re.search(r"\bSWAP\b|\bSWP\b", blob):
            return "FX Swap"
        return "FX Forward"

    # OTC accumulator structures should be labelled separately from plain OTC
    # swaps. Some accumulators (e.g. "ICE Cotton LVL1 0.6625 /Trigger:0.7074
    # Daily Cons Range w/Daily DU OQ109 BP:68.62") never spell "Accum", so the
    # LVL\d+ tier / /Trigger: / Daily Cons|Prod / BP=/OQ= markers also qualify.
    if re.search(
        r"\b(ACCUMULATOR|ACCUM|NO KO|DAILY CONSUMER ACCUM|DAILY PRODUCER ACCUM|LVL\d+)\b"
        r"|/TRIGGER\s*:"
        r"|\bDAILY\s+(?:CONS|PROD|CONSUMER|PRODUCER)\b"
        r"|\b(?:BP|OQ)\s*[:=]",
        desc_u,
    ):
        return "OTC Accumulator"

    # Listed options.
    if option_u in {"CALL", "PUT", "C", "P"} or re.search(r"\b(CALL|PUT|OPTION|EURO OPTION|EUROPEAN OPTION)\b", desc_u):
        return "Option"

    # OTC commodity swaps / structured swaps.
    if re.search(r"\b(SWAP|TRIGGER|BARRIER|RANGE W/)\b", desc_u):
        return "OTC Swap"

    return "Future"


def _infer_position_type_series(df: pd.DataFrame) -> pd.Series:
    """Return the Type classification for each raw/standard/grouped position row."""
    if df is None or df.empty:
        return pd.Series(dtype=object)
    raw_type = _first_existing(df, ["position_type", "Type", "type", "contract_type", "instrument_type"], default=None)
    desc = _first_existing(df, ["contract_description", "Contract Description", "product", "Product"], default=None)
    exchange = _first_existing(df, ["exchange", "Exchange"], default=None)
    product = _first_existing(df, ["product", "product_name", "Product"], default=None)
    option_type = _first_existing(df, ["option_type", "option_type_raw", "Call/Put", "call_put"], default=None)
    source_section = _first_existing(df, ["source_section", "sourceSystem", "source_system"], default=None)
    return pd.Series(
        [
            _position_type_from_row(rt, d, e, p, o, s)
            for rt, d, e, p, o, s in zip(raw_type, desc, exchange, product, option_type, source_section)
        ],
        index=df.index,
    )



def _series_to_numeric_any(series: pd.Series) -> pd.Series:
    """Parse numeric values that may already be floats or statement strings like ($94.37)."""
    if series is None:
        return pd.Series(dtype=float)
    return series.apply(_num_any)


def _coalesced_open_trade_value(df: pd.DataFrame, candidates: list[str] | None = None) -> pd.Series:
    """Return the row-level open trade value used for OTE/NOV allocation.

    Business rule: when Market Value is available, it is the authoritative
    economic value for both non-option OTE and option NOV. Only fall back to
    existing OTE/open-trade-equity fields when a Market Value field is absent
    or blank for that row.
    """
    if df is None or df.empty:
        return pd.Series(dtype=float)
    candidates = candidates or [
        "Market Value",
        "market_value",
        "MarketValue",
        "market_value_signed",
        "open_trade_equity",
        "unrealized_pnl",
        "Unrealised PNL (OTE)",
        "nov",
        "NOV",
    ]
    result = pd.Series([None] * len(df), index=df.index, dtype=object)
    for col in candidates:
        if col not in df.columns:
            continue
        values = df[col].apply(_num_any)
        # Only fill blanks. A zero Market Value is a valid available value and
        # must not be overwritten by another field.
        result = result.where(~result.isna(), values)
    return pd.to_numeric(result, errors="coerce")

def _standard_option_row_mask(out: pd.DataFrame) -> pd.Series:
    """Identify option rows in the standardized position view."""
    if out is None or out.empty:
        return pd.Series(dtype=bool)
    mask = pd.Series(False, index=out.index)
    if "Type" in out.columns:
        mask = mask | out["Type"].astype(str).str.strip().str.upper().isin(["OPTION", "NDO"])
    if "Call/Put" in out.columns:
        cp = out["Call/Put"].astype(str).str.strip().str.upper()
        mask = mask | cp.isin(["CALL", "PUT", "C", "P"])
    if "Contract Description" in out.columns:
        desc = out["Contract Description"].astype(str).str.upper()
        mask = mask | desc.str.contains(r"\b(?:CALL|PUT|OPTION)\b", regex=True, na=False)
    return mask.fillna(False)


def _apply_trade_level_nov_ote_split(out: pd.DataFrame) -> pd.DataFrame:
    """Populate NOV/OTE using Market Value as the authoritative source.

    If Market Value is available on a row, use it directly:
      - options / NDO rows: Market Value -> NOV, OTE blank
      - non-options: Market Value -> Unrealised PNL (OTE), NOV blank

    Existing NOV/OTE values are only used when Market Value is absent or blank.
    """
    if out is None or out.empty:
        return out
    out = out.copy()
    if "NOV" not in out.columns:
        out["NOV"] = None
    if "Unrealised PNL (OTE)" not in out.columns:
        out["Unrealised PNL (OTE)"] = None
    if "Market Value" not in out.columns:
        out["Market Value"] = None

    option_mask = _standard_option_row_mask(out)
    market_value_numeric = _series_to_numeric_any(out["Market Value"])
    existing_nov_numeric = _series_to_numeric_any(out["NOV"])
    existing_ote_numeric = _series_to_numeric_any(out["Unrealised PNL (OTE)"])

    market_available = market_value_numeric.notna()

    # Options / NDOs: NOV is Market Value whenever present.
    option_nov = existing_nov_numeric.where(~market_available, market_value_numeric)
    out.loc[option_mask, "NOV"] = option_nov.loc[option_mask]
    out.loc[option_mask, "Unrealised PNL (OTE)"] = None

    # Non-options: OTE is Market Value whenever present.
    non_option_mask = ~option_mask
    non_option_ote = existing_ote_numeric.where(~market_available, market_value_numeric)
    out.loc[non_option_mask, "Unrealised PNL (OTE)"] = non_option_ote.loc[non_option_mask]
    out.loc[non_option_mask, "NOV"] = None
    return out

def standard_position_view_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Map raw or grouped position rows to the standard Positions schema requested by the user.

    Mapping assumptions:
    - Contract Description = the exact PDF contract description, reconstructed as month/year + product when needed.
    - Product = normalized product name used for grouping/risk views.
    - expiryDate = true option expiration date or FX value date where available.
    - settlementPrice = settlement/closing price where available; blank when the PDF has no row-level settlement price.
    - Trade ID = card from the PDF when available, falling back to trade_id/global_id.
    - Unrealised PNL (OTE) = signed market value when available.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=STANDARD_POSITION_COLUMNS)

    out_source = df.copy()

    # Ensure parsed metadata exists for raw rows.
    if "contract_description" in out_source.columns:
        parsed = out_source["contract_description"].apply(lambda x: pd.Series(parse_contract_product(x)))
        for col in parsed.columns:
            if col not in out_source.columns:
                out_source[col] = parsed[col]
            else:
                current = out_source[col]
                missing = current.isna() | (current.astype(str).str.strip() == "") | (current.astype(str).str.lower().isin(["none", "other", "unknown"])) | (current.astype(str).str.lower() == "unknown")
                out_source[col] = current.where(~missing, parsed[col])

    net_qty = _compute_net_qty_for_view(out_source)
    open_trade_value_numeric = _coalesced_open_trade_value(out_source)
    # Preserve/display Market Value as supplied by the statement/API and use it
    # as the source for P&L allocation when available.
    market_value_signed = open_trade_value_numeric
    market_value_display = _first_existing(out_source, ["market_value", "Market Value", "market_value_signed", "unrealized_pnl", "open_trade_equity"])
    nov_raw = _first_existing(out_source, ["nov", "NOV", "_nov_value"])
    realised_pnl = _first_existing(out_source, ["realised_pnl", "realized_pnl", "Realised PNL", "Realized PNL"])
    day_pnl = _first_existing(out_source, ["day_pnl", "day_pnl_amount", "Day PNL", "Day P&L"])

    contract_description = _pdf_contract_description_for_view(out_source)
    product = _first_existing(out_source, ["product", "product_name"])
    exchange = _first_existing(out_source, ["exchange"])
    out_source = _ensure_fx_position_columns(out_source)
    out_source = _ensure_otc_position_columns(out_source)
    out_source = _ensure_position_type_column(out_source)
    position_type = _first_existing(out_source, ["type", "Type", "position_type", "trade_type"], default=None)
    position_type = position_type.apply(_normalize_position_type_value)

    # Split Market Value into the correct risk buckets at the row level.
    # Options / NDOs feed NOV; non-options feed Unrealised PNL (OTE).
    try:
        option_mask = _raw_option_position_mask(out_source).reindex(out_source.index, fill_value=False)
    except Exception:
        option_mask = pd.Series(False, index=out_source.index)
    option_mask = option_mask | position_type.astype(str).str.strip().str.upper().isin(["OPTION", "NDO"])

    market_value_numeric = pd.Series(open_trade_value_numeric, index=out_source.index)
    nov_numeric = nov_raw.apply(_num_any) if hasattr(nov_raw, "apply") else pd.Series([_num_any(nov_raw)] * len(out_source), index=out_source.index)
    market_value_available = market_value_numeric.notna()
    nov = nov_numeric.where(~market_value_available, market_value_numeric.where(option_mask, None))
    market_value_signed = market_value_numeric.where(~option_mask, None)

    currency = _first_existing(out_source, ["currency", "unit"])
    trigger_barrier = _first_existing(out_source, ["trigger_barrier", "Trigger/Barrier", "triggerBarrier"], default=None)
    ref_price = _first_existing(out_source, ["ref_price", "Ref Price", "bp", "base_price"], default=None)
    original_quantity = _first_existing(out_source, ["original_quantity", "Original Quantity", "oq"], default=None)
    ccy_1 = _first_existing(out_source, ["ccy_1", "ccy1", "primary_currency", "primary_ccy"], default=None)
    ccy_1_amount = _first_existing(out_source, ["ccy_1_amount", "ccy1_amount", "primary_amount", "primary_amount_signed"], default=None)
    ccy_2 = _first_existing(out_source, ["ccy_2", "ccy2", "secondary_currency", "secondary_ccy"], default=None)
    ccy_2_amount = _first_existing(out_source, ["ccy_2_amount", "ccy2_amount", "secondary_amount", "secondary_amount_signed"], default=None)
    expiry_date = _first_existing(out_source, ["expiryDate", "expiry_date", "expiration_date", "value_date", "delivery_date", "contract_date", "end_date"])
    end_date_value = _first_existing(out_source, ["end_date", "End Date"], default=None)
    settlement_price = _first_existing(out_source, ["settlementPrice", "settlement_price", "market_price", "closing_price"])
    trade_id = _first_existing(out_source, ["card", "trade_id", "global_id", "position_id"])
    account_number = _first_existing(out_source, ["account_number"])
    broker_code = _first_existing(out_source, ["broker_code", "salesman", "Broker Code"], default=None)
    # In the StoneX OPEN POSITIONS section, the first TRADE column is the
    # trade date. It is not a contract date. Contract Date should come from
    # a true delivery/contract date when the statement provides one, such as
    # the LME "Delivery / Product" column (delivery_date).
    trade_date = _first_existing(out_source, ["trade_date_iso", "trade_date"])
    last_update = _first_existing(out_source, ["last_update", "statement_date"])
    # Settlement date is retained internally for raw audit tables only.
    settlement_date_raw = _first_existing(
        out_source,
        ["settlement_date", "value_date", "contract_date", "delivery_date", "expiry_date", "end_date"],
        default=None,
    )
    settlement_date = settlement_date_raw.apply(_normalize_any_date) if hasattr(settlement_date_raw, "apply") else settlement_date_raw
    month_year = _combine_month_year(out_source)
    trade_price = _first_existing(out_source, ["trade_price", "rate", "avg_trade_price", "price"])
    avg_fill_price = _first_existing(out_source, ["avg_trade_price", "trade_price", "rate", "price"])
    delta = _first_existing(out_source, ["delta", "Delta"])
    call_put = _first_existing(out_source, ["option_type", "option_type_raw", "call_put"])
    call_put = _normalize_call_put_series(call_put)
    strike = _first_existing(out_source, ["strike", "strikePrice"])

    # Correct NOV / OTE allocation at the trade-row level.
    # Market Value is authoritative whenever present: options / NDOs use it as
    # NOV, and non-options use it as Unrealised PNL (OTE).
    try:
        desc_upper = contract_description.astype(str).str.upper()
    except Exception:
        desc_upper = pd.Series([""] * len(out_source), index=out_source.index)
    type_upper = position_type.astype(str).str.strip().str.upper()
    call_put_nonblank = call_put.notna() & ~call_put.astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "other", "unknown"])
    option_row_mask = (
        type_upper.isin(["OPTION", "NDO"])
        | call_put_nonblank
        | desc_upper.str.contains(r"\b(?:CALL|PUT|OPTION)\b", regex=True, na=False)
    )
    market_value_num = pd.Series(open_trade_value_numeric, index=out_source.index)
    market_value_available = market_value_num.notna()
    nov = nov.where(~(option_row_mask & market_value_available), market_value_num)
    # Keep grouped rows with a real precomputed NOV even if Type is Multiple, but
    # blank NOV for rows/groups without an option component.
    nov_has_value = nov.apply(_num_any).notna()
    nov = nov.where(option_row_mask | nov_has_value, None)
    unrealised_ote_display = market_value_num.where(~option_row_mask & market_value_available, market_value_signed.where(~option_row_mask, None))

    source_system = _first_existing(out_source, ["sourceSystem", "source_system"])
    if source_system.isna().all() if hasattr(source_system, 'isna') else False:
        source_system = pd.Series(["StoneX"] * len(out_source), index=out_source.index)
    direction = _first_existing(out_source, ["direction", "side"])
    if direction.isna().all() if hasattr(direction, 'isna') else False:
        direction = _direction_from_net_qty(net_qty)

    global_id = _first_existing(out_source, ["global_id", "globalId"])

    out = pd.DataFrame({
        "Global ID": global_id,
        "Contract Description": contract_description,
        "Product": product,
        "Type": position_type,
        "Exchange": exchange,
        "Currency": currency,
        "Trigger/Barrier": trigger_barrier,
        "Ref Price": ref_price,
        "Original Quantity": original_quantity,
        "CCY 1": ccy_1,
        "CCY 1 Amount": ccy_1_amount,
        "CCY 2": ccy_2,
        "CCY 2 Amount": ccy_2_amount,
        "expiryDate": expiry_date,
        "End Date": end_date_value,
        "settlementPrice": settlement_price,
        "Trade ID": trade_id,
        "Trade Date": trade_date,
        "Account Number": account_number,
        "Broker Code": broker_code,
        "Net Quantity": net_qty,
        "Last Update": last_update,
        "Settlement Date": settlement_date,
        "Contract Month/Year": month_year,
        "Trade Price": trade_price,
        "Avg Fill Price": avg_fill_price,
        "Delta": delta,
        "Call/Put": call_put,
        "strikePrice": strike,
        "sourceSystem": source_system,
        "Direction": direction,
        "NOV": nov,
        "Unrealised PNL (OTE)": unrealised_ote_display,
        "Realised PNL": realised_pnl,
        "Day PNL": day_pnl,
        "Market Value": market_value_display,
    })
    out = _apply_trade_level_nov_ote_split(out)
    if "Avg Fill Price" in out.columns:
        out["Avg Fill Price"] = pd.to_numeric(out["Avg Fill Price"], errors="coerce").round(2)
    if "expiryDate" in out.columns:
        out["expiryDate"] = out["expiryDate"].apply(_normalize_any_date)
    if "End Date" in out.columns:
        out["End Date"] = out["End Date"].apply(_normalize_any_date)
    return out[STANDARD_POSITION_COLUMNS]


def open_positions_standard_view(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Open Positions display schema.

    User-facing changes:
    - Net Quantity is renamed to Quantity.
    - Contract Description is shown as the PDF-facing instrument detail.
    - Settlement Date is removed from user-facing views.
    - Market Value is hidden/removed from this view.
    - expiryDate is shown when delivery/value/expiration dates are available.
    - Unrealised PNL (OTE) remains available for position P&L.
    """
    df = standard_position_view_from_df(tables.get("Open Positions", pd.DataFrame()))
    if df is None or df.empty:
        return pd.DataFrame(columns=OPEN_POSITION_COLUMNS)
    df = df.rename(columns={"Net Quantity": "Quantity"})
    if "Trade Price" not in df.columns and "Avg Fill Price" in df.columns:
        df["Trade Price"] = df["Avg Fill Price"]
    elif "Trade Price" in df.columns and "Avg Fill Price" in df.columns:
        df["Trade Price"] = df["Trade Price"].where(df["Trade Price"].notna(), df["Avg Fill Price"])
    df = df.drop(columns=["Avg Fill Price", "Market Value", "Direction", "Realised PNL", "Day PNL", "Settlement Date"], errors="ignore")
    # Keep expiryDate in the Open Positions column list even when the current PDF
    # has no expiry/value/end date; the app hides it by default when blank.
    df = _apply_conditional_position_columns(df, drop_empty_expiry=False)
    # Keep the requested display order and preserve any future extra columns after it.
    ordered = [c for c in OPEN_POSITION_COLUMNS if c in df.columns]
    ordered += [c for c in df.columns if c not in ordered]
    return df[ordered]



def _blank_zero_when_no_component(values: pd.Series, counts: pd.Series) -> pd.Series:
    """Show blanks for rows without that component; keep numeric zero when the component exists."""
    out = pd.to_numeric(values, errors="coerce")
    component_counts = pd.to_numeric(counts, errors="coerce").fillna(0)
    return out.where(component_counts > 0, None)


def _add_grouped_risk_pnl_columns(df: pd.DataFrame, grouped_raw: pd.DataFrame | None = None) -> pd.DataFrame:
    """Add grouped risk P&L columns.

    - NOV is populated with option OTE only.
    - Unrealised PNL (OTE) is populated with non-option OTE only.
    Realised PNL and Day PNL are intentionally excluded — they live in the
    dedicated Realised PNL tab, not in position views.
    """
    if df is None:
        df = pd.DataFrame()
    out = df.copy()
    if out.empty:
        if "NOV" not in out.columns:
            out["NOV"] = pd.Series(dtype=float)
        return out

    raw = grouped_raw if grouped_raw is not None else pd.DataFrame(index=out.index)

    if raw is not None and not raw.empty and ("_nov_value" in raw.columns or "nov" in raw.columns or "NOV" in raw.columns):
        option_counts = raw["_option_position_rows"] if "_option_position_rows" in raw.columns else raw.get("option_position_rows", pd.Series([1] * len(out), index=out.index))
        non_option_counts = raw["_non_option_position_rows"] if "_non_option_position_rows" in raw.columns else raw.get("non_option_position_rows", pd.Series([1] * len(out), index=out.index))
        # Align by position; grouped_raw and out are created from the same grouped frame.
        option_counts = pd.Series(option_counts.to_numpy(), index=out.index)
        non_option_counts = pd.Series(non_option_counts.to_numpy(), index=out.index)

        nov_source_col = "_nov_value" if "_nov_value" in raw.columns else ("nov" if "nov" in raw.columns else "NOV")
        nov_values = pd.Series(raw[nov_source_col].to_numpy(), index=out.index)

        if "_ote_non_option_value" in raw.columns:
            ote_source = raw["_ote_non_option_value"]
        elif "market_value" in raw.columns:
            # In grouped rows, market_value has already been renamed from the
            # non-option OTE component. The option component is stored separately
            # in nov/NOV, so do not reclassify market_value by Call/Put.
            ote_source = raw["market_value"]
        elif "Unrealised PNL (OTE)" in raw.columns:
            ote_source = raw["Unrealised PNL (OTE)"]
        else:
            ote_source = pd.Series([None] * len(out), index=out.index)
        ote_non_option_values = pd.Series(ote_source.to_numpy(), index=out.index)

        out["NOV"] = _blank_zero_when_no_component(nov_values, option_counts)
        out["Unrealised PNL (OTE)"] = _blank_zero_when_no_component(ote_non_option_values, non_option_counts)
    else:
        call_blank = out.get("Call/Put", pd.Series([None] * len(out), index=out.index)).apply(lambda x: _normalize_call_put_for_display(x) is None)
        strike_blank = out.get("strikePrice", pd.Series([None] * len(out), index=out.index)).apply(lambda x: _normalize_call_put_for_display(x) is None)
        strike_blank = out.get("strikePrice", pd.Series([None] * len(out), index=out.index)).isna() | (out.get("strikePrice", pd.Series([None] * len(out), index=out.index)).astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "multiple", "other"]))
        is_option = ~(call_blank & strike_blank)
        existing_ote = pd.to_numeric(out.get("Unrealised PNL (OTE)", pd.Series([None] * len(out), index=out.index)), errors="coerce")
        out["NOV"] = existing_ote.where(is_option, None)
        out["Unrealised PNL (OTE)"] = existing_ote.where(~is_option, None)

    return out



def _apply_grouped_nov_ote_display_rules(df: pd.DataFrame) -> pd.DataFrame:
    """Finalize grouped-position P&L columns for display.

    NOV receives option OTE. Unrealised PNL (OTE) remains the futures/forward OTE.
    If a grouped row contains only options, OTE is blank and NOV is populated.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    if "NOV" not in out.columns:
        out["NOV"] = None
    if "Unrealised PNL (OTE)" not in out.columns:
        out["Unrealised PNL (OTE)"] = None

    nov_num = pd.to_numeric(out["NOV"], errors="coerce")
    ote_num = pd.to_numeric(out["Unrealised PNL (OTE)"], errors="coerce")

    # Blank zero NOV values so futures/forwards do not display 0 as an option value.
    out.loc[nov_num.fillna(0) == 0, "NOV"] = None

    # For pure option rows, OTE was moved into NOV before grouping; leave OTE blank.
    nov_num = pd.to_numeric(out["NOV"], errors="coerce")
    out.loc[nov_num.notna() & (ote_num.fillna(0) == 0), "Unrealised PNL (OTE)"] = None
    return out


def _standardized_realized_pnl_for_grouping(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    detail = realized_pnl_summary(tables)
    if detail is None or detail.empty or "realized_pnl" not in detail.columns:
        return pd.DataFrame()
    # Summary-level MTD/YTD lines are not product-specific, so do not merge them into product groups.
    if "pnl_view" in detail.columns:
        detail = detail[detail["pnl_view"].astype(str).str.lower() != "summary"].copy()
    if detail.empty:
        return pd.DataFrame()

    src = detail.copy()
    if "contract_description" in src.columns:
        parsed = src["contract_description"].apply(lambda x: pd.Series(parse_contract_product(x)))
        for col in parsed.columns:
            if col not in src.columns:
                src[col] = parsed[col]
            else:
                current = src[col]
                missing = current.isna() | (current.astype(str).str.strip() == "") | current.astype(str).str.lower().isin(["none", "other", "unknown"])
                src[col] = current.where(~missing, parsed[col])

    out = pd.DataFrame(index=src.index)
    out["Product"] = _first_existing(src, ["product", "product_name"])
    out["Contract Month/Year"] = _first_existing(src, ["ref_month", "Contract Month/Year"])
    if out["Contract Month/Year"].isna().all():
        out["Contract Month/Year"] = _combine_month_year(src)
    out["Account Number"] = _first_existing(src, ["account_number", "Account Number"])
    out["Call/Put"] = _normalize_call_put_series(_first_existing(src, ["option_type", "option_type_raw", "call_put", "Call/Put"]))
    out["strikePrice"] = _first_existing(src, ["strike", "strikePrice"])
    out["Realised PNL"] = pd.to_numeric(src["realized_pnl"], errors="coerce")
    return out.dropna(subset=["Realised PNL"], how="all")


def add_grouped_position_pnl_columns(grouped_df: pd.DataFrame, tables: Dict[str, pd.DataFrame], selected_preset: str | None = None) -> pd.DataFrame:
    """Apply NOV/OTE display rules to a grouped-position view.

    Realised PNL and Day PNL are no longer shown in position views — they have a
    dedicated Realised PNL tab.
    """
    if grouped_df is None or grouped_df.empty:
        return grouped_df
    out = grouped_df.copy()
    out = _apply_grouped_nov_ote_display_rules(out)
    return out


def prepare_grouped_positions_display(grouped_df: pd.DataFrame, tables: Dict[str, pd.DataFrame] | None = None, selected_preset: str | None = None, drop_option_columns: bool = False) -> pd.DataFrame:
    """Apply final grouped-position layout rules used by the Streamlit app and Excel export.

    Account Number is intentionally shown only in the Account Grouping view.
    Product-only and Product + Contract Month/Year are market/risk views, so
    account details are removed from those grouped summaries while remaining
    available in Open Positions and in Account Grouping drill-down.
    """
    if grouped_df is None or grouped_df.empty:
        return grouped_df
    out = grouped_df.copy()
    out = _apply_grouped_nov_ote_display_rules(out)
    out = _apply_grouped_nov_ote_display_rules(out)

    preset = str(selected_preset or "")
    # Drop all trade-level detail columns that don't belong in a grouped summary.
    # Note: Delta is intentionally kept — it is an options-specific risk metric and belongs
    # in the grouped view alongside Call/Put and strikePrice. It will be blank until the
    # parser extracts delta values from the statement, but the column structure is correct.
    out = out.drop(columns=[
        "Trade Price", "Ref Price", "Original Quantity",
        # These survive from STANDARD_POSITION_COLUMNS but are trade-level noise in a grouped view:
        "Global ID", "End Date", "Trade Date",
        "Broker Code", "Last Update", "sourceSystem",
    ], errors="ignore")

    if preset in {"Product", "Product + Contract Month/Year"}:
        out = out.drop(columns=["Account Number"], errors="ignore")

    # Avoid showing 0.00 currency amounts on non-FX grouped rows.
    for ccy_col, amt_col in [("CCY 1", "CCY 1 Amount"), ("CCY 2", "CCY 2 Amount")]:
        if ccy_col in out.columns and amt_col in out.columns:
            ccy_blank = out[ccy_col].isna() | out[ccy_col].astype(str).str.strip().str.lower().isin(["", "none", "nan", "nat", "multiple"])
            out.loc[ccy_blank, amt_col] = None

    if preset == "Product":
        # Product grouping is a pure product-level aggregation. Do not display
        # contract month/year or expiry/value date fields that are not grouping keys.
        out = out.drop(columns=["Contract Month/Year", "expiryDate", "Settlement Date"], errors="ignore")

    if drop_option_columns:
        # Product grouping intentionally hides option detail columns, but NOV must
        # remain visible when the product contains option value. Do not let the
        # conditional option-column helper infer "no options" only because Call/Put
        # and strikePrice were removed for display.
        out = out.drop(columns=["Call/Put", "strikePrice"], errors="ignore")
    out = _apply_conditional_position_columns(out, drop_empty_options=not drop_option_columns)
    # Ensure every column in GROUPED_POSITION_COLUMNS is always present in the
    # dataframe so the customise-columns picker shows the full list regardless of
    # whether FX / option / blank-expiry logic conditionally removed them above.
    # Columns that were dropped by _apply_conditional_position_columns are re-added
    # as empty (None) columns — they will not be pre-selected by default but remain
    # available for the user to toggle on.
    for col in GROUPED_POSITION_COLUMNS:
        if col not in out.columns:
            out[col] = None

    ordered = [c for c in GROUPED_POSITION_COLUMNS if c in out.columns]
    ordered += [c for c in out.columns if c not in ordered]
    return out[ordered]

def grouped_positions_standard_view(tables: Dict[str, pd.DataFrame], group_cols: list[str] | None = None) -> pd.DataFrame:
    """Return grouped positions in the standard schema.

    Avg Fill Price is weighted by absolute Net Quantity inside _group_prepared_positions.
    """
    grouped = grouped_positions_custom(tables, group_cols)
    df = standard_position_view_from_df(grouped)
    df = _add_grouped_risk_pnl_columns(df, grouped)
    df = df.drop(columns=["Market Value", "Trade ID", "Contract Description", "Direction", "Contract Date", "Settlement Date", "Trade Price", "Ref Price", "Original Quantity"], errors="ignore")
    return _apply_conditional_position_columns(df)


def grouped_positions_product_month_auto(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Product + month view that automatically adds strike/call-put only for options.

    Futures/forwards rows group by product + contract month/year.
    Option rows group by product + contract month/year + call/put + strike.
    FX rows group by product + currencies + value date; currency amounts are aggregated and trade price is hidden from grouped views.
    """
    df, price_col, mv_col = _prepared_positions_for_grouping(tables)
    if df is None or df.empty:
        return pd.DataFrame()
    df = _ensure_fx_position_columns(df.copy())
    fx_mask = _fx_position_mask(df)
    frames: list[pd.DataFrame] = []

    non_fx = df.loc[~fx_mask].copy()
    if not non_fx.empty:
        option_signal = pd.Series(False, index=non_fx.index)
        if "option_type" in non_fx.columns:
            option_signal = option_signal | non_fx["option_type"].astype(str).str.lower().isin(["call", "put"])
        if "option_type_raw" in non_fx.columns:
            option_signal = option_signal | non_fx["option_type_raw"].astype(str).str.lower().isin(["call", "put"])

        if "contract_description" in non_fx.columns:
            option_signal = option_signal | non_fx["contract_description"].astype(str).str.upper().str.contains(r"\b(?:CALL|PUT|OPTION)\b", regex=True, na=False)

        if "option_type" in non_fx.columns:
            non_fx["option_type"] = non_fx["option_type"].where(option_signal, None)
        elif "option_type_raw" in non_fx.columns:
            non_fx["option_type"] = non_fx["option_type_raw"].where(option_signal, None)
        else:
            non_fx["option_type"] = None

        if "strike" in non_fx.columns:
            non_fx["strike"] = non_fx["strike"].where(option_signal, None)
        else:
            non_fx["strike"] = None

        product_month_cols = ["product", "ref_month", "option_type", "strike"]
        if "trigger_barrier" in non_fx.columns and _series_nonblank(non_fx["trigger_barrier"]).any():
            product_month_cols.append("trigger_barrier")
        frames.append(_group_prepared_positions(non_fx, _expiry_aware_group_cols(product_month_cols, non_fx), price_col, mv_col))

    fx = df.loc[fx_mask].copy()
    if not fx.empty:
        frames.append(_group_prepared_positions(fx, _fx_group_cols_for_base(None, mode="product_month"), price_col, mv_col))

    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)

def grouped_positions_product_month_auto_standard_view(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Standard-schema view for the automatic futures/options grouping."""
    grouped = grouped_positions_product_month_auto(tables)
    df = standard_position_view_from_df(grouped)
    df = _add_grouped_risk_pnl_columns(df, grouped)
    df = df.drop(columns=["Market Value", "Trade ID", "Contract Description", "Direction", "Contract Date", "Settlement Date", "Trade Price", "Ref Price", "Original Quantity"], errors="ignore")
    return _apply_conditional_position_columns(df)


def merge_extracted_tables(tables_list: list[Dict[str, pd.DataFrame]], position_mode: str = "append_all") -> Dict[str, pd.DataFrame]:
    """Merge extracted tables from multiple PDFs.

    position_mode:
      - append_all: keep every open-position row from every PDF. This is now the only UI behavior.
    """
    sheet_names = [
        "Executed Trades", "Purchase & Sale", "Closed Positions", "Receives Delivers", "Journal Entries",
        "Realized Gain and Loss", "Realized PNL Summary", "Open Positions", "Notes", "Exceptions"
    ]
    merged: Dict[str, pd.DataFrame] = {}
    for name in sheet_names:
        frames = []
        for i, tables in enumerate(tables_list, start=1):
            df = tables.get(name, pd.DataFrame())
            if df is not None and not df.empty:
                temp = df.copy()
                if "source_pdf_index" not in temp.columns:
                    temp["source_pdf_index"] = i
                frames.append(temp)
        merged[name] = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()

    # De-duplicate trades where the same PDF/statement is uploaded twice.
    for name in ["Executed Trades", "Purchase & Sale", "Closed Positions", "Receives Delivers", "Journal Entries", "Realized Gain and Loss"]:
        df = merged.get(name, pd.DataFrame())
        if df.empty:
            continue
        keys = [c for c in ["account_number", "statement_date", "trade_date", "trade_date_iso", "trade_id", "global_id", "contract_description", "ref_month", "quantity", "trade_price", "price"] if c in df.columns]
        if keys:
            merged[name] = df.drop_duplicates(subset=keys, keep="first").reset_index(drop=True)

    pos = merged.get("Open Positions", pd.DataFrame())
    if pos is not None and not pos.empty:
        # Open-position rows are card/trade-line level. Do NOT de-duplicate only on
        # product/qty/price/value because statements can legitimately contain many
        # separate cards with identical qty/price/value. The rubber/SCM statement is
        # a good example: many one-lot rows at the same price.
        #
        # This strict key removes true duplicate uploads/duplicate extracted lines,
        # while preserving unique cards/rows on continuation pages.
        strict_keys = [c for c in [
            "account_number", "statement_date", "trade_date", "trade_date_iso",
            "card", "account_type", "contract_month", "contract_year",
            "contract_description", "ref_month", "delivery_date", "settlement_date",
            "long", "short", "quantity", "trade_price", "price",
            "currency", "primary_currency", "primary_amount", "secondary_currency", "secondary_amount",
            "ccy_1", "ccy_1_amount", "ccy_2", "ccy_2_amount",
            "market_value", "market_value_signed", "drcr",
            "page", "source_line"
        ] if c in pos.columns]
        if strict_keys:
            merged["Open Positions"] = pos.drop_duplicates(subset=strict_keys, keep="first").reset_index(drop=True)
            pos = merged["Open Positions"]

    # Add a merge summary/notes sheet.
    note_rows = []
    for i, tables in enumerate(tables_list, start=1):
        # The app adds source_pdf_name to tables in app.py; fall back to index.
        summary = tables.get("Summary", pd.DataFrame())
        source_name = None
        for df in tables.values():
            if isinstance(df, pd.DataFrame) and not df.empty and "source_pdf" in df.columns:
                source_name = str(df["source_pdf"].iloc[0])
                break
        note_rows.append({"source_pdf_index": i, "source_pdf": source_name or f"PDF {i}", "position_mode": position_mode})
    merged["Merge Notes"] = pd.DataFrame(note_rows)
    merged["Summary"] = build_summary(merged)
    return merged


def realized_pnl_summary(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Absolute-basic realized PNL view.

    Uses row-level Realized Gain and Loss / Purchase & Sale where available.
    If the statement only has a summary-level realized P&L section, includes MTD/YTD rows.
    """
    frames = []
    realized = tables.get("Realized Gain and Loss", pd.DataFrame())
    if realized is not None and not realized.empty:
        df = realized.copy()
        df["source_sheet"] = "Realized Gain and Loss"
        if "cash_flow" in df.columns:
            df["realized_pnl"] = df["cash_flow"].apply(_num_any)
        elif "amount_signed" in df.columns:
            df["realized_pnl"] = df["amount_signed"].apply(_num_any)
        df["pnl_view"] = "Detail"
        frames.append(df)

    closed = tables.get("Closed Positions", pd.DataFrame())
    if closed is not None and not closed.empty:
        df = closed.copy()
        df["source_sheet"] = "Closed Positions"
        if "realized_pnl" not in df.columns and "amount_signed" in df.columns:
            df["realized_pnl"] = df["amount_signed"].apply(_num_any)
        df["pnl_view"] = "Closed Position Detail"
        frames.append(df)

    ps = tables.get("Purchase & Sale", pd.DataFrame())
    if ps is not None and not ps.empty:
        df = ps.copy()
        df["source_sheet"] = "Purchase & Sale"
        if "amount_signed" in df.columns:
            df["realized_pnl"] = df["amount_signed"].apply(_num_any)
        elif "amount" in df.columns:
            drcr_series = df["drcr"] if "drcr" in df.columns else [None] * len(df)
            df["realized_pnl"] = [_signed(a, d) for a, d in zip(df["amount"], drcr_series)]
        # Raw P&S rows are useful audit detail but are not the closed-position total rows.
        df["pnl_view"] = "P&S Trade Detail"
        frames.append(df)

    summary = tables.get("Realized PNL Summary", pd.DataFrame())
    if summary is not None and not summary.empty:
        df = summary.copy()
        df["source_sheet"] = "Statement Summary"
        df["pnl_view"] = "Summary"
        # Use MTD as the primary realized_pnl displayed in total metric.
        if "mtd_realized_pnl" in df.columns:
            df["realized_pnl"] = df["mtd_realized_pnl"].apply(_num_any)
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    detail = pd.concat(frames, ignore_index=True, sort=False)
    if "contract_description" in detail.columns:
        parsed = detail["contract_description"].apply(parse_contract_product).apply(pd.Series)
        for col in parsed.columns:
            if col not in detail.columns or detail[col].isna().all():
                detail[col] = parsed[col]

    wanted = [
        "pnl_view", "statement_date", "account_number", "currency",
        "mtd_realized_pnl", "ytd_realized_pnl", "realized_pnl",
        "trade_date", "trade_date_iso", "close_date", "trade_id", "contract_description",
        "product", "exchange", "ref_month", "option_type", "strike",
        "quantity", "long", "short", "trade_price", "price", "close_price", "cash_flow", "amount_signed",
        "source_sheet", "source_section", "source_pdf", "page", "source_line"
    ]
    cols = [c for c in wanted if c in detail.columns]
    return detail[cols].copy()


def closed_positions_standard_view(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Normalise closed-position data to the same display-column format as
    open_positions_standard_view.

    Sources (merged):
    - tables["Purchase & Sale"]  — individual closed trade rows (date, price, qty per leg)
    - tables["Closed Positions"] — GROSS PROFIT OR LOSS summary rows (net realised P&L per group)

    Individual P&S rows carry trade-level detail but no per-row P&L.
    The GROSS PROFIT OR LOSS row carries the net P&L for the group.
    Both are included so the tab shows the 3 closed trades AND the net result.
    """
    ps = tables.get("Purchase & Sale", pd.DataFrame())
    closed = tables.get("Closed Positions", pd.DataFrame())

    frames = []
    if ps is not None and not ps.empty:
        df = ps.copy()
        df["_row_type"] = "Trade"
        frames.append(df)
    if closed is not None and not closed.empty:
        df = closed.copy()
        df["_row_type"] = "Net P&L"
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    src = pd.concat(frames, ignore_index=True, sort=False)

    # Fill in product metadata from contract_description where columns are absent / blank.
    if "contract_description" in src.columns:
        parsed = src["contract_description"].apply(
            lambda x: pd.Series(parse_contract_product(str(x) if x else ""))
        )
        for col in parsed.columns:
            if col not in src.columns or src[col].isna().all():
                src[col] = parsed[col]
            else:
                missing = (
                    src[col].isna()
                    | (src[col].astype(str).str.strip() == "")
                    | src[col].astype(str).str.lower().isin(["none", "other", "unknown"])
                )
                src[col] = src[col].where(~missing, parsed[col])

    out = pd.DataFrame(index=src.index)
    out["Product"]              = _first_existing(src, ["product", "product_name"])

    # "Type" mirrors the Aggregated Positions column: Options if the row has an option_type,
    # otherwise Futures (closed FX/swap rows show as FX via Exchange).
    raw_option_type = _first_existing(src, ["option_type", "option_type_raw"])
    has_option = raw_option_type.notna() & (raw_option_type.astype(str).str.strip() != "")
    out["Type"]                 = has_option.map({True: "Options", False: "Futures"})

    out["Exchange"]             = _first_existing(src, ["exchange", "Exchange"])
    out["Contract Month/Year"]  = _first_existing(src, ["ref_month", "Contract Month/Year"])
    if out["Contract Month/Year"].isna().all():
        out["Contract Month/Year"] = _combine_month_year(src)
    out["expiryDate"]           = _first_existing(src, ["close_date", "expiryDate", "expiry_date"])
    out["Trigger/Barrier"]      = _first_existing(src, ["trigger_barrier", "Trigger/Barrier"], default=None)
    out["CCY 1"]                = _first_existing(src, ["ccy_1", "primary_currency"], default=None)
    out["CCY 1 Amount"]         = _first_existing(src, ["ccy_1_amount"], default=None)
    out["CCY 2"]                = _first_existing(src, ["ccy_2", "secondary_currency"], default=None)
    out["CCY 2 Amount"]         = _first_existing(src, ["ccy_2_amount"], default=None)
    out["Call/Put"]             = _normalize_call_put_series(raw_option_type)
    out["strikePrice"]          = pd.to_numeric(
        _first_existing(src, ["strike", "strikePrice"]), errors="coerce"
    )
    out["Account Number"]       = _first_existing(src, ["account_number"])

    # Quantity columns — Net P&L rows carry Long+Short (closed volume); Trade rows carry
    # the individual leg quantity.  Map to the same "Net Quantity" column name as Aggregated
    # Positions so the two views are structurally identical.
    row_type  = src.get("_row_type", pd.Series("Trade", index=src.index))
    trade_qty = pd.to_numeric(_first_existing(src, ["quantity"]), errors="coerce")
    long_qty  = pd.to_numeric(_first_existing(src, ["long"]),     errors="coerce")
    short_qty = pd.to_numeric(_first_existing(src, ["short"]),    errors="coerce")
    is_net_row = row_type == "Net P&L"
    # Net Quantity: for Net P&L rows show Long (total contracts closed); for Trade rows show
    # the individual leg quantity.
    out["Net Quantity"]         = long_qty.where(is_net_row, trade_qty)
    # Keep Long/Short separately so users can cross-check.
    out["Long"]                 = long_qty.where(is_net_row, None)
    out["Short"]                = short_qty.where(is_net_row, None)

    close_price = pd.to_numeric(
        _first_existing(src, ["close_price", "trade_price", "price"]), errors="coerce"
    )
    out["Avg Fill Price"]       = close_price   # mirrors Aggregated Positions column name
    out["settlementPrice"]      = close_price   # close price is the best proxy for settlement

    # NOV is not applicable for closed positions — leave blank so the column exists but is empty.
    out["NOV"]                  = None

    # Realised PNL — this replaces "Unrealised PNL (OTE)" in the same visual position.
    out["Realised PNL"]         = pd.to_numeric(
        _first_existing(src, ["realized_pnl", "amount_signed"]), errors="coerce"
    )
    # Debit/Credit indicator from the statement (DR = debit / loss, CR = credit / gain).
    out["Debit/Credit"]         = _first_existing(src, ["drcr"], default=None)

    # Supplementary columns (not in GROUPED_POSITION_COLUMNS but useful for drill-down).
    out["Contract Description"] = _first_existing(src, ["contract_description"])
    out["Currency"]             = _first_existing(src, ["currency", "unit"])
    out["Trade Date"]           = _first_existing(src, ["trade_date_iso", "trade_date"])
    out["Close Date"]           = _first_existing(src, ["close_date"])
    out["Row Type"]             = row_type
    out["source_section"]       = _first_existing(src, ["source_section"], default="Purchase & Sale")

    return _apply_conditional_position_columns(out)


def grouped_realized_pnl_view(
    tables: Dict[str, pd.DataFrame],
    group_cols: list[str] | None,
    mode: str = "custom",
) -> pd.DataFrame:
    """Aggregate closed positions by the given keys, summing Realised PNL.

    Uses the same column normalisation as closed_positions_standard_view and
    applies _apply_conditional_position_columns so CCY / option visibility
    mirrors the Aggregated Positions view.

    mode="auto_futures_options" uses product + exchange + ref_month and also
    adds option_type + strike when option rows are present, matching the
    Product + Contract Month/Year preset in Aggregated Positions.
    """
    df = closed_positions_standard_view(tables)
    if df is None or df.empty:
        return pd.DataFrame()

    display_group_col_map = {
        "product":        "Product",
        "exchange":       "Exchange",
        "ref_month":      "Contract Month/Year",
        "account_number": "Account Number",
        "trigger_barrier":"Trigger/Barrier",
        "option_type":    "Call/Put",
        "strike":         "strikePrice",
    }

    # auto_futures_options: mirror the Product + Contract Month/Year preset —
    # group by product + exchange + ref_month, add option cols when options present.
    if mode == "auto_futures_options":
        base = ["product", "exchange", "ref_month"]
        has_options = (
            "Call/Put" in df.columns
            and df["Call/Put"].notna().any()
            and df["Call/Put"].astype(str).str.strip().str.lower().ne("").any()
        )
        if has_options:
            base += ["option_type", "strike"]
        effective_cols = base
    else:
        effective_cols = list(group_cols or [])

    resolved = [display_group_col_map.get(c, c) for c in effective_cols]
    available = [c for c in resolved if c in df.columns]
    if not available:
        available = [c for c in ["Product", "Exchange", "Contract Month/Year"] if c in df.columns]

    # Aggregate: sum Realised PNL, Net Quantity, Long, Short.
    # NOV / settlementPrice / Avg Fill Price use _unique_or_multiple (same as other descriptive cols).
    numeric_sum_cols = [c for c in ["Realised PNL", "Net Quantity", "Long", "Short"] if c in df.columns]
    agg: dict = {c: "sum" for c in numeric_sum_cols}
    for col in df.columns:
        if col not in available and col not in agg:
            agg[col] = _unique_or_multiple

    try:
        grouped = df.groupby(available, dropna=False).agg(agg).reset_index()
    except Exception:
        return df

    # Null-out Long/Short where no Net P&L rows contributed (trade-only groups).
    for col in ["Long", "Short"]:
        if col in grouped.columns:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce")
            grouped.loc[grouped[col].fillna(0) == 0, col] = None

    return _apply_conditional_position_columns(grouped)


def grouped_fx_trades_view(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate FX trade rows by Account, Product, Type, CCY 1, CCY 2, and Expiry Date.

    Takes the output of ``standard_position_view_from_df`` (or the trades view
    dataframe, which uses "Quantity" instead of "Net Quantity") and returns one
    aggregated row per unique (Account Number, Product, Type, CCY 1, CCY 2,
    expiryDate) combination, with:

      - CCY 1 Amount / CCY 2 Amount — summed net notional
      - Quantity — summed net quantity
      - Unrealised PNL (OTE) — summed OTE (when available)
      - Trade Count — number of underlying individual trades

    Handles both "Quantity" (trades view) and "Net Quantity" (standard view)
    column names transparently.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    # Work on a copy; normalise to a single quantity column name.
    work = df.copy()
    qty_col = "Quantity" if "Quantity" in work.columns else "Net Quantity" if "Net Quantity" in work.columns else None

    # Filter to FX rows only.
    if "Type" in work.columns:
        type_upper = work["Type"].astype(str).str.strip().str.upper()
        fx_mask = type_upper.str.startswith("FX") | type_upper.isin(["NDO"])
    else:
        fx_mask = pd.Series(False, index=work.index)

    fx_df = work[fx_mask].copy()
    if fx_df.empty:
        return pd.DataFrame()

    # Group keys — only include columns that actually exist.
    group_cols = [c for c in ["Account Number", "Product", "Type", "CCY 1", "CCY 2", "expiryDate"] if c in fx_df.columns]

    # Build aggregation spec: sum numeric risk columns, count rows.
    fx_df["_trade_count"] = 1
    agg: Dict[str, Any] = {"_trade_count": "sum"}

    for col in ["CCY 1 Amount", "CCY 2 Amount", "Unrealised PNL (OTE)"]:
        if col in fx_df.columns:
            agg[col] = "sum"

    if qty_col and qty_col in fx_df.columns:
        agg[qty_col] = "sum"

    try:
        grouped = fx_df.groupby(group_cols, dropna=False).agg(agg).reset_index()
    except Exception:
        return fx_df.drop(columns=["_trade_count"], errors="ignore").reset_index(drop=True)

    grouped = grouped.rename(columns={"_trade_count": "Trade Count"})

    # Reorder columns for legibility.
    leading = [c for c in ["Account Number", "Product", "Type", "CCY 1", "CCY 2", "expiryDate"] if c in grouped.columns]
    numeric = [c for c in ["CCY 1 Amount", "CCY 2 Amount", qty_col or "", "Unrealised PNL (OTE)", "Trade Count"] if c and c in grouped.columns]
    rest = [c for c in grouped.columns if c not in leading and c not in numeric]
    return grouped[leading + numeric + rest].reset_index(drop=True)


def statement_dates_by_account(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Return statement-date coverage by account for audit after multi-PDF merge."""
    frames = []
    for name in ["Executed Trades", "Open Positions", "Purchase & Sale", "Closed Positions", "Receives Delivers", "Realized Gain and Loss", "Realized PNL Summary"]:
        df = tables.get(name, pd.DataFrame())
        if df is not None and not df.empty:
            cols = [c for c in ["source_pdf", "account_number", "statement_date"] if c in df.columns]
            if cols:
                temp = df[cols].drop_duplicates().copy()
                temp["sheet"] = name
                frames.append(temp)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True, sort=False).drop_duplicates()
    return out.sort_values([c for c in ["account_number", "statement_date", "source_pdf", "sheet"] if c in out.columns]).reset_index(drop=True)

def to_excel_bytes(tables: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    ordered = {
        "Summary": tables.get("Summary", pd.DataFrame()),
        "Aggregated Positions": prepare_grouped_positions_display(
            grouped_positions_product_month_auto_standard_view(tables),
            tables=tables,
            selected_preset="Product + Contract Month/Year",
            drop_option_columns=False,
        ),
        "Realized PNL": realized_pnl_summary(tables),
        "Statement Dates": statement_dates_by_account(tables),
    }
    for name, df in tables.items():
        if name == "Summary":
            continue
        if name == "Open Positions":
            ordered["Trades"] = open_positions_standard_view(tables)
        elif name == "Grouped Trades":
            continue
        else:
            ordered[name] = df

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in ordered.items():
            if df is None or df.empty:
                df = pd.DataFrame([{"message": "No rows extracted"}])
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            ws = writer.book[sheet_name[:31]]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
                ws.column_dimensions[col[0].column_letter].width = width
    output.seek(0)
    return output.getvalue()